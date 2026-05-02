"""
Kineo AG – Dashboard Update Script
===================================
Verwendung:
  python update_dashboard.py

Das Skript liest alle Quelldateien aus dem Ordner 'input/'
(rekursiv, inkl. input/hyrox/...), erkennt automatisch die neuesten
Versionen, und schreibt das Dashboard nach
'output/archive/Kineo_Dashboard_<Zeitstempel>.xlsx' sowie eine Kopie
als 'output/Kineo_Dashboard_AKTUELL.xlsx'.

Voraussetzungen:
  pip install openpyxl pandas pdfplumber
"""

import os
import re
import glob
import sys
import shutil
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import warnings
warnings.filterwarnings('ignore')
try:
    import pdfplumber
except ImportError:
    pdfplumber = None
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, LineChart, Reference, Series

# ═══════════════════════════════════════════════════════════════════════
# KONFIGURATION — hier kannst du Dateimuster anpassen
# ═══════════════════════════════════════════════════════════════════════

INPUT_DIR  = Path("input")
OUTPUT_DIR = Path("output")
# Zeitgestempelte Dashboard-Kopien landen hier; „Kineo_Dashboard_AKTUELL.xlsx“ bleibt direkt unter output/
OUTPUT_ARCHIVE_DIR = OUTPUT_DIR / "archive"
# Empfohlene Ablage für Hyrox-Uploads (optional — das Skript sucht rekursiv unter input/)
INPUT_HYROX_SCHEDULES = INPUT_DIR / "hyrox" / "schedules"
INPUT_HYROX_INVOICES = INPUT_DIR / "hyrox" / "invoices"
INPUT_LEGACY_DIR = INPUT_DIR / "_legacy"  # alte flache Ablage, weiterhin unterstützt
DOWNLOAD_DIR = Path.home() / "Downloads"

# Dateimuster: Schlüssel → Suchmuster (Glob) im input/-Ordner
# Das Skript nimmt bei mehreren Treffern immer die neueste Datei (nach Datum im Namen oder Änderungsdatum)
# Zusatzmonate (> Feb): Felder pro Therapeut/in für Ranking (aus by_month übernommen)
MONTH_THER_KEYS = {
    "Mär26": ("mar_umsatz", "mar_zielB"),
    "Apr26": ("apr_umsatz", "apr_zielB"),
    "Mai26": ("mai_umsatz", "mai_zielB"),
    "Jun26": ("jun_umsatz", "jun_zielB"),
    "Jul26": ("jul_umsatz", "jul_zielB"),
    "Aug26": ("aug_umsatz", "aug_zielB"),
    "Sep26": ("sep_umsatz", "sep_zielB"),
    "Okt26": ("okt_umsatz", "okt_zielB"),
    "Nov26": ("nov_umsatz", "nov_zielB"),
    "Dez26": ("dez_umsatz", "dez_zielB"),
}
MONTH_KPI_LABEL = {
    "Jan26": "Jan", "Feb26": "Feb", "Mär26": "Mär", "Apr26": "Apr", "Mai26": "Mai",
    "Jun26": "Jun", "Jul26": "Jul", "Aug26": "Aug", "Sep26": "Sep", "Okt26": "Okt",
    "Nov26": "Nov", "Dez26": "Dez",
}

FILE_PATTERNS = {
    "umsatzanalyse":    "*msatzanalyse*2026*.xlsx",
    "hyrox_schedule":   "*yrox*schedules*.xlsx",          # SportsNow: Kursplan Export (aktuell)
    "hyrox_dashboard":  "*yrox*dashboard*.xlsx",          # Altes aufbereitetes Dashboard (historisch)
    "hyrox_active":     "*yrox*active*passes*.xlsx",      # SportsNow: Aktive Abonnemente
    "hyrox_past":       "*yrox*past*passes*.xlsx",        # SportsNow: Abgelaufene Abonnemente
    "hyrox_invoices":   "*yrox*invoices*.xlsx",           # SportsNow: Rechnungen (Umsatz = System)
    "runnerslab":       "*nnerslab*.xlsx",
    "ausstehend":       "*usstehend*Zahlung*.xlsx",
    "fitness":          "*itness*Thalwil*.xlsx",
    "berichte":         "*erichte*Aerzte*.xlsx",
    "intrum":           "*ntrum*.pdf",
}

# Hinweis SportsNow-Exports:
# Kursplan:   Mein Adminbereich → Stundenplan → "Exportieren" (ohne TN-Details)
# Abonnemente: Mein Adminbereich → Abonnemente → Aktive / Abgelaufene → "Exportieren"
# Rechnungen: Umsatz gesamt / Monat (falls vorhanden) — sonst Fallback Pässe

# ═══════════════════════════════════════════════════════════════════════
# STIL-HELFER
# ═══════════════════════════════════════════════════════════════════════

DARK   = "1E293B"; BLUE  = "2563EB"; GREEN  = "16A34A"; RED   = "DC2626"
ORANGE = "D97706"; TEAL  = "0891B2"; PURPLE = "7C3AED"; WHITE = "FFFFFF"
ALT    = "F8FAFC"; LGRAY = "F1F5F9"; HDRBG  = "EFF6FF"; LBLUE = "DBEAFE"

def fill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, sz=10, color="1E293B", italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bb(c="D9D9D9"):
    s = Side(style="thin", color=c); return Border(bottom=s)
def fb(c="CBD5E1"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def title_bar(ws, row, text, bg=DARK, fg=WHITE, sz=14, cols=16):
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{gcl(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", bold=True, size=sz, color=fg)
    c.fill = fill(bg); c.alignment = aln("center", "center")

def sec(ws, row, text, cols=16, bg=LGRAY):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"A{row}:{gcl(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", bold=True, size=9, color="475569")
    c.fill = fill(bg)

def hdr(ws, row, vals, bg="1E3A5F", fg=WHITE, sz=9, h=18):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = Font(name="Arial", bold=True, size=sz, color=fg)
        c.fill = fill(bg); c.alignment = aln("center", "center"); c.border = fb("2C4A7C")

def dc(ws, row, col, val, bg=WHITE, bold=False, fmt=None, h="right", sz=10, color="1E293B"):
    c = ws.cell(row, col, val)
    c.font = Font(name="Arial", bold=bold, size=sz, color=color)
    c.fill = fill(bg); c.border = bb(); c.alignment = aln(h)
    if fmt: c.number_format = fmt
    return c

def zell(ws, row, col, val, bg=WHITE):
    """Zielerreichungs-Zelle mit Ampelfarbe."""
    if val is None or val == 0:
        ws.cell(row, col, "–").fill = fill(bg)
        ws.cell(row, col).border = bb()
        ws.cell(row, col).alignment = aln("center")
        return
    fc, tc = (("D1FAE5", GREEN)  if val >= 1.0  else
              ("DCFCE7", GREEN)  if val >= 0.9  else
              ("FEF3C7", ORANGE) if val >= 0.8  else
              ("FED7AA", ORANGE) if val >= 0.7  else
              ("FEE2E2", RED))
    c = ws.cell(row, col, val)
    c.fill = fill(fc); c.font = Font(name="Arial", bold=True, size=10, color=tc)
    c.number_format = "0.0%"; c.border = bb(); c.alignment = aln("center")

def gap(ws, row, cols=16):
    ws.row_dimensions[row].height = 5
    ws.merge_cells(f"A{row}:{gcl(cols)}{row}")
    ws.cell(row, 1).fill = fill("E2E8F0")

def kpi_tile(ws, row, col, label, val, sub, color=BLUE, span=1):
    """KPI-Kachel. span=N bedeutet: über N Spalten gemerged (schmale Spalten füllen)."""
    for r, h in [(row, 13), (row+1, 22), (row+2, 11)]:
        ws.row_dimensions[r].height = h
    end_col = col + span - 1
    for r in [row, row+1, row+2]:
        for cc in range(col, end_col + 1):
            ws.cell(r, cc).fill = fill(HDRBG)
            ws.cell(r, cc).border = fb("BFDBFE")
        if span > 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=end_col)
    ws.cell(row, col, label).font = fnt(sz=9, color="64748B")
    ws.cell(row+1, col, val).font = Font(name="Arial", bold=True, size=13, color=color)
    ws.cell(row+2, col, sub).font = fnt(sz=8, color="94A3B8", italic=True)

# ═══════════════════════════════════════════════════════════════════════
# DATEI-FINDER
# ═══════════════════════════════════════════════════════════════════════

_TEST_OR_TESTER_RE = re.compile(r"(?i)\btester\b|\btest\b")


def text_has_test_or_tester(val) -> bool:
    """True wenn im Text das Wort 'test' oder 'tester' vorkommt (Wortgrenzen, case-insensitive)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    return bool(_TEST_OR_TESTER_RE.search(str(val)))


def drop_rows_with_test_or_tester(df: pd.DataFrame, columns: list[str], label: str = "") -> pd.DataFrame:
    """Entfernt Zeilen, in denen eine der Spalten 'test' oder 'tester' als Wort enthält."""
    if df is None or df.empty:
        return df
    present = [c for c in columns if c in df.columns]
    if not present:
        return df
    mask_bad = pd.Series(False, index=df.index)
    for c in present:
        mask_bad |= df[c].map(text_has_test_or_tester)
    n_drop = int(mask_bad.sum())
    if n_drop:
        sfx = f" ({label})" if label else ""
        print(f"  → Zeilen mit Test/Tester ignoriert{sfx}: {n_drop}")
    return df.loc[~mask_bad].reset_index(drop=True)


def _is_temp_office_file(p: Path) -> bool:
    """Excel/Word temp files (~$...) — never use as data source."""
    return p.name.startswith("~$")


def _is_under_legacy_dir(p: Path) -> bool:
    """True wenn Datei unter input/_legacy/ liegt (Archiv, nicht für aktive Auswertung)."""
    try:
        return str(p.resolve()).startswith(str(INPUT_LEGACY_DIR.resolve()) + os.sep)
    except Exception:
        return False


def ensure_input_output_layout() -> None:
    """Legt empfohlene Unterordner an (leer ok)."""
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_LEGACY_DIR.mkdir(exist_ok=True)
    INPUT_HYROX_SCHEDULES.mkdir(parents=True, exist_ok=True)
    INPUT_HYROX_INVOICES.mkdir(parents=True, exist_ok=True)


def collect_input_matches(pattern: str) -> list[Path]:
    """Alle Treffer für ein Glob-Muster unter input/ (flach + rekursiv), ohne Office-Tempfiles."""
    matches: list[Path] = []
    if not INPUT_DIR.exists():
        return matches
    for p in INPUT_DIR.glob(pattern):
        if p.is_file() and not _is_temp_office_file(p) and not _is_under_legacy_dir(p):
            matches.append(p)
    for p in INPUT_DIR.rglob(pattern):
        if p.is_file() and not _is_temp_office_file(p) and not _is_under_legacy_dir(p):
            matches.append(p)
    uniq: dict[str, Path] = {}
    for p in matches:
        uniq[str(p.resolve())] = p
    return list(uniq.values())


def _regex_fallback_input_files(pattern: str) -> list[Path]:
    """Fallback wie bisher, aber rekursiv unter input/."""
    if not pattern:
        return []
    rx = re.compile(pattern.replace("*", ".*").replace("?", "."), re.IGNORECASE)
    out: list[Path] = []
    for p in INPUT_DIR.rglob("*"):
        if (
            p.is_file()
            and not _is_temp_office_file(p)
            and not _is_under_legacy_dir(p)
            and rx.search(p.name)
        ):
            out.append(p)
    return out


def abo_is_personal_training(abo) -> bool:
    """True wenn Abo-Text PT, Personal Training oder 1:1 / 1 zu 1 meint."""
    if abo is None or (isinstance(abo, float) and pd.isna(abo)):
        return False
    a = str(abo)
    al = a.lower().replace("–", "-").replace("—", "-")
    if "personal training" in al:
        return True
    if re.search(r"(?i)1\s*(?::|zu|-|)\s*1", a) or "1zu1" in al.replace(" ", ""):
        return True
    return "PT" in a


def find_file(key: str) -> Path | None:
    """Findet die neueste Datei für ein Muster. Gibt None zurück wenn nicht gefunden."""
    pattern = FILE_PATTERNS.get(key, "")
    matches = collect_input_matches(pattern)
    if key == "hyrox_invoices" and DOWNLOAD_DIR.exists():
        matches += [
            p for p in DOWNLOAD_DIR.glob(pattern)
            if p.is_file() and not _is_temp_office_file(p)
        ]
    if not matches:
        matches = _regex_fallback_input_files(pattern)
    if not matches:
        return None
    def sort_key(p):
        # Timestamp im SportsNow-Format: 20260325_0620
        m = re.search(r'(\d{8}_?\d{4})', p.name)
        if m: return m.group(0)
        m = re.search(r'(\d{4})[\._\-]?(\d{2})[\._\-]?(\d{2})', p.name)
        if m: return m.group(0)
        return str(p.stat().st_mtime)
    return sorted(matches, key=sort_key, reverse=True)[0]

def find_all_files(key: str) -> list[Path]:
    """Findet alle Dateien für ein Muster, neueste zuerst."""
    pattern = FILE_PATTERNS.get(key, "")
    matches = collect_input_matches(pattern)
    if key == "hyrox_invoices" and DOWNLOAD_DIR.exists():
        matches += [
            p for p in DOWNLOAD_DIR.glob(pattern)
            if p.is_file() and not _is_temp_office_file(p)
        ]
    if not matches:
        matches = _regex_fallback_input_files(pattern)
    def sort_key(p):
        m = re.search(r'(\d{4})[\._\-]?(\d{2})[\._\-]?(\d{2})', p.name)
        return m.group(0) if m else str(p.stat().st_mtime)
    return sorted(matches, key=sort_key, reverse=True)


def _extract_export_ts(name: str) -> int:
    """
    Extrahiert Export-Timestamp aus Dateinamen wie:
    kineo-hyrox_schedules_202604090829.xlsx -> 202604090829
    Fallback: 0 wenn nicht erkennbar.
    """
    m = re.search(r"(20\d{6})(\d{4})", str(name))
    if m:
        return int(m.group(1) + m.group(2))
    m2 = re.search(r"(20\d{6})", str(name))
    if m2:
        return int(m2.group(1)) * 10000
    return 0

# ═══════════════════════════════════════════════════════════════════════
# DATEN-LOADER
# ═══════════════════════════════════════════════════════════════════════

def load_umsatzanalyse(path: Path) -> dict:
    """Liest alle verfügbaren Monatsdaten aus der Umsatzanalyse."""
    print(f"  Lese Umsatzanalyse: {path.name}")
    ua = pd.read_excel(path, sheet_name=None, header=None)
    result = {
        "months_available": [],  # z.B. ["Jan26","Feb26"]
        "by_month": {},          # month -> list of therapeuten dicts
        "month_summary": {},     # month -> KPI summary (FTE/Umsatz/Umsatz pro FTE/Ziel Ø)
        "tracking": [],          # Umsatztracking
        "overview": {},          # Übersicht-Daten
        "input_params": {},      # Parameterblatt
    }

    # Input Parameter
    if "Input Parameter" in ua:
        ip = ua["Input Parameter"].dropna(how="all")
        params = {}
        for _, row in ip.iterrows():
            if pd.notna(row[0]) and pd.notna(row[1]):
                params[str(row[0]).strip()] = row[1]
        result["input_params"] = params

    # Umsatztracking
    if "Umsatztracking" in ua:
        tr = ua["Umsatztracking"]
        tracking = []
        for i in range(2, len(tr)):
            row = tr.iloc[i]
            if pd.notna(row[0]) and str(row[0]).strip() not in ["Summe","NaN","nan",""]:
                tracking.append({
                    "name": str(row[0]).strip(),
                    "stand1": float(row[1]) if pd.notna(row[1]) else None,
                    "stand2": float(row[2]) if pd.notna(row[2]) else None,
                    "delta": float(row[3]) if pd.notna(row[3]) else 0,
                })
        result["tracking"] = tracking

    # Übersicht (YTD-Daten)
    if "Übersicht" in ua:
        ov = ua["Übersicht"]
        overview = {"therapeuten": []}
        skip_names = {"xxx","Summe","Check","NaN","nan","","0","0.0"}
        for i in range(4, len(ov)):
            row = ov.iloc[i]
            name = str(row[0]).strip() if pd.notna(row[0]) else ""

            # Rechte Seite: Summary-Werte in col 29/30/31 (unabhängig von col0)
            label29 = str(row[29]).strip() if pd.notna(row[29]) else ""
            val30 = float(row[30]) if pd.notna(row[30]) else None
            val31 = float(row[31]) if pd.notna(row[31]) else None
            if label29 == "FTE":
                overview["fte_jan"] = val30; overview["fte_feb"] = val31
            elif label29 == "Ferien":
                overview["ferien_jan"] = val30; overview["ferien_feb"] = val31
            elif label29 == "Umsatz":
                overview["umsatz_jan"] = val30; overview["umsatz_feb"] = val31
            elif label29 == "Umsatz pro FTE":
                overview["umsatz_fte_jan"] = val30; overview["umsatz_fte_feb"] = val31
            elif label29 == "Zielerreichung Ø":
                overview["ziel_jan"] = val30; overview["ziel_feb"] = val31
            elif label29 == "Krank & andere":
                overview["krank_jan"] = val30; overview["krank_feb"] = val31

            # Therapeuten-Zeilen
            if not name or name in skip_names:
                continue
            bg = float(row[1]) if pd.notna(row[1]) else None
            jan_u  = float(row[2])  if pd.notna(row[2])  else None
            feb_u  = float(row[3])  if pd.notna(row[3])  else None
            jan_zB = float(row[15]) if pd.notna(row[15]) else None
            feb_zB = float(row[16]) if pd.notna(row[16]) else None
            overview["therapeuten"].append({
                "name": name, "bg": bg,
                "jan_umsatz": jan_u, "feb_umsatz": feb_u,
                "jan_zielB": jan_zB, "feb_zielB": feb_zB,
            })
        result["overview"] = overview

    # Monatliche Detail-Sheets
    month_map = {
        "Jan26": "Umsatz pro Mitarbeiter Jan26",
        "Feb26": "Umsatz pro Mitarbeiter Feb26",
        "Mär26": "Umsatz pro Mitarbeiter Mrz26",
        "Apr26": "Umsatz pro Mitarbeiter Apr26",
        "Mai26": "Umsatz pro Mitarbeiter Mai26",
        "Jun26": "Umsatz pro Mitarbeiter Jun26",
        "Jul26": "Umsatz pro Mitarbeiter Jul26",
        "Aug26": "Umsatz pro Mitarbeiter Aug26",
        "Sep26": "Umsatz pro Mitarbeiter Sep26",
        "Okt26": "Umsatz pro Mitarbeiter Okt26",
        "Nov26": "Umsatz pro Mitarbeiter Nov26",
        "Dez26": "Umsatz pro Mitarbeiter Dez26",
    }
    taetigkeiten_map = {
        "Jan26": "Input Tätigkeiten Jan26",
        "Feb26": "Input Tätigkeiten Feb26",
        "Mär26": "Input Tätigkeiten Mrz26",
        "Apr26": "Input Tätigkeiten Apr26",
        "Mai26": "Input Tätigkeiten Mai26",
    }

    for month_key, sheet_name in month_map.items():
        if sheet_name not in ua:
            continue
        df = ua[sheet_name]
        ther_list = []
        has_data = False
        # Monats-KPIs aus der Zusammenfassung im Sheet auslesen
        month_summary = {}
        for i2 in range(len(df)):
            lbl = str(df.iloc[i2, 0]).strip() if pd.notna(df.iloc[i2, 0]) else ""
            val = df.iloc[i2, 1] if pd.notna(df.iloc[i2, 1]) else None
            if lbl == "FTE" and val is not None:
                month_summary["fte"] = float(val)
            elif lbl == "Umsatz" and val is not None:
                month_summary["umsatz"] = float(val)
            elif lbl == "Umsatz pro FTE" and val is not None:
                month_summary["umsatz_fte"] = float(val)
            elif lbl in ["Zielerreichung Ø", "Zielerreichung O"] and val is not None:
                month_summary["ziel"] = float(val)
        if month_summary:
            result["month_summary"][month_key] = month_summary
        # Summen-Zeile 27: col4=Ferien, col5=Kurse, col6=Mktg, col7=Lauf, col8=Mgmt, col9=Krank
        try:
            sum_row = df.iloc[27]
            result["by_month_totals"] = result.get("by_month_totals", {})
            result["by_month_totals"][month_key] = {
                "ferien":  float(sum_row[4]) if pd.notna(sum_row[4]) else 0,
                "kurse":   float(sum_row[5]) if pd.notna(sum_row[5]) else 0,
                "mktg":    float(sum_row[6]) if pd.notna(sum_row[6]) else 0,
                "lauf":    float(sum_row[7]) if pd.notna(sum_row[7]) else 0,
                "mgmt":    float(sum_row[8]) if pd.notna(sum_row[8]) else 0,
                "krank":   float(sum_row[9]) if pd.notna(sum_row[9]) else 0,
            }
        except: pass
        # Summenzeilen am Ende der Monats-Sheets (inkl. "Krank & andere",
        # "Umsatz", "Umsatz pro FTE", "Zielerreichung Ø", "Summe (Tage ...)")
        SKIP_LABELS = {
            "summe", "nan", "", "fte", "ferien", "0", "0.0",
            "krank & andere", "umsatz", "umsatz pro fte",
            "zielerreichung ø", "zielerreichung o",
            "total", "ziel", "ziele", "zielerreichung",
        }
        for i in range(3, len(df)):
            row = df.iloc[i]
            name = str(row[0]).strip() if pd.notna(row[0]) else ""
            nm_key = name.lower()
            if (not name
                    or nm_key in SKIP_LABELS
                    or nm_key.startswith("summe")
                    or nm_key.startswith("zielerreichung")):
                continue
            bg = float(row[1]) if pd.notna(row[1]) else None
            umsatz = float(row[14]) if pd.notna(row[14]) else None
            if umsatz and umsatz > 0:
                has_data = True
            zA = float(row[16]) if pd.notna(row[16]) else None
            zB = float(row[18]) if pd.notna(row[18]) else None
            zC = float(row[20]) if pd.notna(row[20]) else None
            t_entry = {
                "name": name, "bg": bg,
                "umsatz": umsatz,
                "zielA": zA, "zielB": zB, "zielC": zC,
                "ferien": float(row[4]) if pd.notna(row[4]) else 0,
                "krank": float(row[9]) if pd.notna(row[9]) else 0,
                "mgmt": float(row[8]) if pd.notna(row[8]) else 0,
                "kurs_tage": float(row[5]) if pd.notna(row[5]) else 0,
                "at": float(row[2]) if pd.notna(row[2]) else 0,
                "sollat": float(row[3]) if pd.notna(row[3]) else 0,
                "tageB": float(row[11]) if pd.notna(row[11]) else 0,
                "kurs_anz": 0, "mktg_h": 0, "lauf_anz": 0,  # from Tätigkeiten
            }
            ther_list.append(t_entry)

        # Tätigkeiten-Daten hinzufügen
        taet_sheet = taetigkeiten_map.get(month_key)
        if taet_sheet and taet_sheet in ua:
            td = ua[taet_sheet]
            for i in range(3, len(td)):
                trow = td.iloc[i]
                tname = str(trow[0]).strip() if pd.notna(trow[0]) else ""
                if not tname or tname in ["0","Summe","NaN","nan",""]:
                    continue
                for t in ther_list:
                    if t["name"] == tname:
                        t["kurs_anz"] = int(trow[3]) if pd.notna(trow[3]) else 0
                        t["mktg_h"] = float(trow[4]) if pd.notna(trow[4]) else 0
                        t["lauf_anz"] = int(trow[5]) if pd.notna(trow[5]) else 0
                        break

        if has_data:
            result["months_available"].append(month_key)
            result["by_month"][month_key] = ther_list

    _merge_extra_months_into_overview_therapeuten(result)
    return result


def _merge_extra_months_into_overview_therapeuten(result: dict) -> None:
    """Hängt Umsatz/Ziel(B) für Monate > Feb an die Therapeuten-Liste in Übersicht an (Ranking).

    Wenn Therapeuten erst in einem späteren Monat auftauchen (z. B. Neuzugänge ab März),
    werden sie automatisch neu in die Liste aufgenommen, damit sie im Ranking erscheinen.
    """
    ov = result.setdefault("overview", {})
    thers = ov.setdefault("therapeuten", [])

    # Neuzugänge aus späten Monaten (Mär+) als Stub in die Liste aufnehmen.
    # Summen-Labels im Monats-Sheet filtern (z. B. "Krank & andere", "Umsatz",
    # "Zielerreichung Ø"), sonst landen die als Pseudo-Therapeuten im Ranking.
    SUMMARY_LABELS = {
        "krank & andere", "umsatz", "umsatz pro fte",
        "zielerreichung ø", "zielerreichung o",
        "fte", "ferien", "summe", "total", "ziel",
    }
    existing_names = {t["name"] for t in thers}
    for mk in result.get("months_available", []):
        if mk in ("Jan26", "Feb26"):
            continue
        for r in result.get("by_month", {}).get(mk, []):
            nm = r.get("name")
            if not nm or nm in existing_names:
                continue
            if nm.strip().lower() in SUMMARY_LABELS:
                continue
            # Ein echter Therapeut hat mindestens Umsatz oder Ziel(B) > 0
            if not (r.get("umsatz") or r.get("zielB")):
                continue
            thers.append({
                "name": nm,
                "bg": r.get("bg"),
                "jan_umsatz": None, "feb_umsatz": None,
                "jan_zielB": None,  "feb_zielB": None,
            })
            existing_names.add(nm)

    if not thers:
        return

    for mk, (uk, zk) in MONTH_THER_KEYS.items():
        if mk not in result.get("months_available", []):
            continue
        bm = result.get("by_month", {}).get(mk, [])
        by_name = {t["name"]: t for t in bm}
        for t in thers:
            r = by_name.get(t["name"])
            if r and r.get("umsatz") is not None:
                t[uk] = float(r["umsatz"])
            else:
                t[uk] = None
            zb = r.get("zielB") if r else None
            t[zk] = float(zb) if zb is not None else None


def load_runnerslab(path: Path) -> dict:
    """Liest Runnerslab Umsätze."""
    print(f"  Lese Runnerslab: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    years = {}
    months = ["Januar","Februar","März","April","Mai","Juni",
              "Juli","August","September","Oktober","November","Dezember"]
    months_short = ["Jan","Feb","Mär","Apr","Mai","Jun",
                    "Jul","Aug","Sep","Okt","Nov","Dez"]
    for i, row in df.iterrows():
        val = str(row[0]).strip()
        if re.match(r'^20\d\d$', val):
            yr = int(val)
            pos_row = df.iloc[i+1]
            onl_row = df.iloc[i+2]
            tot_row = df.iloc[i+3]
            years[yr] = {}
            for j, m in enumerate(months_short):
                years[yr][m] = {
                    "pos":    float(pos_row[j+1]) if pd.notna(pos_row[j+1]) else 0,
                    "online": float(onl_row[j+1]) if pd.notna(onl_row[j+1]) else 0,
                    "total":  float(tot_row[j+1]) if pd.notna(tot_row[j+1]) else 0,
                }
    return years


def load_hyrox(schedule_path: Path, active_path: Path, past_path: Path,
               dashboard_path: Path = None) -> dict:
    """
    Liest Hyrox-Daten aus SportsNow-Exporten + optionalem historischen Dashboard.
    Strategie: Historische Wochen aus altem Dashboard, aktuelle aus neuem Export.
      - dashboard_path: Altes hyrox_dashboard*.xlsx (historische Wochendaten)
      - schedule_path:  Neuer SportsNow Kursplan-Export (aktuelle Wochen)
      - active_path:    SportsNow Aktive Abonnemente
      - past_path:      SportsNow Abgelaufene Abonnemente
    """
    result = {
        "weekly": [], "best_classes": [], "worst_classes": [],
        "by_kurstyp": {}, "by_time": {}, "by_day": {},
        "total_revenue": 0, "pass_sales_monthly": {},
        "revenue_source": "",  # "invoices" | "passes" | ""
        "revenue_hyrox_courses": 0.0,       # ohne Personal Training
        "revenue_personal_training": 0.0,
        "revenue_monthly_hyrox": {},         # YYYY-MM -> CHF
        "revenue_monthly_pt": {},
        "conversion": {"trial": 0, "converted": 0, "rate": 0},
        "non_returning": [], "non_returning_by_type": [],
    }

    # ── Kursplan: historisch + aktuell kombinieren ────────────────────

    # Sammle alle Kursdaten in einem DataFrame
    all_sc_frames = []

    # 1. Historische Daten aus altem Dashboard (hyrox_dashboard*.xlsx)
    if dashboard_path and dashboard_path.exists():
        print(f"  Lese Hyrox historische Daten: {dashboard_path.name}")
        try:
            hist = pd.read_excel(dashboard_path, sheet_name="Source Weekly Classes", header=0)
            hist.columns = hist.columns.str.strip()
            hist["WeekStart"] = pd.to_datetime(hist["WeekStart"], errors="coerce")
            for _, row in hist.iterrows():
                if pd.notna(row["WeekStart"]):
                    kw = row["WeekStart"].isocalendar().week
                    result["weekly"].append({
                        "date": row["WeekStart"].strftime("%d.%m.%y"),
                        "kw": int(kw),
                        "kurse": int(row.get("Anzahl Kurse", 0) or 0),
                        "tn": int(row.get("Gesamtteilnehmer", 0) or 0),
                        "plaetze": int(row.get("Gesamtplätze", 0) or 0),
                        "auslastung": float(row.get("Auslastung %", 0) or 0),
                        "avg_tn": float(row.get("Durchschnitt Teilnehmer pro Kurs", 0) or 0),
                        "neue_kunden": 0,
                        "k1":0,"trial":0,"er10":0,"er30":0,"unlim":0,"pt":0,
                    })
            print(f"  → {len(result['weekly'])} historische Wochen geladen")
        except Exception as e:
            print(f"  ⚠️  Historische Daten nicht lesbar: {e}")

    # Bestimme bereits geladene Wochen anhand Datum (nicht KW, da KW sich wiederholen kann)
    loaded_week_dates = set()
    for w in result["weekly"]:
        loaded_week_dates.add(w.get("date",""))

    # 2. Alle SportsNow Kursplan-Exporte einlesen und zusammenfügen
    all_sc_parts = []
    schedule_months: dict[str, list[str]] = {}  # yyyy-mm -> [filenames]
    DEFAULT_HYROX_SEATS = 12  # in deinen bisherigen aggregierten Exports ist 'Plätze' konstant 12
    for sp in find_all_files("hyrox_schedule"):
        try:
            df_try = pd.read_excel(sp, header=0)
            df_try.columns = df_try.columns.str.strip()

            if "Datum" not in df_try.columns:
                print(f"  ⚠️ Kursplan-Datei übersprungen ({sp.name}): keine 'Datum'-Spalte")
                continue

            df_try["_source"] = sp.name
            df_try = drop_rows_with_test_or_tester(
                df_try,
                [
                    "Vorname", "Name", "Abonnement", "E-Mail", "Email",
                    "Stunde", "Beschreibung", "Ort/Raum", "Team",
                    "Zusatzinformationen (öffentlich)",
                    "Zusatzinformationen (intern)",
                ],
                sp.name,
            )
            if df_try.empty:
                print(f"  ⚠️ Kursplan nach Test/Tester-Filter leer — übersprungen ({sp.name})")
                continue

            # Monat(e) aus Datum-Spalte ableiten (auch bei Detail-Exports ohne Anmeldungen/Plätze)
            df_dates = pd.to_datetime(df_try["Datum"], dayfirst=True, errors="coerce").dropna()
            if not df_dates.empty:
                months_in_file = sorted(df_dates.dt.to_period("M").astype(str).unique())
                months_label = ", ".join(months_in_file)
                print(f"  Lese Kursplan: {sp.name} ({len(df_try)} Zeilen)  → Monate: {months_label}")
                for m in months_in_file:
                    schedule_months.setdefault(m, []).append(sp.name)
            else:
                print(f"  Lese Kursplan: {sp.name} ({len(df_try)} Zeilen)  → Monate: (Datum nicht auswertbar)")

            # Fall A: aggregierter Export (hat Anmeldungen/Plätze)
            if "Anmeldungen" in df_try.columns and "Plätze" in df_try.columns:
                if "Ort/Raum" not in df_try.columns:
                    df_try["Ort/Raum"] = ""
                all_sc_parts.append(df_try)
                continue

            # Fall B: Detail-Export (Teilnehmende je Zeile; braucht Aggregation)
            needed = ["Beginnt um", "Stunde"]
            if all(c in df_try.columns for c in needed):
                # gültige Teilnehmenden bestimmen (Name/Vorname vorhanden und nicht 'Ohne Abonnement')
                valid = pd.Series(False, index=df_try.index)
                if "Name" in df_try.columns:
                    valid |= df_try["Name"].notna()
                if "Vorname" in df_try.columns:
                    valid |= df_try["Vorname"].notna()
                # Wichtig: 'Ohne Abonnement' zählt im System typischerweise trotzdem als Buchung/Teilnahme
                # (belegt Platz) → deshalb NICHT herausfiltern.

                keys = ["Datum", "Beginnt um", "Stunde"]
                agg = (
                    df_try.loc[valid]
                    .groupby(keys)
                    .size()
                    .reset_index(name="Anmeldungen")
                )
                if agg.empty:
                    print(f"  ⚠️ Detail-Kursplan ohne auswertbare Teilnehmenden ({sp.name})")
                    continue

                # Plätze aus bisherigen Daten: pro Kursinstanz 12 Seats (konstant)
                agg["Plätze"] = DEFAULT_HYROX_SEATS
                agg["Ort/Raum"] = ""
                agg["_source"] = sp.name
                all_sc_parts.append(agg)
                continue

            print(f"  ⚠️ Kursplan-Datei übersprungen ({sp.name}): Schema nicht erkannt")

        except Exception as e:
            print(f"  ⚠️ Kursplan-Datei übersprungen ({sp.name}): {e}")

    # Wenn der neueste Monats-Export mitten in der Woche startet (z.B. Mittwoch),
    # ergänzen wir fehlende Wochenstart-Tage (Mo/Di) aus dem letzten älteren Export.
    # So werden KW-Werte über Monatsgrenzen korrekt vollständig.
    if schedule_path and schedule_path.exists() and all_sc_parts:
        try:
            latest_ts = _extract_export_ts(schedule_path.name)
            latest_df = pd.read_excel(schedule_path, header=0)
            latest_df.columns = latest_df.columns.str.strip()
            if "Datum" in latest_df.columns:
                latest_df["Datum"] = pd.to_datetime(latest_df["Datum"], dayfirst=True, errors="coerce")
                latest_df = latest_df[latest_df["Datum"].notna()].copy()
                if not latest_df.empty:
                    min_date = latest_df["Datum"].min()
                    weekday = int(min_date.weekday())  # 0=Mo
                    if weekday > 0:
                        week_start = min_date - pd.to_timedelta(weekday, unit="d")
                        missing_dates = {(week_start + pd.to_timedelta(i, unit="d")).normalize() for i in range(weekday)}
                        if missing_dates:
                            # bereits vorhandene Kurs-Keys aus allen aktuell geladenen Schedules
                            present = set()
                            for part in all_sc_parts:
                                p2 = part.copy()
                                if "Datum" not in p2.columns:
                                    continue
                                p2["Datum"] = pd.to_datetime(p2["Datum"], dayfirst=True, errors="coerce")
                                p2 = p2[p2["Datum"].notna()].copy()
                                for _, rr in p2.iterrows():
                                    present.add((
                                        rr["Datum"].normalize(),
                                        str(rr.get("Beginnt um", "")),
                                        str(rr.get("Stunde", "")),
                                    ))

                            # ältere Schedule-Exporte aus Downloads durchsuchen (neueste zuerst)
                            dl_candidates = []
                            if DOWNLOAD_DIR.exists():
                                for dp in DOWNLOAD_DIR.glob(FILE_PATTERNS["hyrox_schedule"]):
                                    if not dp.is_file() or _is_temp_office_file(dp):
                                        continue
                                    ts = _extract_export_ts(dp.name)
                                    if ts and latest_ts and ts < latest_ts:
                                        dl_candidates.append(dp)
                            dl_candidates = sorted(dl_candidates, key=lambda p: _extract_export_ts(p.name), reverse=True)

                            carry_rows = []
                            for dp in dl_candidates:
                                try:
                                    ddf = pd.read_excel(dp, header=0)
                                    ddf.columns = ddf.columns.str.strip()
                                    if "Datum" not in ddf.columns:
                                        continue
                                    ddf["Datum"] = pd.to_datetime(ddf["Datum"], dayfirst=True, errors="coerce")
                                    ddf = ddf[ddf["Datum"].notna()].copy()
                                    if ddf.empty:
                                        continue
                                    ddf["_source"] = dp.name
                                    ddf = ddf[ddf["Datum"].dt.normalize().isin(missing_dates)].copy()
                                    if ddf.empty:
                                        continue
                                    for _, rr in ddf.iterrows():
                                        key = (
                                            rr["Datum"].normalize(),
                                            str(rr.get("Beginnt um", "")),
                                            str(rr.get("Stunde", "")),
                                        )
                                        if key not in present:
                                            carry_rows.append(rr)
                                            present.add(key)
                                    # sobald wir für mindestens einen fehlenden Tag Zeilen gefunden haben, reicht 1 Datei
                                    if carry_rows:
                                        break
                                except Exception:
                                    continue

                            if carry_rows:
                                carry_df = pd.DataFrame(carry_rows)
                                if "Ort/Raum" not in carry_df.columns:
                                    carry_df["Ort/Raum"] = ""
                                all_sc_parts.append(carry_df)
                                print(
                                    f"  → Monatsgrenze ergänzt: {len(carry_df)} Kurszeilen "
                                    f"aus älterem Schedule-Export für KW-Startwoche"
                                )
        except Exception as e:
            print(f"  ⚠️ Monatsgrenzen-Ergänzung übersprungen: {e}")

    # Warnung bei mehrfach abgedeckten Monaten
    dup_months = {m: files for m, files in schedule_months.items() if len(files) > 1}
    if dup_months:
        print("  ⚠️ Hinweis: Mehrere Kursplan-Dateien für denselben Monat gefunden:")
        for m, files in sorted(dup_months.items()):
            print(f"     - {m}: {', '.join(files)}")

    sc_to_use = pd.concat(all_sc_parts, ignore_index=True) if all_sc_parts else None

    if sc_to_use is not None and not sc_to_use.empty:
        sc = sc_to_use.copy()
        sc["Datum"] = pd.to_datetime(sc["Datum"], dayfirst=True, errors="coerce")
        sc = sc[sc["Datum"].notna()].copy()
        # Abgesagte/annullierte Klassen nicht mitzählen (wurden bisher als Kurse gezählt).
        if "Status" in sc.columns:
            status_norm = sc["Status"].astype(str).str.strip().str.lower()
            sc = sc[~status_norm.isin({"abgesagt", "annulliert", "cancelled", "canceled"})].copy()
        # Duplikate entfernen (gleicher Kurs in mehreren Exporten) — neueste Quelle gewinnt
        sc = sc.sort_values("_source").drop_duplicates(
            subset=["Datum","Beginnt um","Stunde"], keep="last"
        ).reset_index(drop=True)
        sc["Auslastung"] = sc["Anmeldungen"] / sc["Plätze"].replace(0, 1)
        sc["Wochentag_DE"] = sc["Datum"].dt.day_name().map({
            "Monday":"Montag","Tuesday":"Dienstag","Wednesday":"Mittwoch",
            "Thursday":"Donnerstag","Friday":"Freitag","Saturday":"Samstag","Sunday":"Sonntag"
        })
        sc["WeekStart"] = sc["Datum"] - pd.to_timedelta(sc["Datum"].dt.weekday, unit="d")
        # Nur Wochen hinzufügen die noch nicht aus historischen Daten geladen wurden
        for ws_date, grp in sc.groupby("WeekStart"):
            kw = int(ws_date.isocalendar().week)
            ws_label = ws_date.strftime("%d.%m.%y")
            if ws_label not in loaded_week_dates:
                result["weekly"].append({
                    "date": ws_date.strftime("%d.%m.%y"),
                    "kw": kw,
                    "kurse": len(grp),
                    "tn": int(grp["Anmeldungen"].sum()),
                    "plaetze": int(grp["Plätze"].sum()),
                    "auslastung": float(grp["Anmeldungen"].sum() / grp["Plätze"].sum()) if grp["Plätze"].sum() > 0 else 0,
                    "avg_tn": float(grp["Anmeldungen"].mean()),
                    "neue_kunden": 0,
                    "k1":0,"trial":0,"er10":0,"er30":0,"unlim":0,"pt":0,
                })
        # Best / Worst aus allen verfügbaren Kursen
        ran = sc[sc["Anmeldungen"] > 0].copy()
        if len(ran) >= 3:
            result["best_classes"] = ran.nlargest(3, "Auslastung")[
                ["Datum","Beginnt um","Stunde","Ort/Raum","Anmeldungen","Plätze","Auslastung","Wochentag_DE"]
            ].rename(columns={"Wochentag_DE":"Wochentag"}).to_dict("records")
            result["worst_classes"] = ran.nsmallest(3, "Auslastung")[
                ["Datum","Beginnt um","Stunde","Ort/Raum","Anmeldungen","Plätze","Auslastung","Wochentag_DE"]
            ].rename(columns={"Wochentag_DE":"Wochentag"}).to_dict("records")
        # Auslastung-Analysen
        for key, col, attr in [("by_kurstyp","Stunde","by_kurstyp"),
                                 ("by_time","Beginnt um","by_time"),
                                 ("by_day","Wochentag_DE","by_day")]:
            result[attr] = ran.groupby(col).agg(
                n=("Anmeldungen","count"),
                avg_tn=("Anmeldungen","mean"),
                avg_ausl=("Auslastung","mean")
            ).sort_values("avg_ausl", ascending=False).to_dict("index")

    # Neue Schedule-Wochen überschreiben historische (neuere Daten sind genauer)
    # Merge: historische als Basis, neue Exporte überschreiben bei Überlappung
    if all_sc_parts:
        sc_weekly_map = {}  # date -> entry
        sc["WeekStart"] = sc["Datum"] - pd.to_timedelta(sc["Datum"].dt.weekday, unit="D")
        for ws_date, grp in sc.groupby("WeekStart"):
            ws_label = ws_date.strftime("%d.%m.%y")
            sc_weekly_map[ws_label] = {
                "date": ws_label,
                "kw": int(ws_date.isocalendar().week),
                "kurse": len(grp),
                "tn": int(grp["Anmeldungen"].sum()),
                "plaetze": int(grp["Plätze"].sum()),
                "auslastung": float(grp["Anmeldungen"].sum() / grp["Plätze"].sum()) if grp["Plätze"].sum() > 0 else 0,
                "avg_tn": float(grp["Anmeldungen"].mean()),
                "neue_kunden": 0,
                "k1":0,"trial":0,"er10":0,"er30":0,"unlim":0,"pt":0,
            }
        # Historische Wochen die NICHT in neuen Exporten sind behalten
        # Wochen die in neuen Exporten sind → neue Version verwenden
        hist_only = [w for w in result["weekly"] if w["date"] not in sc_weekly_map]
        new_weeks = list(sc_weekly_map.values())
        result["weekly"] = hist_only + new_weeks

    # Sortieren nach Datum
    result["weekly"].sort(key=lambda x: x.get("date", ""))

    # ── Hyrox Umsatz: Rechnungen (bevorzugt, entspricht meist dem System) ─
    all_inv_parts = []
    for inv_path in find_all_files("hyrox_invoices"):
        if not inv_path or not inv_path.exists():
            continue
        try:
            print(f"  Lese Hyrox Rechnungen: {inv_path.name}")
            idf = pd.read_excel(inv_path, header=0, engine="openpyxl")
            idf.columns = idf.columns.str.strip()
            idf["_source"] = inv_path.name
            all_inv_parts.append(idf)
        except Exception as e:
            print(f"  ⚠️ Hyrox Rechnungen übersprungen ({inv_path.name}): {e}")

    if all_inv_parts:
        inv_df = pd.concat(all_inv_parts, ignore_index=True)
        inv_df = drop_rows_with_test_or_tester(
            inv_df,
            [
                "Vorname", "Name", "Abonnement", "Beschreibung",
                "E-Mail", "Email", "Stunde", "Kunde",
            ],
            "Hyrox-Rechnungen",
        )
        if "Total" not in inv_df.columns:
            print("  ⚠️ Rechnungs-Export ohne Spalte 'Total' — Umsatz aus Rechnungen nicht möglich")
        else:
            inv_df["Total"] = pd.to_numeric(inv_df["Total"], errors="coerce").fillna(0)
            dedup_col = "Rechnungsnummer" if "Rechnungsnummer" in inv_df.columns else None
            if dedup_col:
                # Prefer the newest export when the same Rechnungnummer appears in
                # multiple invoice files (input/ + Downloads/).
                def _inv_source_sortkey(src_name: str) -> int:
                    # Typical: kineo-hyrox_invoices_202604021058.xlsx
                    # Extract yyyymmdd + hhmm if present.
                    m = re.search(r"(\d{8})[_\-]?\d{0,4}(\d{4})", str(src_name))
                    if m:
                        # group(1)=yyyymmdd, group(2)=hhmm
                        return int(m.group(1) + m.group(2))
                    m2 = re.search(r"(\d{8})", str(src_name))
                    if m2:
                        return int(m2.group(1)) * 10000
                    return 0

                inv_df["_inv_sort"] = inv_df["_source"].apply(_inv_source_sortkey)
                inv_df = (
                    inv_df.sort_values("_inv_sort")
                    .drop_duplicates(subset=[dedup_col], keep="last")
                    .reset_index(drop=True)
                )
                if "_inv_sort" in inv_df.columns:
                    del inv_df["_inv_sort"]

            # "Annulliert" unbedingt bei Umsatz rausrechnen (bezahlt/offen soll stimmen)
            if "Rechnungsstatus" in inv_df.columns:
                rs_norm = inv_df["Rechnungsstatus"].astype(str).str.strip().str.lower()
                mask_cancel = rs_norm.eq("annulliert")
            else:
                mask_cancel = pd.Series(False, index=inv_df.index)

            inv_rev = inv_df.loc[~mask_cancel].copy()
            if int(mask_cancel.sum()):
                print(f"  → Rechnungen 'Annulliert' ignoriert: {int(mask_cancel.sum())} Zeilen")

            result["total_revenue"] = float(inv_rev["Total"].sum())
            result["revenue_source"] = "invoices"
            if "Abonnement" in inv_df.columns:
                mask_pt = inv_rev["Abonnement"].apply(abo_is_personal_training)
            else:
                mask_pt = pd.Series(False, index=inv_rev.index)
            result["revenue_hyrox_courses"] = float(inv_rev.loc[~mask_pt, "Total"].sum())
            result["revenue_personal_training"] = float(inv_rev.loc[mask_pt, "Total"].sum())
            if "Kaufdatum" in inv_df.columns:
                kd = pd.to_datetime(inv_rev["Kaufdatum"], dayfirst=True, errors="coerce")
                ok = inv_rev.loc[kd.notna()].copy()
                ok["Monat"] = kd[kd.notna()].dt.to_period("M").astype(str)
                ok["Total"] = pd.to_numeric(ok["Total"], errors="coerce").fillna(0)
                if "Abonnement" in ok.columns:
                    ok["_is_pt"] = ok["Abonnement"].apply(abo_is_personal_training)
                else:
                    ok["_is_pt"] = False
                result["pass_sales_monthly"] = ok.groupby("Monat")["Total"].sum().to_dict()
                result["revenue_monthly_hyrox"] = (
                    ok[~ok["_is_pt"]].groupby("Monat")["Total"].sum().to_dict()
                )
                result["revenue_monthly_pt"] = (
                    ok[ok["_is_pt"]].groupby("Monat")["Total"].sum().to_dict()
                )
            print(
                f"  → Hyrox Umsatz (Rechnungen): CHF {result['total_revenue']:,.0f} "
                f"(Hyrox o. PT: CHF {result['revenue_hyrox_courses']:,.0f}, "
                f"PT: CHF {result['revenue_personal_training']:,.0f}) "
                f"— {len(inv_rev)} Zeilen nach Dedupe (Annulliert ausgeschlossen)"
            )

    # ── Pass Sales (Aktiv + Abgelaufen): Conversion, Produktmix, Nicht-Wiederkehrer ─
    # WICHTIG: active/past Exporte sind oft Momentaufnahmen.
    # Damit der Gesamtumsatz seit Start stimmt, lesen wir ALLE passenden Exporte aus input/ ein
    # und deduplizieren nach ID.
    all_passes = []
    pass_files = []
    for key, label in [("hyrox_active", "aktiv"), ("hyrox_past", "abgelaufen")]:
        for path in find_all_files(key):
            if path and path.exists():
                pass_files.append((path, label))

    # Falls keine Dateien gefunden wurden, fallback auf die übergebenen Pfade
    if not pass_files:
        for path, label in [(active_path, "aktiv"), (past_path, "abgelaufen")]:
            if path and path.exists():
                pass_files.append((path, label))

    for path, label in pass_files:
        try:
            print(f"  Lese Hyrox Passes ({label}): {path.name}")
            # engine explizit setzen (verhindert sporadische read timeouts)
            df = pd.read_excel(path, header=0, engine="openpyxl")
            df.columns = df.columns.str.strip()
            df["_source"] = path.name
            all_passes.append(df)
        except Exception as e:
            print(f"  ⚠️ Hyrox Passes übersprungen ({path.name}): {e}")

    if all_passes:
        passes = pd.concat(all_passes, ignore_index=True)
        passes = drop_rows_with_test_or_tester(
            passes,
            [
                "Vorname", "Name", "Abonnement", "Beschreibung",
                "E-Mail", "Email", "Kunde",
            ],
            "Hyrox-Pässe",
        )
        # Deduplizieren nach ID (gleicher Kauf in mehreren Exports)
        if "ID" in passes.columns:
            passes = passes.sort_values("_source").drop_duplicates(subset=["ID"], keep="last").reset_index(drop=True)

        # Umsatz nur aus Pässen, wenn keine Rechnungen vorhanden
        if result.get("revenue_source") != "invoices":
            if "Total" in passes.columns:
                passes["Total"] = pd.to_numeric(passes["Total"], errors="coerce").fillna(0)
                result["total_revenue"] = float(passes["Total"].sum())
                result["revenue_source"] = "passes"
        else:
            if "Total" in passes.columns:
                passes["Total"] = pd.to_numeric(passes["Total"], errors="coerce").fillna(0)

        # Produktgruppen vereinfachen
        def produktgruppe(abo):
            abo = str(abo)
            if "Trial" in abo:      return "Trial"
            if "10 Class" in abo:   return "10er"
            if "30 Class" in abo:   return "30er"
            if "Unlimited" in abo:  return "Unlimited"
            if "Single" in abo:     return "1 Karte"
            if abo_is_personal_training(abo): return "PT"
            return "Andere"
        passes["Produktgruppe"] = passes["Abonnement"].apply(produktgruppe)
        result["pass_sales_by_product"] = passes.groupby("Produktgruppe")["Total"].sum().to_dict()

        # Umsatz Hyrox vs PT aus Pässen (nur wenn keine Rechnungen)
        if result.get("revenue_source") == "passes" and "Total" in passes.columns:
            result["revenue_hyrox_courses"] = float(
                passes.loc[passes["Produktgruppe"] != "PT", "Total"].sum()
            )
            result["revenue_personal_training"] = float(
                passes.loc[passes["Produktgruppe"] == "PT", "Total"].sum()
            )
            if "Kaufdatum" in passes.columns:
                passes["Kaufdatum"] = pd.to_datetime(
                    passes["Kaufdatum"], dayfirst=True, errors="coerce"
                )
                p2 = passes[passes["Kaufdatum"].notna()].copy()
                if not p2.empty:
                    p2["Monat"] = p2["Kaufdatum"].dt.to_period("M").astype(str)
                    result["pass_sales_monthly"] = p2.groupby("Monat")["Total"].sum().to_dict()
                    result["revenue_monthly_hyrox"] = (
                        p2[p2["Produktgruppe"] != "PT"].groupby("Monat")["Total"].sum().to_dict()
                    )
                    result["revenue_monthly_pt"] = (
                        p2[p2["Produktgruppe"] == "PT"].groupby("Monat")["Total"].sum().to_dict()
                    )

        # Trial Conversion (nur wenn 'Kunde' existiert)
        converted = 0
        trial_kunden = 0
        if "Kunde" in passes.columns:
            trial_kunden = passes[passes["Produktgruppe"]=="Trial"]["Kunde"].nunique()
            trial_ids = set(passes[passes["Produktgruppe"]=="Trial"]["Kunde"])
            converted_ids = set(passes[passes["Produktgruppe"].isin(["10er","30er","Unlimited"])]["Kunde"])
            converted = len(trial_ids & converted_ids)
        result["conversion"] = {
            "trial": trial_kunden,
            "converted": converted,
            "rate": converted/trial_kunden if trial_kunden > 0 else 0,
        }

        # Non-returning customers: abgelaufen aber kein aktives Abo
        # (hier verwenden wir weiterhin die jeweils neuesten Dateien, falls vorhanden)
        if active_path and active_path.exists() and past_path and past_path.exists():
            active_df = pd.read_excel(active_path, header=0, engine="openpyxl")
            past_df   = pd.read_excel(past_path,   header=0, engine="openpyxl")
            active_df.columns = active_df.columns.str.strip()
            past_df.columns   = past_df.columns.str.strip()
            _pass_test_cols = [
                "Vorname", "Name", "Abonnement", "Beschreibung",
                "E-Mail", "Email", "Kunde",
            ]
            active_df = drop_rows_with_test_or_tester(
                active_df, _pass_test_cols, "aktiv (Nicht-Wiederkehrer)"
            )
            past_df = drop_rows_with_test_or_tester(
                past_df, _pass_test_cols, "abgelaufen (Nicht-Wiederkehrer)"
            )
            active_ids_set = set(active_df["Kunde"].unique())
            past_ids_set   = set(past_df["Kunde"].unique())
            nr_ids = past_ids_set - active_ids_set
            nr = past_df[past_df["Kunde"].isin(nr_ids)].copy()
            nr["Gültig bis"] = pd.to_datetime(nr["Gültig bis"], dayfirst=True, errors="coerce")
            nr = nr.sort_values("Gültig bis", ascending=False).drop_duplicates("Kunde", keep="first")
            nr = nr.sort_values("Gültig bis", ascending=False)
            result["non_returning"] = nr[
                [c for c in ["Vorname","Name","E-Mail","Abonnement","Gültig bis","Total"] if c in nr.columns]
            ].to_dict("records")
            result["non_returning_by_type"] = nr.groupby("Abonnement").agg(
                Anzahl=("Kunde","count"), Umsatz=("Total","sum")
            ).sort_values("Anzahl", ascending=False).reset_index().to_dict("records")
            print(f"  → Nicht-wiederkehrende Kunden: {len(nr)}")

        src = result.get("revenue_source", "")
        if src == "invoices":
            print(f"  → Hyrox KPI-Umsatz (Rechnungen, siehe oben): CHF {result['total_revenue']:,.0f}")
        else:
            print(f"  → Hyrox Total Revenue (Pässe): CHF {result['total_revenue']:,.0f}")
        print(f"  → Trial Conversion: {converted}/{trial_kunden} = {result['conversion']['rate']:.1%}")

    return result

def load_ausstehend(path: Path) -> list:
    """Liest ausstehende Zahlungen."""
    print(f"  Lese Ausstehende Zahlungen: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    invoices = []
    header_row = None
    for i, row in df.iterrows():
        if str(row[0]).strip() in ["Re.- Nr.", "Re.-Nr."]:
            header_row = i
            break
    if header_row is None:
        return []
    for i in range(header_row+1, len(df)):
        row = df.iloc[i]
        if pd.notna(row[0]) and str(row[0]).strip() not in ["NaN","nan",""]:
            try:
                invoices.append({
                    "renr":      str(row[0]).strip(),
                    "kunde":     str(row[1]).strip() if pd.notna(row[1]) else "",
                    "therapeut": str(row[2]).strip() if pd.notna(row[2]) else "",
                    "an":        str(row[3]).strip() if pd.notna(row[3]) else "",
                    "an_wen":    str(row[4]).strip() if pd.notna(row[4]) else "",
                    "re_datum":  str(row[6]).strip() if pd.notna(row[6]) else "",
                    "letzte_m":  str(row[7]).strip() if pd.notna(row[7]) else "",
                    "mahnstufe": int(row[8]) if pd.notna(row[8]) else 0,
                    "tage":      int(row[9]) if pd.notna(row[9]) else 0,
                    "standort":  int(row[10]) if pd.notna(row[10]) else 0,
                    "betrag":    float(row[12]) if pd.notna(row[12]) else 0,
                })
            except: pass
    return invoices


def load_fitness(path: Path) -> list:
    """Liest Fitness-Abo-Daten."""
    print(f"  Lese Fitness-Abos: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    rows = []
    # Header ist Zeile mit "Was" oder KW-Muster
    header_idx = None
    for i, row in df.iterrows():
        if str(row[0]).strip() == "Was":
            header_idx = i; break
    if header_idx is None:
        return []
    labels = df.iloc[header_idx+1:header_idx+10, 0].tolist()
    kw_cols = df.iloc[header_idx].tolist()
    for j, kw in enumerate(kw_cols[1:], 1):
        if pd.isna(kw): continue
        kw_label = str(kw).strip()
        entry = {"kw": kw_label}
        for k, label in enumerate(labels):
            label_str = str(label).strip() if pd.notna(label) else ""
            val = df.iloc[header_idx+1+k, j]
            entry[label_str] = val if pd.notna(val) else None
        rows.append(entry)
    return rows


# Intrum Auftragsbestätigung (PDF) → Inkasso-Sheet
_INTRUM_FALLSTATUS_RE = re.compile(
    r"^(Mahnphase|Eröffnungsphase|Einstellung|Inkassophase|Rechtsverfolgung|Zahlungseingang)",
    re.IGNORECASE,
)
_INTRUM_ADR_CODES = re.compile(
    r"\s(10|20|26|27|28|29|30|31|32|33)\s",
)
_INTRUM_SCORE_ERFOLG = {
    "A": "75–100%",
    "B": "60–75%",
    "C": "25–60%",
    "D": "10–25%",
    "E": "0–10%",
    "M": "–",
    "X": "–",
    "Z": "–",
    "–": "–",
}


def _intrum_parse_chf_amounts(text: str) -> list[float]:
    out: list[float] = []
    for m in re.finditer(r"(?:\d{1,3}(?:'\d{3})+|\d{1,3}(?:,\d{3})+|\d+)\.\d{2}", text):
        s = m.group(0).replace("'", "")
        if "," in s:
            if s.rfind(",") < s.rfind("."):
                s = s.replace(",", "")
            else:
                s = s.replace(",", "X").replace(".", "").replace("X", ".")
        try:
            out.append(float(s))
        except ValueError:
            pass
    return out


def _intrum_split_blocks(text: str) -> list[str]:
    text = text.replace("\r\n", "\n")
    starts = [m.start() for m in re.finditer(r"(?m)^(\d{8})(?=\s)", text)]
    blocks: list[str] = []
    for i, s in enumerate(starts):
        blk = text[s : starts[i + 1] if i + 1 < len(starts) else len(text)].strip()
        if len(blk) < 15:
            continue
        fl = blk.split("\n", 1)[0].strip()
        if fl.startswith("Inkassonummer") or not re.match(r"^\d{8}\s", fl):
            continue
        blocks.append(blk)
    return blocks


def _intrum_parse_block(block: str) -> dict | None:
    lines = [l.strip() for l in block.splitlines() if l.strip()]
    if not lines:
        return None
    first = lines[0]
    if first.lower().startswith("anzahl fälle"):
        return None

    m = re.match(
        r"^(\d{8})\s+(.+?)\s+([A-E]|M|X|Z)\s+(Deutsch|Englisch)\s+(.*)$",
        first,
    )
    m2 = re.match(r"^(\d{8})\s+(.+?)\s+-\s+(Englisch|Deutsch)\s+(.*)$", first)
    if m:
        ink, name, score, _lang, _tail = m.groups()
    elif m2:
        ink, name, _lang, _tail = m2.groups()
        score = "–"
    else:
        return None

    name = re.sub(r"\s+", " ", name).strip()
    addr_parts: list[str] = []
    adr_stat = None
    ref_nr = ""
    re_datum = None
    fallstatus = ""

    for line in lines[1:]:
        if _INTRUM_FALLSTATUS_RE.match(line):
            fallstatus = _INTRUM_FALLSTATUS_RE.match(line).group(1)
            break
        addr_parts.append(line)
        if adr_stat is None:
            am = _INTRUM_ADR_CODES.search(f" {line} ")
            if am:
                adr_stat = am.group(1)
        if re_datum is None:
            dm = re.search(r"(\d{2}\.\d{2}\.\d{4})", line)
            if dm:
                re_datum = dm.group(1)
        rm = re.search(r"\b(\d{5})\s+\d{2}\.\d{2}\.\d{4}", line)
        if rm:
            ref_nr = rm.group(1)

    adresse = re.sub(r"\s+", " ", " ".join(addr_parts)).strip() or "–"

    kapital = None
    for line in reversed(lines):
        if _INTRUM_FALLSTATUS_RE.match(line):
            ams = _intrum_parse_chf_amounts(line)
            if ams:
                kapital = ams[-1]
            break
    if kapital is None:
        ams = _intrum_parse_chf_amounts(block)
        kapital = ams[-1] if ams else None

    return {
        "inkasso_nr": ink,
        "schuldner": name,
        "adresse": adresse[:240],
        "score": score,
        "adress_status": adr_stat or "–",
        "ref_nr": ref_nr or "–",
        "datum": re_datum or "–",
        "kapital": kapital,
        "fallstatus": fallstatus or "–",
        "erfolgsquote": _INTRUM_SCORE_ERFOLG.get(score, "–"),
    }


def _intrum_extract_pdf_text(path: Path) -> str:
    if pdfplumber is None:
        return ""
    chunks: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            head = t[:1400]
            if "Legende Scorewerte" in head and "Inkassonummer" not in head[:500]:
                continue
            chunks.append(t)
    return "\n".join(chunks)


def parse_intrum_pdf(path: Path) -> list[dict]:
    """Liest eine Intrum-Auftragsbestätigung-PDF; liefert Fälle mit Schlüsseln für das Inkasso-Sheet."""
    full = _intrum_extract_pdf_text(path)
    if not full.strip():
        return []
    cases: list[dict] = []
    for blk in _intrum_split_blocks(full):
        row = _intrum_parse_block(blk)
        if row:
            row["_source_file"] = path.name
            cases.append(row)
    return cases


def load_intrum_cases(paths: list[Path]) -> list[dict]:
    """
    Alle Intrum-PDFs einlesen, nach Inkasso-Nr. deduplizieren (neuere Datei überschreibt ältere).
    """
    if not paths or pdfplumber is None:
        return []
    ordered = sorted(paths, key=lambda p: p.stat().st_mtime)
    by_nr: dict[str, dict] = {}
    for p in ordered:
        for c in parse_intrum_pdf(p):
            by_nr[c["inkasso_nr"]] = c
    return sorted(by_nr.values(), key=lambda x: int(x["inkasso_nr"]))


def load_berichte(path: Path) -> dict:
    """Liest Berichte an Ärzte."""
    print(f"  Lese Berichte an Ärzte: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    if len(df) < 2:
        return {}
    header_row = df.iloc[0].tolist()
    therapists = [str(h) for h in header_row[1:] if pd.notna(h)]
    data = {}
    for i in range(1, len(df)):
        row = df.iloc[i]
        kw = str(row[0]).strip() if pd.notna(row[0]) else ""
        if not kw: continue
        vals = {}
        for j, t in enumerate(therapists):
            v = row[j+1]
            vals[t] = int(v) if pd.notna(v) and v != "" else 0
        data[kw] = vals
    return {"therapists": therapists, "data": data}


# ═══════════════════════════════════════════════════════════════════════
# DASHBOARD-WRITER  — baut alle 8 Sheets
# ═══════════════════════════════════════════════════════════════════════

def write_uebersicht(wb, ua_data, rl_data, hy_data, inv_data, fit_data, today_str):
    ws = wb.active; ws.title = "📊 Übersicht"
    ws.sheet_view.showGridLines = False
    ov = ua_data.get("overview", {})
    months_avail = ua_data.get("months_available", [])
    month_order = ["Jan26","Feb26","Mär26","Apr26","Mai26","Jun26","Jul26","Aug26","Sep26","Okt26","Nov26","Dez26"]
    month_labels = {
        "Jan26": "Jan 2026",
        "Feb26": "Feb 2026",
        "Mär26": "Mär 2026",
        "Apr26": "Apr 2026",
        "Mai26": "Mai 2026",
        "Jun26": "Jun 2026",
        "Jul26": "Jul 2026",
        "Aug26": "Aug 2026",
        "Sep26": "Sep 2026",
        "Okt26": "Okt 2026",
        "Nov26": "Nov 2026",
        "Dez26": "Dez 2026",
    }
    months_sorted = [m for m in month_order if m in months_avail]

    # Gesamtbreite des Sheets: Sektion A braucht Label + Monatsspalten + Total,
    # übrige Sektionen nutzen bis zu 12 Spalten. So bleibt der Header-Balken
    # immer bündig mit der breitesten Tabelle, egal wie viele Monate dazukommen.
    n_sec_a = 2 + len(months_sorted)  # Label + n_months + Total
    n_full = max(12, n_sec_a)
    mt = ua_data.get("by_month_totals", {})
    jan_t = mt.get("Jan26", {})
    feb_t = mt.get("Feb26", {})

    month_summary = ua_data.get("month_summary", {}) or {}

    def month_umsatz(month_key: str) -> float:
        rows = ua_data.get("by_month", {}).get(month_key, []) or []
        return float(sum(float(t.get("umsatz") or 0) for t in rows))

    def month_kpi(month_key: str, key: str):
        return (month_summary.get(month_key, {}) or {}).get(key)

    jan_u   = month_umsatz("Jan26") or (ov.get("umsatz_jan", 0) or 0)
    feb_u   = month_umsatz("Feb26") or (ov.get("umsatz_feb", 0) or 0)
    mar_u   = month_umsatz("Mär26")
    jan_zB  = ov.get("ziel_jan",   0) or 0
    feb_zB  = ov.get("ziel_feb",   0) or 0
    jan_fte = ov.get("fte_jan",    0) or 0
    feb_fte = ov.get("fte_feb",    0) or 0

    # Dynamische Periode-KPIs (skalieren mit jedem neuen Monat)
    period_u_sum = sum(month_umsatz(m) for m in months_sorted)
    ziel_vals = [month_kpi(m, "ziel") for m in months_sorted if month_kpi(m, "ziel") is not None]
    # Fallback auf ov.ziel_jan/ziel_feb falls month_summary leer ist
    if not ziel_vals:
        ziel_vals = [v for v in [jan_zB, feb_zB] if v]
    period_ziel_avg = (sum(ziel_vals) / len(ziel_vals)) if ziel_vals else 0
    last_month = months_sorted[-1] if months_sorted else "Jan26"
    last_lbl = MONTH_KPI_LABEL.get(last_month, last_month)
    last_u = month_umsatz(last_month)
    last_ziel = month_kpi(last_month, "ziel")
    if last_ziel is None:
        last_ziel = {"Jan26": jan_zB, "Feb26": feb_zB}.get(last_month, 0)
    range_lbl_kpi = (
        last_lbl if len(months_sorted) <= 1
        else f"{MONTH_KPI_LABEL.get(months_sorted[0], months_sorted[0])} – {last_lbl}"
    )

    # Runnerslab: alle verfügbaren Monate 2026 (skaliert mit den vorhandenen Therapeuten-Monaten)
    rl_month_labels = [MONTH_KPI_LABEL.get(m, m) for m in months_sorted]
    rl_period = sum(
        rl_data.get(2026, {}).get(lbl, {}).get("total", 0) for lbl in rl_month_labels
    )
    hy_total = hy_data.get("total_revenue", 0) or 0
    hy_rx = float(hy_data.get("revenue_hyrox_courses", 0) or 0)
    hy_pt = float(hy_data.get("revenue_personal_training", 0) or 0)
    if hy_total and (hy_rx + hy_pt) == 0:
        hy_rx = float(hy_total)
    hy_rev_src = hy_data.get("revenue_source", "")
    hy_sub = "Rechnungen" if hy_rev_src == "invoices" else ("Pässe" if hy_rev_src == "passes" else "Hyrox")
    inv_total = sum(i["betrag"] for i in inv_data) if inv_data else 0
    inv_count = len(inv_data)
    fitness_total = None
    for r in reversed(fit_data or []):
        g = r.get("Mitglieder Gesamt")
        if g: fitness_total = int(g); break

    title_bar(ws, 1, f"Kineo AG – Management Dashboard  ·  Stand {today_str}", cols=n_full)
    ws.merge_cells(f"A2:{gcl(n_full)}2"); ws["A2"].fill = fill(BLUE); ws.row_dimensions[2].height = 4

    kpis = [
        (1, f"Therapeuten Σ",        f"CHF {period_u_sum:,.0f}".replace(",","'"),  range_lbl_kpi,      BLUE),
        (2, f"Therap. {last_lbl}",   f"CHF {last_u:,.0f}".replace(",","'"),        "Letzter Monat",    BLUE),
        (3, "Ø Zielerr.",            f"{period_ziel_avg:.1%}" if period_ziel_avg else "–", range_lbl_kpi, ORANGE),
        (4, f"Zielerr. {last_lbl}",  f"{last_ziel:.1%}" if last_ziel else "–",     "Letzter Monat",    ORANGE),
        (5, "Runnerslab",            f"CHF {rl_period:,.0f}".replace(",","'"),     range_lbl_kpi,      GREEN),
        (6, "Hyrox o. PT",           f"CHF {hy_rx:,.0f}".replace(",","'"),         hy_sub,             TEAL),
        (7, "Hyrox PT",              f"CHF {hy_pt:,.0f}".replace(",","'"),         "Pers. Training",   PURPLE),
        (8, "Ausstehend",            f"CHF {inv_total:,.0f}".replace(",","'"),     f"{inv_count} Re.", RED),
        (9, "Fitness-Abos",          f"{fitness_total or '–'} Mitgl.",             "Thalwil",          GREEN),
    ]
    for col, label, val, sub, color in kpis:
        kpi_tile(ws, 4, col, label, val, sub, color)
    gap(ws, 7, n_full)

    # A – Therapeuten Kennzahlen
    sec(ws, 8, "A │ THERAPEUTEN – Kennzahlen aus Umsatzanalyse", cols=n_full)
    # Bereichs-Label, z.B. "Jan – Mär" (wächst automatisch mit jedem neuen Monat)
    if months_sorted:
        first_lbl = MONTH_KPI_LABEL.get(months_sorted[0], months_sorted[0])
        last_lbl  = MONTH_KPI_LABEL.get(months_sorted[-1], months_sorted[-1])
        range_lbl = first_lbl if len(months_sorted) == 1 else f"{first_lbl} – {last_lbl}"
    else:
        range_lbl = "YTD"
    month_cols = [month_labels.get(m, m) for m in months_sorted]
    extra_months = [m for m in months_sorted if m not in ["Jan26", "Feb26"]]
    hdr(ws, 9, ["Kennzahl"] + month_cols + [f"Σ/Ø {range_lbl}"])

    def r(v): return round(v, 1) if isinstance(v, float) else v

    # Tage + Stunden Format (8.4h pro Tag)
    def th(tage, std): return f"{r(tage)} T / {r(std)} h" if tage else "0 T"
    j_krank_h = r(jan_t.get("krank",0)*8.4); f_krank_h = r(feb_t.get("krank",0)*8.4)
    j_kurs_h  = r(jan_t.get("kurse",0)*8.4); f_kurs_h  = r(feb_t.get("kurse",0)*8.4)
    j_mktg_h  = r(jan_t.get("mktg",0)*8.4);  f_mktg_h  = r(feb_t.get("mktg",0)*8.4)
    j_lauf_h  = r(jan_t.get("lauf",0)*8.4);  f_lauf_h  = r(feb_t.get("lauf",0)*8.4)
    j_mgmt_h  = r(jan_t.get("mgmt",0)*8.4);  f_mgmt_h  = r(feb_t.get("mgmt",0)*8.4)

    # YTD = sum over all available months
    def ytd(key, default_jan=0, default_feb=0):
        """Sum a key over all available months from by_month_totals"""
        total = 0
        for m in months_avail:
            total += mt.get(m, {}).get(key, 0)
        return round(total, 1)

    # YTD-Umsatz direkt aus verfügbaren Monats-Detaildaten (inkl. März usw.)
    ytd_umsatz = sum(month_umsatz(m) for m in months_avail)

    a_rows = [
        ("Gesamtumsatz (CHF)",    jan_u,               feb_u,               ytd_umsatz, "#'##0.00"),
        ("FTE",                   jan_fte,              feb_fte,             (jan_fte+feb_fte)/2,  "0.0"),
        ("Umsatz pro FTE (CHF)",  ov.get("umsatz_fte_jan",0) or 0,
                                  ov.get("umsatz_fte_feb",0) or 0,
                                  ((ov.get("umsatz_fte_jan",0) or 0)+(ov.get("umsatz_fte_feb",0) or 0))/2, "#'##0.00"),
        ("Ø Zielerreichung (B)",  jan_zB,              feb_zB,              (jan_zB+feb_zB)/2,   "0.0%"),
        ("Ferientage",
         th(jan_t.get("ferien",0), r(jan_t.get("ferien",0)*8.4)),
         th(feb_t.get("ferien",0), r(feb_t.get("ferien",0)*8.4)),
         th(sum(mt.get(m,{}).get("ferien",0) for m in months_avail),
            r(sum(mt.get(m,{}).get("ferien",0)*8.4 for m in months_avail))), "0"),
        ("Krankheitstage",
         th(jan_t.get("krank",0), j_krank_h),
         th(feb_t.get("krank",0), f_krank_h),
         th(ytd("krank"), r(ytd("krank")*8.4)), "0"),
        ("Fitnesskurse",
         th(jan_t.get("kurse",0), j_kurs_h),
         th(feb_t.get("kurse",0), f_kurs_h),
         th(ytd("kurse"), r(ytd("kurse")*8.4)), "0"),
        ("Marketing",
         th(jan_t.get("mktg",0), j_mktg_h),
         th(feb_t.get("mktg",0), f_mktg_h),
         th(ytd("mktg"), r(ytd("mktg")*8.4)), "0"),
        ("Laufanalyse",
         th(jan_t.get("lauf",0), j_lauf_h),
         th(feb_t.get("lauf",0), f_lauf_h),
         th(ytd("lauf"), r(ytd("lauf")*8.4)), "0"),
        ("Management",
         th(jan_t.get("mgmt",0), j_mgmt_h),
         th(feb_t.get("mgmt",0), f_mgmt_h),
         th(ytd("mgmt"), r(ytd("mgmt")*8.4)), "0"),
        ("Abwesenheiten Total",
         th(sum(jan_t.get(k,0) for k in ["ferien","krank","kurse","mktg","lauf","mgmt"]),
            r(sum(jan_t.get(k,0)*8.4 for k in ["ferien","krank","kurse","mktg","lauf","mgmt"]))),
         th(sum(feb_t.get(k,0) for k in ["ferien","krank","kurse","mktg","lauf","mgmt"]),
            r(sum(feb_t.get(k,0)*8.4 for k in ["ferien","krank","kurse","mktg","lauf","mgmt"]))),
         th(ytd("ferien")+ytd("krank")+ytd("kurse")+ytd("mktg")+ytd("lauf")+ytd("mgmt"),
            r((ytd("ferien")+ytd("krank")+ytd("kurse")+ytd("mktg")+ytd("lauf")+ytd("mgmt"))*8.4)), "0"),
    ]
    n_months_data = len([m for m in months_sorted if "26" in m])

    r_start = 10
    for i, (label, jv, fv, tot, fmt) in enumerate(a_rows, r_start):
        bg = ALT if i%2==0 else WHITE
        dc(ws, i, 1, label, bg, bold=True, h="left")
        # Monatswerte dynamisch schreiben: Jan, Feb, (optional Mär/Apr/...), Total
        month_values = [jv, fv]
        if extra_months:
            if label == "Gesamtumsatz (CHF)":
                month_values += [month_umsatz(m) for m in extra_months]
            elif label == "FTE":
                month_values += [month_kpi(m, "fte") for m in extra_months]
            elif label == "Umsatz pro FTE (CHF)":
                month_values += [month_kpi(m, "umsatz_fte") for m in extra_months]
            elif label == "Ø Zielerreichung (B)":
                month_values += [month_kpi(m, "ziel") for m in extra_months]
            elif label == "Ferientage":
                month_values += [th(mt.get(m, {}).get("ferien", 0), r(mt.get(m, {}).get("ferien", 0) * 8.4)) for m in extra_months]
            elif label == "Krankheitstage":
                month_values += [th(mt.get(m, {}).get("krank", 0), r(mt.get(m, {}).get("krank", 0) * 8.4)) for m in extra_months]
            elif label == "Fitnesskurse":
                month_values += [th(mt.get(m, {}).get("kurse", 0), r(mt.get(m, {}).get("kurse", 0) * 8.4)) for m in extra_months]
            elif label == "Marketing":
                month_values += [th(mt.get(m, {}).get("mktg", 0), r(mt.get(m, {}).get("mktg", 0) * 8.4)) for m in extra_months]
            elif label == "Laufanalyse":
                month_values += [th(mt.get(m, {}).get("lauf", 0), r(mt.get(m, {}).get("lauf", 0) * 8.4)) for m in extra_months]
            elif label == "Management":
                month_values += [th(mt.get(m, {}).get("mgmt", 0), r(mt.get(m, {}).get("mgmt", 0) * 8.4)) for m in extra_months]
            elif label == "Abwesenheiten Total":
                month_values += [
                    th(
                        sum(mt.get(m, {}).get(k, 0) for k in ["ferien", "krank", "kurse", "mktg", "lauf", "mgmt"]),
                        r(sum(mt.get(m, {}).get(k, 0) * 8.4 for k in ["ferien", "krank", "kurse", "mktg", "lauf", "mgmt"])),
                    )
                    for m in extra_months
                ]
            else:
                month_values += [None for _ in extra_months]

        for j, v in enumerate(month_values, start=2):
            is_str = isinstance(v, str)
            if v is not None:
                dc(ws, i, j, v, bg, fmt=None if is_str else fmt, h="left" if is_str else "right")
            else:
                dc(ws, i, j, "–", bg, h="center")

        total_col = 2 + len(month_values)
        if tot is not None:
            dc(
                ws, i, total_col, tot, bg, bold=True,
                fmt=None if isinstance(tot, str) else fmt,
                h="left" if isinstance(tot, str) else "right"
            )
        else:
            dc(ws, i, total_col, "–", bg, h="center")

    # Info-Zeile: Datengrundlage (Anzahl Monate)
    info_row = r_start + len(a_rows)
    info_text = f"ℹ️  Datengrundlage: {n_months_data} Monat{'e' if n_months_data != 1 else ''} ({', '.join(month_labels.get(m, m) for m in months_sorted)})"
    ws.merge_cells(f"A{info_row}:{gcl(max(n_full, 2 + len(months_sorted)))}{info_row}")
    info_c = ws.cell(info_row, 1, info_text)
    info_c.font = fnt(sz=8, color="475569", italic=True)
    info_c.fill = fill(LBLUE)
    info_c.alignment = aln("left")
    ws.row_dimensions[info_row].height = 14

    # Dynamic row positions
    r_b  = r_start + len(a_rows) + 2   # Section B start (after info row + gap)
    r_b1 = r_b + 1                      # B header
    r_b2 = r_b + 2                      # B data start
    gap(ws, r_b - 1, n_full)

    # B – Runnerslab Jahresvergleich
    sec(ws, r_b, "B │ RUNNERSLAB – Jan+Feb Jahresvergleich", cols=n_full)
    hdr(ws, r_b1, ["Jahr","Jan Total","Feb Total","Jan+Feb","Δ% vs. Vorjahr"])
    years_to_show = sorted([y for y in rl_data.keys() if y >= 2021])
    prev_jf = None
    for i, yr in enumerate(years_to_show, r_b2):
        bg = ALT if i%2==0 else WHITE
        rl_jan = rl_data[yr].get("Jan",{}).get("total",0)
        rl_feb = rl_data[yr].get("Feb",{}).get("total",0)
        jf = rl_jan + rl_feb
        dc(ws, i, 1, str(yr), bg, bold=True, h="center")
        dc(ws, i, 2, rl_jan, bg, fmt="#'##0")
        dc(ws, i, 3, rl_feb, bg, fmt="#'##0")
        dc(ws, i, 4, jf, bg, bold=True, fmt="#'##0")
        if prev_jf and prev_jf > 0:
            delta = (jf - prev_jf) / prev_jf
            c1 = ws.cell(i, 5, delta); c1.number_format="0.0%"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,size=10,color=GREEN if delta>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
        else:
            dc(ws, i, 5, "–", bg, h="center")
        prev_jf = jf

    r_c_gap = r_b2 + len(years_to_show)
    gap(ws, r_c_gap, n_full)

    # C – Runnerslab Monat für Monat
    r_c = r_c_gap + 1
    months_s = ["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"]
    show_years = sorted([y for y in rl_data.keys() if y >= 2024])
    sec(ws, r_c, "C │ RUNNERSLAB – Monat für Monat (alle verfügbaren Jahre)", cols=n_full)
    hdr(ws, r_c+1, ["Monat"] + [str(y) for y in show_years] + ["Δ letzt.2J","Δ%"])
    for i, m in enumerate(months_s):
        rr = r_c+2+i; bg = ALT if rr%2==0 else WHITE
        dc(ws, rr, 1, m, bg, bold=True, h="left")
        vals = []
        for j, yr in enumerate(show_years):
            v = rl_data[yr].get(m,{}).get("total",0) or 0
            dc(ws, rr, j+2, v if v else None, bg, fmt="#'##0", bold=(j==len(show_years)-1))
            vals.append(v)
        if len(vals) >= 2 and vals[-2] and vals[-1]:
            d = vals[-1]-vals[-2]; dp = d/vals[-2]
            c1=ws.cell(rr,len(show_years)+2,d); c1.number_format="#'##0"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,size=10,color=GREEN if d>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
            c2=ws.cell(rr,len(show_years)+3,dp); c2.number_format="0.0%"; c2.fill=fill(bg)
            c2.border=bb(); c2.alignment=aln("right")
        else:
            for cn in [len(show_years)+2, len(show_years)+3]:
                dc(ws, rr, cn, "–", bg, h="center")

    r_d_gap = r_c + 2 + 12
    gap(ws, r_d_gap, n_full)

    # D – Hyrox (Monatsübersicht: NUR effektive Kursbuchungen / Auslastung)
    r_d = r_d_gap + 1
    sec(ws, r_d, "D │ HYROX – Kursbuchungen & Auslastung (aus Schedule)", cols=n_full)
    hdr(ws, r_d+1, ["Monat","Kurse","Buchungen (TN)","Plätze","Auslastung Ø","Ø TN/Kurs"])

    # Wochen-Daten (aus Schedule) → Monatsaggregation
    weekly = hy_data.get("weekly", []) or []

    def month_label(period_str: str) -> str:
        """Wandelt 'YYYY-MM' in 'Monat YYYY' (deutsch) um."""
        try:
            y, m = period_str.split("-")
            m = int(m)
            monate = ["Januar","Februar","März","April","Mai","Juni",
                      "Juli","August","September","Oktober","November","Dezember"]
            return f"{monate[m-1]} {y}"
        except Exception:
            return period_str

    # Hilfsstruktur: aggregiere Weekly-Daten nach Monat
    weekly_by_month: dict[str, dict] = {}
    for w in weekly:
        d_str = w.get("date")
        if not d_str:
            continue
        try:
            # Erwartetes Format: 'dd.mm.yy'
            day, mon, yr = d_str.split(".")
            mon_i = int(mon)
            yr_i = int("20" + yr) if len(yr) == 2 else int(yr)
            key = f"{yr_i}-{mon_i:02d}"
        except Exception:
            continue
        agg = weekly_by_month.setdefault(key, {"kurse": 0, "tn": 0, "plaetze": 0, "ausl_sum": 0.0, "cnt": 0})
        kurse = w.get("kurse") or 0
        tn = w.get("tn") or 0
        plaetze = w.get("plaetze") or 0
        ausl = w.get("auslastung") or 0
        agg["kurse"] += kurse
        agg["tn"] += tn
        agg["plaetze"] += plaetze
        agg["ausl_sum"] += ausl
        agg["cnt"] += 1

    # Sortiere nach Monatsschlüssel (YYYY-MM)
    month_keys = sorted(weekly_by_month.keys())

    row = r_d + 2
    for key in month_keys:
        w_agg = weekly_by_month.get(key, {})
        kurse = w_agg.get("kurse", 0)
        tn = w_agg.get("tn", 0)
        plaetze = w_agg.get("plaetze", 0)
        cnt = w_agg.get("cnt", 0)
        avg_ausl = (w_agg.get("ausl_sum", 0.0) / cnt) if cnt else None
        avg_tn = (tn / kurse) if kurse else None

        bg = ALT if row % 2 == 0 else WHITE
        dc(ws, row, 1, month_label(key), bg, bold=True, h="left")
        dc(ws, row, 2, kurse if kurse else None, bg, h="center")
        dc(ws, row, 3, tn if tn else None, bg, h="center")
        dc(ws, row, 4, plaetze if plaetze else None, bg, h="center")

        if avg_ausl is not None:
            c_ = ws.cell(row, 5, avg_ausl)
            c_.number_format = "0.0%"
            c_.border = bb()
            c_.alignment = aln("center")
            c_.fill = fill("D1FAE5" if avg_ausl >= 0.65 else ("FEF3C7" if avg_ausl >= 0.5 else "FEE2E2"))
            c_.font = Font(
                name="Arial",
                bold=True,
                size=10,
                color=GREEN if avg_ausl >= 0.65 else (ORANGE if avg_ausl >= 0.5 else RED),
            )
        else:
            dc(ws, row, 5, "–", bg, h="center")

        dc(ws, row, 6, avg_tn, bg, fmt="0.00", h="center") if avg_tn is not None else dc(ws, row, 6, "–", bg, h="center")
        row += 1

    # D2 – Umsatz: Hyrox (Kurse/Pässe) vs Personal Training (pro Monat)
    r_d2 = row + 1
    gap(ws, r_d2, n_full)
    r_d2 += 1
    sec(ws, r_d2, "D2 │ HYROX – Umsatz: Kurse/Pässe vs Personal Training (pro Monat)", cols=n_full)
    hdr(ws, r_d2 + 1, ["Monat", "Hyrox o. PT (CHF)", "Personal Training (CHF)", "Total (CHF)"])
    mh = hy_data.get("revenue_monthly_hyrox") or {}
    mp = hy_data.get("revenue_monthly_pt") or {}
    months_rev = sorted(set(mh.keys()) | set(mp.keys()))
    r2 = r_d2 + 2
    if months_rev:
        for mk in months_rev:
            vh = float(mh.get(mk, 0) or 0)
            vp = float(mp.get(mk, 0) or 0)
            bg = ALT if r2 % 2 == 0 else WHITE
            dc(ws, r2, 1, month_label(mk), bg, bold=True, h="left")
            dc(ws, r2, 2, vh if vh else None, bg, fmt="#'##0.00")
            dc(ws, r2, 3, vp if vp else None, bg, fmt="#'##0.00")
            dc(ws, r2, 4, vh + vp, bg, bold=True, fmt="#'##0.00")
            r2 += 1
    else:
        dc(
            ws,
            r2,
            1,
            "Keine Monatsumsätze (Kaufdatum in Rechnungen/Pässen fehlt oder keine Hyrox-Daten).",
            ALT,
            h="left",
        )
        ws.merge_cells(f"A{r2}:D{r2}")
        r2 += 1

    # Dynamische Spaltenbreiten – wachsen automatisch mit jedem neuen Monat mit
    # - Spalte 1: Label (z.B. "Gesamtumsatz (CHF)", Jahr, "Januar 2026")
    # - Monats-Spalten: breit genug für "CHF 282'897" oder "21.0 T / 176.4 h"
    # - Total-Spalte: etwas breiter, fett formatiert
    # - Restliche Spalten: breit genug für CHF-KPI-Kacheln wie "CHF 37'486" (13pt bold)
    n_mo = len(months_sorted)
    ws.column_dimensions['A'].width = 24
    for i in range(n_mo):
        ws.column_dimensions[gcl(2 + i)].width = 15
    ws.column_dimensions[gcl(2 + n_mo)].width = 17  # Total-Spalte
    for c in range(2 + n_mo + 1, n_full + 1):
        ws.column_dimensions[gcl(c)].width = 13

def write_therapeuten(wb, ua_data, today_str):
    ws = wb.create_sheet("👥 Therapeuten")
    ws.sheet_view.showGridLines = False
    months = ua_data.get("months_available", [])
    ov = ua_data.get("overview", {})
    ther_list = ov.get("therapeuten", [])
    month_summary = ua_data.get("month_summary") or {}
    mt = ua_data.get("by_month_totals") or {}

    month_order_tp = [
        "Jan26", "Feb26", "Mär26", "Apr26", "Mai26", "Jun26", "Jul26", "Aug26",
        "Sep26", "Okt26", "Nov26", "Dez26",
    ]
    months_sorted = [m for m in month_order_tp if m in months]
    extra_months = [m for m in months_sorted if m not in ("Jan26", "Feb26")]
    n_months = len(months_sorted)

    def month_umsatz(month_key: str) -> float:
        rows = ua_data.get("by_month", {}).get(month_key, []) or []
        return float(sum(float(t.get("umsatz") or 0) for t in rows))

    def total_umsatz_kpi(mk: str) -> float:
        s = month_umsatz(mk)
        if s:
            return s
        if mk == "Jan26":
            return float(ov.get("umsatz_jan") or 0)
        if mk == "Feb26":
            return float(ov.get("umsatz_feb") or 0)
        return 0.0

    def lbl(mk: str) -> str:
        return MONTH_KPI_LABEL.get(mk, mk)

    def kpi_ziel_agg(mk: str):
        if mk == "Jan26":
            return ov.get("ziel_jan")
        if mk == "Feb26":
            return ov.get("ziel_feb")
        return (month_summary.get(mk) or {}).get("ziel")

    def kpi_fte(mk: str):
        if mk == "Jan26":
            return ov.get("fte_jan")
        if mk == "Feb26":
            return ov.get("fte_feb")
        return (month_summary.get(mk) or {}).get("fte")

    def kpi_ufte(mk: str):
        if mk == "Jan26":
            return ov.get("umsatz_fte_jan")
        if mk == "Feb26":
            return ov.get("umsatz_fte_feb")
        return (month_summary.get(mk) or {}).get("umsatz_fte")

    def kpi_krank(mk: str):
        if mk == "Jan26":
            return ov.get("krank_jan")
        if mk == "Feb26":
            return ov.get("krank_feb")
        return mt.get(mk, {}).get("krank", 0)

    def ther_u(t, mk: str):
        if mk == "Jan26":
            return t.get("jan_umsatz")
        if mk == "Feb26":
            return t.get("feb_umsatz")
        keys = MONTH_THER_KEYS.get(mk)
        return t.get(keys[0]) if keys else None

    def ther_zb(t, mk: str):
        if mk == "Jan26":
            return t.get("jan_zielB")
        if mk == "Feb26":
            return t.get("feb_zielB")
        keys = MONTH_THER_KEYS.get(mk)
        return t.get(keys[1]) if keys else None

    # ── 5 feste Overall-KPI-Kacheln (unabhängig von Monatsanzahl) ───────
    total_u_all = sum(total_umsatz_kpi(mk) for mk in months_sorted)
    ziel_vals = [kpi_ziel_agg(mk) for mk in months_sorted]
    ziel_vals = [z for z in ziel_vals if z is not None]
    avg_ziel_all = (sum(ziel_vals) / len(ziel_vals)) if ziel_vals else None
    fte_vals = [kpi_fte(mk) for mk in months_sorted]
    fte_vals = [f for f in fte_vals if f is not None]
    avg_fte_all = (sum(fte_vals) / len(fte_vals)) if fte_vals else None
    ufte_vals = [kpi_ufte(mk) for mk in months_sorted]
    ufte_vals = [f for f in ufte_vals if f is not None]
    avg_ufte_all = (sum(ufte_vals) / len(ufte_vals)) if ufte_vals else None
    total_krank_all = sum(float(kpi_krank(mk) or 0) for mk in months_sorted)

    if months_sorted:
        first_lbl_t = lbl(months_sorted[0])
        last_lbl_t  = lbl(months_sorted[-1])
        range_lbl_t = first_lbl_t if len(months_sorted) == 1 else f"{first_lbl_t} – {last_lbl_t}"
        period_label = f"{range_lbl_t} 2026"
    else:
        range_lbl_t = "YTD"
        period_label = "2026"

    # Header-Breite = Ranking-Breite (dynamisch mit Monatszahl)
    hdr_cols = ["Rang", "Therapeut/in", "BG%"]
    for mk in months_sorted:
        hdr_cols.append(f"{lbl(mk)} Umsatz")
        hdr_cols.append(f"{lbl(mk)} Ziel(B)")
    hdr_cols.extend([f"Σ Umsatz {range_lbl_t}", f"Ø Ziel(B) {range_lbl_t}", "Trend", "Note"])
    n_hdr = len(hdr_cols)

    title_bar(ws, 1, f"Therapeuten – Umsatz & Zielerreichung  ·  {today_str}", cols=n_hdr)
    ws.merge_cells(f"A2:{gcl(n_hdr)}2")
    ws["A2"].fill = fill(BLUE)
    ws.row_dimensions[2].height = 4

    # 5 Kacheln je 2 Spalten breit (cols 1-2, 3-4, 5-6, 7-8, 9-10) – unabhängig von Monatsanzahl
    overall_kpis = [
        (1, f"Umsatz {period_label}", f"CHF {total_u_all:,.0f}".replace(",", "'"), "alle verfügb. Monate", BLUE),
        (3, "Ø Zielerr. (B)", f"{avg_ziel_all:.1%}" if avg_ziel_all is not None else "–", "Monats-Ø", ORANGE),
        (5, "Ø FTE", f"{avg_fte_all:.1f}" if avg_fte_all is not None else "–", "Monats-Ø", GREEN),
        (7, "Ø Umsatz/FTE", f"CHF {avg_ufte_all:,.0f}".replace(",", "'") if avg_ufte_all is not None else "–", "Monats-Ø", TEAL),
        (9, "Krank Total", f"{total_krank_all:.0f} T", "Summe alle Monate", RED),
    ]
    for c_, lab, val, sub, color in overall_kpis:
        kpi_tile(ws, 4, c_, lab, val, sub, color, span=2)
    gap(ws, 7, n_hdr)

    # ── Kompakte Monatsübersicht (wächst horizontal mit jedem Monat) ────
    r_ms = 8
    sec(ws, r_ms, "MONATSÜBERSICHT – Kennzahlen pro Monat", cols=n_hdr)
    ms_headers = ["Kennzahl"] + [lbl(m) for m in months_sorted] + [f"Σ/Ø {range_lbl_t}"]
    n_ms = len(ms_headers)
    hdr(ws, r_ms + 1, ms_headers)
    # Restliche Header-Spalten füllen (konsistente Optik bis n_hdr)
    for c in range(n_ms + 1, n_hdr + 1):
        cc = ws.cell(r_ms + 1, c, "")
        cc.fill = fill("1E3A5F"); cc.border = fb("2C4A7C")

    ms_rows = [
        ("Gesamtumsatz (CHF)", [total_umsatz_kpi(mk) for mk in months_sorted], total_u_all, "#'##0"),
        ("Ø Zielerreichung (B)", [kpi_ziel_agg(mk) for mk in months_sorted], avg_ziel_all, "0.0%"),
        ("FTE", [kpi_fte(mk) for mk in months_sorted], avg_fte_all, "0.0"),
        ("Umsatz / FTE (CHF)", [kpi_ufte(mk) for mk in months_sorted], avg_ufte_all, "#'##0"),
        ("Krankheitstage", [float(kpi_krank(mk) or 0) for mk in months_sorted], total_krank_all, "0.0"),
        ("Ferientage",
         [float(mt.get(mk, {}).get("ferien", 0) or 0) for mk in months_sorted],
         float(sum(mt.get(mk, {}).get("ferien", 0) for mk in months_sorted)),
         "0.0"),
    ]
    r_cur = r_ms + 2
    for label, vals, tot, fmt in ms_rows:
        bg = ALT if r_cur % 2 == 0 else WHITE
        dc(ws, r_cur, 1, label, bg, bold=True, h="left")
        for i_v, v in enumerate(vals, start=2):
            if v is None:
                dc(ws, r_cur, i_v, "–", bg, h="center")
            else:
                dc(ws, r_cur, i_v, v, bg, fmt=fmt, h="center")
        total_c = 2 + len(vals)
        if tot is None:
            dc(ws, r_cur, total_c, "–", bg, h="center")
        else:
            dc(ws, r_cur, total_c, tot, bg, bold=True, fmt=fmt, h="center")
        # Leere Zellen bis n_hdr für saubere Optik
        for c in range(total_c + 1, n_hdr + 1):
            cc = ws.cell(r_cur, c)
            cc.fill = fill(bg); cc.border = bb()
        r_cur += 1

    gap(ws, r_cur, n_hdr)

    def avg_zB(t):
        vals = [ther_zb(t, mk) for mk in months_sorted]
        vals = [v for v in vals if v is not None and v != 0]
        return sum(vals) / len(vals) if vals else 0

    ranked = sorted(ther_list, key=avg_zB, reverse=True)

    r_rank_sec = r_cur + 1
    r_rank_hdr = r_rank_sec + 1
    sec(ws, r_rank_sec, "RANKING – Ø Zielerreichung (B) · pensenbereinigt · fairster Vergleich", cols=n_hdr)
    hdr(ws, r_rank_hdr, hdr_cols, h=20)

    total_col = 4 + 2 * n_months
    avg_col = total_col + 1
    trend_col = avg_col + 1
    note_col = trend_col + 1

    for rank, t in enumerate(ranked, 1):
        i = r_rank_hdr + rank
        bg = ALT if i % 2 == 0 else WHITE
        medal = "🥇" if rank == 1 else ("🥈" if rank == 2 else ("🥉" if rank == 3 else str(rank)))
        avg_z = avg_zB(t)
        zseq = [ther_zb(t, mk) for mk in months_sorted if ther_zb(t, mk) is not None]
        if len(zseq) >= 2:
            z_prev, z_last = zseq[-2], zseq[-1]
            trend = "▲" if z_last > z_prev else ("▼" if z_last < z_prev else "=")
        else:
            trend = "–"
        trend_c = GREEN if trend == "▲" else (RED if trend == "▼" else "94A3B8")
        ws.cell(i, 1, rank if rank > 3 else medal).fill = fill(bg)
        ws.cell(i, 1).font = fnt(bold=True, sz=10)
        ws.cell(i, 1).border = bb()
        ws.cell(i, 1).alignment = aln("center")
        dc(ws, i, 2, t["name"], bg, bold=True, h="left")
        dc(ws, i, 3, (t.get("bg") or 0) / 100, bg, fmt="0%", h="center")
        for mi, mk in enumerate(months_sorted):
            c_u = 4 + 2 * mi
            c_z = 5 + 2 * mi
            uu = ther_u(t, mk)
            if uu is not None:
                dc(ws, i, c_u, uu, bg, fmt="#'##0.00")
            else:
                dc(ws, i, c_u, "–", bg, h="center")
            zz = ther_zb(t, mk)
            zell(ws, i, c_z, zz, bg)
        tot_u = sum((ther_u(t, mk) or 0) for mk in months_sorted)
        dc(ws, i, total_col, tot_u, bg, bold=True, fmt="#'##0.00")
        zell(ws, i, avg_col, avg_z if avg_z > 0 else None, bg)
        dc(ws, i, trend_col, trend, bg, h="center", bold=True, color=trend_c)
        note = "Teilzeit" if (t.get("bg") or 100) <= 50 else ""
        dc(ws, i, note_col, note, bg, h="center", sz=9, color="64748B")

    r_rt = r_rank_hdr + len(ranked) + 1
    # TOTAL/Ø als feste Werte (inkl. aller Monate wie Mär) — sichtbar auch ohne Excel-Neuberechnung
    footer_vals: list = ["", "TOTAL/Ø", ""]
    for mk in months_sorted:
        us = [ther_u(t, mk) for t in ranked if ther_u(t, mk) is not None]
        zs = [ther_zb(t, mk) for t in ranked if ther_zb(t, mk) is not None]
        sum_u = sum(us) if us else None
        avg_z_m = (sum(zs) / len(zs)) if zs else None
        footer_vals.append(sum_u)
        footer_vals.append(avg_z_m)
    tot_um_all = sum(
        sum((ther_u(t, mk) or 0) for mk in months_sorted) for t in ranked
    ) or None
    footer_vals.append(tot_um_all)
    avg_z_rank_footer = (
        sum(avg_zB(t) for t in ranked) / len(ranked) if ranked else None
    )
    footer_vals.append(avg_z_rank_footer)
    footer_vals.extend(["", ""])
    # None → "–" damit leere Monate sichtbar sind statt 0.00
    footer_vals = [("–" if v is None else v) for v in footer_vals]
    hdr(ws, r_rt, footer_vals)
    for c in range(4, total_col + 1, 2):
        ws.cell(r_rt, c).number_format = "#'##0.00"
    for c in range(5, avg_col + 1, 2):
        ws.cell(r_rt, c).number_format = "0.0%"
    ws.cell(r_rt, total_col).number_format = "#'##0.00"
    ws.cell(r_rt, avg_col).number_format = "0.0%"
    gap(ws, r_rt + 1, n_hdr)

    # Detail-Tabellen pro Monat
    row_cur = r_rt+2
    for month in months:
        month_data = ua_data["by_month"].get(month, [])
        if not month_data: continue
        sec(ws,row_cur,f"{month} – Detail pro Therapeut/in",cols=11)
        hdr(ws,row_cur+1,["Therapeut/in","BG%","Ferien","Krank","Mgmt(T)",
                          "Kurse(T)","TageB","Umsatz","CHF/TagB","Ziel%(A)","Ziel%(B)"],h=18,sz=8)
        r=row_cur+2
        for t in month_data:
            bg=ALT if r%2==0 else WHITE
            dc(ws,r,1,t["name"],bg,bold=True,h="left")
            dc(ws,r,2,(t.get("bg") or 0)/100,bg,fmt="0%",h="center")
            dc(ws,r,3,t.get("ferien",0),bg,fmt="0.0",h="center")
            dc(ws,r,4,t.get("krank",0),bg,h="center")
            dc(ws,r,5,t.get("mgmt",0),bg,h="center")
            dc(ws,r,6,round(t.get("kurs_tage",0),2),bg,fmt="0.00",h="center")
            dc(ws,r,7,round(t.get("tageB",0),1),bg,fmt="0.0")
            u=t.get("umsatz"); dc(ws,r,8,u,bg,bold=True,fmt="#'##0.00") if u else dc(ws,r,8,"–",bg,h="center")
            fte=round(u/t["tageB"],2) if u and t.get("tageB") else None
            dc(ws,r,9,fte,bg,fmt="#'##0.00") if fte else dc(ws,r,9,"–",bg,h="center")
            zell(ws,r,10,t.get("zielA"),bg); zell(ws,r,11,t.get("zielB"),bg)
            r+=1
        # Totals
        r_mt=r
        hdr(ws,r_mt,[f"{month} TOTAL","","",
                     f"=SUM(D{row_cur+2}:D{r_mt-1})",f"=SUM(E{row_cur+2}:E{r_mt-1})","","",
                     f"=SUM(H{row_cur+2}:H{r_mt-1})","",
                     f"=AVERAGE(J{row_cur+2}:J{r_mt-1})",f"=AVERAGE(K{row_cur+2}:K{r_mt-1})"])
        for c in [8]: ws.cell(r_mt,c).number_format="#'##0.00"
        for c in [10,11]: ws.cell(r_mt,c).number_format="0.0%"
        row_cur = r_mt+2
        gap(ws,row_cur-1,11)

    # Umsatztracking
    tracking = ua_data.get("tracking", [])
    if tracking:
        sec(ws, row_cur, "UMSATZTRACKING – Monatsabschluss-Vergleich", cols=11)
        hdr(ws, row_cur + 1, ["Therapeut/in", "Stand 1", "Stand 2 (final)", "Delta", "Trend"])
        for i, t in enumerate(tracking, row_cur + 2):
            bg = ALT if i % 2 == 0 else WHITE
            dc(ws, i, 1, t["name"], bg, bold=True, h="left")
            dc(ws, i, 2, t.get("stand1"), bg, fmt="#'##0.00")
            dc(ws, i, 3, t.get("stand2"), bg, bold=True, fmt="#'##0.00")
            d = t.get("delta", 0)
            dc(
                ws, i, 4, d, bg, fmt="#'##0.00", bold=True,
                color=GREEN if d > 0 else (RED if d < 0 else "94A3B8"),
            )
            dc(
                ws, i, 5, "▲" if d > 0 else ("▼" if d < 0 else "="), bg, h="center", bold=True,
                color=GREEN if d > 0 else (RED if d < 0 else "94A3B8"),
            )

    note_row = row_cur + len(tracking) + 3
    merge_w = max(n_hdr, 11)
    ws.merge_cells(f"A{note_row}:{gcl(merge_w)}{note_row}")
    ws.cell(
        note_row, 1,
        "ℹ️  Ranking-Basis: Ziel(B) = Umsatz ÷ (TageB × CHF 1'040)  |  "
        "TageB = SollAT − Ferien, inkl. Kurse & Weiterb., exkl. Krank & Mgmt  |  "
        "Pensenbereinigt = fairer Vergleich Vollzeit/Teilzeit",
    ).font = fnt(sz=8, color="475569", italic=True)
    ws.cell(note_row, 1).fill = fill(LBLUE)
    ws.row_dimensions[note_row].height = 14

    # Ranking: col 1=Rang, col 2=Name, col 3=BG%, col 4..(3+2*n_months)=abwechselnd Umsatz/Ziel,
    # col total_col=Total Umsatz, avg_col=Ø Ziel, trend_col=Trend, note_col=Note
    for c in range(1, merge_w + 1):
        if c == 1:
            w = 8   # Rang
        elif c == 2:
            w = 18  # Name
        elif c == 3:
            w = 8   # BG%
        elif c < total_col and (c - 4) % 2 == 0:
            w = 13  # Monats-Umsatz (CHF ~ 11 Zeichen)
        elif c < total_col:
            w = 9   # Monats-Ziel(B) (84.7% ~ 6 Zeichen)
        elif c == total_col:
            w = 15  # Total Umsatz (größer, fett)
        elif c == avg_col:
            w = 10  # Ø Ziel(B)
        elif c == trend_col:
            w = 7
        elif c == note_col:
            w = 10
        else:
            w = 10
        ws.column_dimensions[gcl(c)].width = w


def write_runnerslab(wb, rl_data):
    ws = wb.create_sheet("👟 Runnerslab")
    ws.sheet_view.showGridLines = False
    title_bar(ws,1,"Runnerslab – Jahresvergleich & Monat-für-Monat",bg="1D4ED8",cols=10)

    months_s=["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"]
    years_sorted=sorted(rl_data.keys())

    # Jahrestabelle
    sec(ws,3,"Jahresgesamtumsatz",cols=10)
    hdr(ws,4,["Jahr","POS (CHF)","Online (CHF)","Total (CHF)","Δ CHF","Δ%"])
    prev_tot=None
    for i,yr in enumerate(years_sorted,5):
        bg=ALT if i%2==0 else WHITE
        pos=sum(rl_data[yr].get(m,{}).get("pos",0) for m in months_s)
        onl=sum(rl_data[yr].get(m,{}).get("online",0) for m in months_s)
        tot=sum(rl_data[yr].get(m,{}).get("total",0) for m in months_s)
        dc(ws,i,1,str(yr),bg,bold=True,h="center")
        dc(ws,i,2,pos,bg,fmt="#'##0"); dc(ws,i,3,onl,bg,fmt="#'##0")
        dc(ws,i,4,tot,bg,bold=True,fmt="#'##0")
        if prev_tot and prev_tot>0:
            d=tot-prev_tot; dp=d/prev_tot
            c1=ws.cell(i,5,d); c1.number_format="#'##0"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,sz=10,color=GREEN if d>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
            c2=ws.cell(i,6,dp); c2.number_format="0.0%"; c2.fill=fill(bg)
            c2.border=bb(); c2.alignment=aln("right")
        else:
            dc(ws,i,5,"–",bg,h="center"); dc(ws,i,6,"–",bg,h="center")
        prev_tot=tot

    gap(ws,5+len(years_sorted),10)

    # Monat für Monat
    r_mfm=6+len(years_sorted)
    show_yrs=sorted([y for y in years_sorted if y>=2024])
    sec(ws,r_mfm,"Monat für Monat",cols=10)
    hdr(ws,r_mfm+1,["Monat"]+[str(y) for y in show_yrs]+["Δ letzt.","Δ%"])
    for i,m in enumerate(months_s):
        r=r_mfm+2+i; bg=ALT if r%2==0 else WHITE
        dc(ws,r,1,m,bg,bold=True,h="left")
        vals=[]
        for j,yr in enumerate(show_yrs):
            v=rl_data[yr].get(m,{}).get("total",0) or 0
            dc(ws,r,j+2,v if v else None,bg,fmt="#'##0",bold=(j==len(show_yrs)-1))
            vals.append(v)
        if len(vals)>=2 and vals[-2] and vals[-1]:
            d=vals[-1]-vals[-2]; dp=d/vals[-2]
            c1=ws.cell(r,len(show_yrs)+2,d); c1.number_format="#'##0"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,sz=10,color=GREEN if d>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
            c2=ws.cell(r,len(show_yrs)+3,dp); c2.number_format="0.0%"; c2.fill=fill(bg)
            c2.border=bb(); c2.alignment=aln("right")
        else:
            for c_ in [len(show_yrs)+2,len(show_yrs)+3]:
                dc(ws,r,c_,"–",bg,h="center")

    # Chart Monatsverlauf
    r_chart=r_mfm+2+12+2
    chart=LineChart(); chart.title="Monatsverlauf"; chart.width=20; chart.height=12
    for j,yr in enumerate(show_yrs):
        s=Series(Reference(ws,min_col=j+2,min_row=r_mfm+1,max_row=r_mfm+13),title=str(yr))
        chart.series.append(s)
    chart.set_categories(Reference(ws,min_col=1,min_row=r_mfm+2,max_row=r_mfm+13))
    ws.add_chart(chart,"A"+str(r_chart))

    for c,w in [(1,12)]+[(i,12) for i in range(2,10)]:
        ws.column_dimensions[gcl(c)].width=w


def write_hyrox(wb, hy_data):
    ws = wb.create_sheet("🏃 Hyrox")
    ws.sheet_view.showGridLines = False
    title_bar(ws,1,"Hyrox – Dashboard",bg=TEAL,cols=14)
    ws.merge_cells("A2:N2"); ws["A2"].fill=fill("06B6D4"); ws.row_dimensions[2].height=4

    weekly = hy_data.get("weekly", [])
    # Sortierung für die Hyrox-Wochen-Tabelle: nach KW (inkl. Jahr), damit KW-Nummern aus
    # verschiedenen Jahren nicht vertauscht werden.
    def _week_sort_key(w: dict):
        kw = w.get("kw", 0) or 0
        d_str = str(w.get("date", "")).strip()
        # Erwartet Format: 'dd.mm.yy' (aus WeekStart)
        try:
            d = datetime.strptime(d_str[:8] + "." + d_str[-2:], "%d.%m.%y")
            return (d.year, int(kw), d)
        except Exception:
            return (0, int(kw), d_str)

    weekly = sorted(weekly, key=_week_sort_key)
    total_tn=sum(w.get("tn",0) for w in weekly)
    total_kurse=sum(w.get("kurse",0) for w in weekly)
    avg_ausl=sum(w.get("auslastung",0) for w in weekly)/len(weekly) if weekly else 0
    neue_k=sum(w.get("neue_kunden",0) for w in weekly)
    cv=hy_data.get("conversion",{})

    hy_um_sub = "Rechnungen" if hy_data.get("revenue_source") == "invoices" else (
        "Pässe" if hy_data.get("revenue_source") == "passes" else "aktuell")
    hy_rx = float(hy_data.get("revenue_hyrox_courses", 0) or 0)
    hy_pt = float(hy_data.get("revenue_personal_training", 0) or 0)
    hy_tot = float(hy_data.get("total_revenue", 0) or 0)
    if hy_tot and (hy_rx + hy_pt) == 0:
        hy_rx = hy_tot
    kpis = [
        (1, "Umsatz Hyrox o. PT", f"CHF {hy_rx:,.0f}".replace(",", "'"), "Kurse/Pässe", TEAL),
        (3, "Umsatz PT", f"CHF {hy_pt:,.0f}".replace(",", "'"), "Personal Training", PURPLE),
        (5, "Umsatz gesamt", f"CHF {hy_tot:,.0f}".replace(",", "'"), hy_um_sub, TEAL),
        (7, "Teilnehmer", str(total_tn), "Schedule", TEAL),
        (9, "Kurse", str(total_kurse), "Schedule", TEAL),
        (11, "Ø Auslastung", f"{avg_ausl:.1%}", "Schedule", TEAL),
        (
            13,
            "Trial Conv.",
            f"{cv.get('rate',0):.1%}",
            f"{cv.get('converted',0)} von {cv.get('trial',0)} (Pässe)",
            TEAL,
        ),
    ]
    for col, lbl, val, sub, color in kpis:
        kpi_tile(ws, 4, col, lbl, val, sub, color)
    gap(ws,7,14)

    # Wochentabelle
    sec(ws,8,"Wöchentliche Übersicht",cols=14)
    hdr(ws,9,["Woche ab","KW","Kurse","TN","Plätze","Ausl.%","Ø TN","Neue K.",
              "1Karte","Trial","10er","30er","Unlim.","PT"],bg="0E4F5C")
    for i,w in enumerate(weekly,10):
        bg=ALT if i%2==0 else WHITE
        dc(ws,i,1,w.get("date","")[:10],bg,h="left"); dc(ws,i,2,w.get("kw"),bg,h="center")
        dc(ws,i,3,w.get("kurse"),bg,h="center"); dc(ws,i,4,w.get("tn"),bg,h="center")
        dc(ws,i,5,w.get("plaetze"),bg,h="center")
        au=w.get("auslastung",0)
        c_=ws.cell(i,6,au); c_.number_format="0.0%"; c_.border=bb(); c_.alignment=aln("center")
        c_.fill=fill("D1FAE5" if au>=0.65 else("FEF3C7" if au>=0.5 else"FEE2E2"))
        c_.font=Font(name="Arial",bold=True,sz=10,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED))
        dc(ws,i,7,round(w.get("avg_tn",0),2),bg,fmt="0.00",h="center")
        dc(ws,i,8,w.get("neue_kunden"),bg,h="center")
        for c_n,k in zip(range(9,15),["k1","trial","er10","er30","unlim","pt"]):
            dc(ws,i,c_n,w.get(k,0) or None,bg,h="center")
    r_ht=10+len(weekly)
    hdr(ws,r_ht,["TOTAL","",
                 f"=SUM(C10:C{r_ht-1})",f"=SUM(D10:D{r_ht-1})","",
                 f"=AVERAGE(F10:F{r_ht-1})",f"=AVERAGE(G10:G{r_ht-1})",
                 f"=SUM(H10:H{r_ht-1})"]+[f"=SUM({gcl(c)}10:{gcl(c)}{r_ht-1})" for c in range(9,15)],bg="0E4F5C")
    ws.cell(r_ht,6).number_format="0.0%"
    gap(ws,r_ht+1,14)

    # Best / Worst Kurse
    r_bw=r_ht+2
    sec(ws,r_bw,"TOP 3 BESTE KURSE (höchste Auslastung)",cols=14,bg="D1FAE5")
    hdr(ws,r_bw+1,["Datum","Uhrzeit","Kurstyp","Standort","Anm.","Plätze","Auslastung","Wochentag"],bg="14532D")
    for i,cls in enumerate(hy_data.get("best_classes",[]),r_bw+2):
        bg="F0FDF4"
        dc(ws,i,1,str(cls.get("Datum",""))[:10],bg,bold=True,h="left")
        dc(ws,i,2,cls.get("Beginnt um",""),bg,h="center")
        dc(ws,i,3,cls.get("Stunde",""),bg,h="left")
        dc(ws,i,4,cls.get("Ort/Raum",""),bg,h="left")
        dc(ws,i,5,cls.get("Anmeldungen"),bg,h="center"); dc(ws,i,6,cls.get("Plätze"),bg,h="center")
        c_=ws.cell(i,7,cls.get("Auslastung",0)); c_.number_format="0.0%"; c_.fill=fill("D1FAE5")
        c_.font=Font(name="Arial",bold=True,sz=10,color=GREEN); c_.border=bb(); c_.alignment=aln("center")
        dc(ws,i,8,cls.get("Wochentag",""),bg,h="center")

    r_ww=r_bw+6
    sec(ws,r_ww,"FLOP 3 SCHLECHTESTE KURSE (tiefste Auslastung, mind. 1 TN)",cols=14,bg="FEE2E2")
    hdr(ws,r_ww+1,["Datum","Uhrzeit","Kurstyp","Standort","Anm.","Plätze","Auslastung","Wochentag","Hinweis"],bg="7F1D1D")
    hints={"06:30":"Früh-Slot","06:00":"Früh-Slot"}
    for i,cls in enumerate(hy_data.get("worst_classes",[]),r_ww+2):
        bg="FFF1F1"
        dc(ws,i,1,str(cls.get("Datum",""))[:10],bg,bold=True,h="left")
        dc(ws,i,2,cls.get("Beginnt um",""),bg,h="center")
        dc(ws,i,3,cls.get("Stunde",""),bg,h="left")
        dc(ws,i,4,cls.get("Ort/Raum",""),bg,h="left")
        dc(ws,i,5,cls.get("Anmeldungen"),bg,h="center"); dc(ws,i,6,cls.get("Plätze"),bg,h="center")
        c_=ws.cell(i,7,cls.get("Auslastung",0)); c_.number_format="0.0%"; c_.fill=fill("FEE2E2")
        c_.font=Font(name="Arial",bold=True,sz=10,color=RED); c_.border=bb(); c_.alignment=aln("center")
        dc(ws,i,8,cls.get("Wochentag",""),bg,h="center")
        hint=hints.get(cls.get("Beginnt um",""),"Zu wenig Nachfrage")
        dc(ws,i,9,hint,bg,sz=9,color=ORANGE,h="left")

    r_an=r_ww+6
    sec(ws,r_an,"AUSLASTUNGS-ANALYSE",cols=14)
    hdr(ws,r_an+1,["Kategorie","Anz. Kurse","Ø TN/Kurs","Ø Auslastung","Empfehlung"])
    row_cur=r_an+2
    for label, data_dict in [("nach Kurstyp", hy_data.get("by_kurstyp",{})),
                              ("nach Uhrzeit", hy_data.get("by_time",{})),
                              ("nach Wochentag", hy_data.get("by_day",{}))]:
        ws.cell(row_cur,1,f"— {label} —").font=fnt(bold=True,sz=9,color=TEAL)
        ws.cell(row_cur,1).fill=fill("E0F7FA"); ws.row_dimensions[row_cur].height=14
        row_cur+=1
        for key, vals in data_dict.items():
            bg=ALT if row_cur%2==0 else WHITE
            au=vals.get("avg_ausl",0)
            dc(ws,row_cur,1,str(key),bg,bold=True,h="left")
            dc(ws,row_cur,2,vals.get("n",0),bg,h="center")
            dc(ws,row_cur,3,round(vals.get("avg_tn",0),1),bg,fmt="0.0")
            c_=ws.cell(row_cur,4,au); c_.number_format="0.0%"; c_.fill=fill("D1FAE5" if au>=0.65 else("FEF3C7" if au>=0.5 else"FEE2E2"))
            c_.font=Font(name="Arial",bold=True,sz=10,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED))
            c_.border=bb(); c_.alignment=aln("center")
            rec="⭐ Top-Slot" if au>=0.65 else("Gut" if au>=0.5 else"⚠️ Prüfen")
            dc(ws,row_cur,5,rec,bg,sz=9,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED),h="left")
            row_cur+=1

    for c,w in [(1,14),(2,7),(3,8),(4,10),(5,8),(6,8),(7,9),(8,12),(9,14),(10,8),(11,8),(12,8),(13,8),(14,8)]:
        ws.column_dimensions[gcl(c)].width=w

    # ── Nicht-wiederkehrende Kunden ───────────────────────────────────
    # Erste freie Zeile bestimmen
    r_nr_start = ws.max_row + 2
    gap(ws, r_nr_start - 1, 14)
    nr_list      = hy_data.get("non_returning", [])
    nr_type_list = hy_data.get("non_returning_by_type", [])

    sec(ws, r_nr_start,
        f"NICHT-WIEDERKEHRENDE KUNDEN – Abgelaufenes Abo, kein Folge-Kauf · Total: {len(nr_list)} Kunden",
        cols=14, bg="FEE2E2")

    # Zusammenfassung nach Typ
    hdr(ws, r_nr_start+1, ["Abo-Typ","Anzahl Kunden","Umsatz letztes Abo (CHF)","Empfehlung"], bg="7F1D1D")
    for idx, row_d in enumerate(nr_type_list, r_nr_start+2):
        bg = ALT if idx % 2 == 0 else WHITE
        dc(ws, idx, 1, row_d.get("Abonnement",""), bg, bold=True, h="left")
        dc(ws, idx, 2, row_d.get("Anzahl", 0), bg, h="center")
        dc(ws, idx, 3, row_d.get("Umsatz", 0), bg, fmt="#'##0.00")
        rec = "Re-Aktivierung per E-Mail" if "Trial" in str(row_d.get("Abonnement","")) else "Angebot senden"
        dc(ws, idx, 4, rec, bg, h="left", sz=9, color=ORANGE)

    r_detail_start = r_nr_start + 2 + len(nr_type_list) + 1
    gap(ws, r_detail_start - 1, 14)
    sec(ws, r_detail_start, "DETAIL – Retargeting-Liste (Name, E-Mail, letztes Abo)", cols=14, bg="FEF2F2")
    hdr(ws, r_detail_start+1,
        ["Vorname","Name","E-Mail","Letztes Abo","Gültig bis","Betrag (CHF)"], bg="7F1D1D")
    for idx, row_d in enumerate(nr_list, r_detail_start+2):
        bg = ALT if idx % 2 == 0 else WHITE
        dc(ws, idx, 1, row_d.get("Vorname",""), bg, h="left")
        dc(ws, idx, 2, row_d.get("Name",""), bg, h="left")
        dc(ws, idx, 3, row_d.get("E-Mail",""), bg, h="left", sz=9)
        dc(ws, idx, 4, row_d.get("Abonnement",""), bg, h="left", sz=9)
        gueltig = row_d.get("Gültig bis")
        dc(ws, idx, 5, str(gueltig)[:10] if gueltig else "–", bg, h="center")
        dc(ws, idx, 6, row_d.get("Total", 0), bg, fmt="#'##0.00")

    r_total = r_detail_start + 2 + len(nr_list)
    hdr(ws, r_total,
        ["TOTAL", "", f"{len(nr_list)} Kunden reaktivieren", "", "",
         f"=SUM(F{r_detail_start+2}:F{r_total-1})"], bg="7F1D1D")
    ws.cell(r_total, 6).number_format = "#'##0.00"


def write_simple_sheet(wb, title, bg_color, data_rows, col_headers, col_widths=None, cols=12):
    """Generischer Sheet-Writer für einfache Datentabellen."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    title_bar(ws,1,title.replace("💳 ","").replace("🏋 ","").replace("⚖️ ","").replace("📋 ",""),
              bg=bg_color,cols=cols)
    ws.merge_cells(f"A2:{gcl(cols)}2"); ws["A2"].fill=fill(bg_color); ws.row_dimensions[2].height=4
    hdr(ws,3,col_headers)
    for i,row in enumerate(data_rows,4):
        bg=ALT if i%2==0 else WHITE
        for j,val in enumerate(row,1):
            dc(ws,i,j,val,bg,h="left" if j==1 else "right")
    if col_widths:
        for c,w in enumerate(col_widths,1):
            ws.column_dimensions[gcl(c)].width=w
    return ws


def _fmt_source(path: Path | None) -> str:
    """Formatiert eine Datei als 'name (dd.mm.yyyy hh:mm)'."""
    if not path:
        return "nicht gefunden"
    try:
        ts = datetime.fromtimestamp(path.stat().st_mtime).strftime("%d.%m.%Y %H:%M")
        return f"{path.name} ({ts})"
    except Exception:
        return path.name


def _fmt_source_compact(path: Path | None) -> str:
    """Kompakte Version: nur Datum der Datei, ohne langen Dateinamen."""
    if not path:
        return "–"
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%d.%m.%Y")
    except Exception:
        return path.name


def add_data_basis(ws, cols: int, lines: list[str], bg: str = "E2E8F0"):
    """Schreibt die Datengrundlage in Zeile 2 eines Sheets (mit mehrzeiligem Umbruch)."""
    txt = "Datengrundlage: " + (" | ".join(lines) if lines else "keine Quelldatei")
    # Bestehende Merges in Zeile 2 entfernen, damit der neue Merge nicht kollidiert
    # (write_* setzt oft schon einen schmalen Separator A2:<n_hdr>2 — Excel repariert
    # sonst den Datei und verwirft die Merges beim Öffnen).
    existing = [r for r in list(ws.merged_cells.ranges) if r.min_row <= 2 <= r.max_row]
    for r in existing:
        ws.unmerge_cells(str(r))
    ws.merge_cells(f"A2:{gcl(cols)}2")
    c = ws.cell(2, 1, txt)
    c.fill = fill(bg)
    c.font = Font(name="Arial", size=9, bold=False, color="334155")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = bb()
    # Höhe an Textlänge anpassen: kurze Infos 22, lange bis 44
    approx_chars_per_line = max(cols * 12, 80)
    n_lines = max(1, min(3, -(-len(txt) // approx_chars_per_line)))
    ws.row_dimensions[2].height = 18 + 12 * n_lines


def add_all_sheets_data_basis(
    wb,
    files_found: dict,
    hy_sched: Path | None,
    hy_dashboard: Path | None,
    hy_active: Path | None,
    hy_past: Path | None,
    hy_invoices: list[Path],
    intrum_files: list[Path] | None = None,
    inkasso_count: int = 0,
):
    """Ergänzt Datengrundlage-Infos auf allen Dashboard-Sheets."""
    newest_inv = hy_invoices[0] if hy_invoices else None
    inv_count = len(hy_invoices)
    intrum_sorted = sorted(
        intrum_files or [],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if intrum_sorted:
        intrum_basis = (
            f"Intrum PDFs ({len(intrum_sorted)}): "
            + " | ".join(_fmt_source(p) for p in intrum_sorted)
        )
    else:
        intrum_basis = f"Intrum PDF: {_fmt_source(files_found.get('intrum'))}"

    sheet_sources = {
        "📊 Übersicht": (12, [
            # Kompakte Übersicht: nur Stand-Datum je Quelle (Details auf den jeweiligen Sheets)
            f"Umsatzanalyse {_fmt_source_compact(files_found.get('umsatzanalyse'))}",
            f"Runnerslab {_fmt_source_compact(files_found.get('runnerslab'))}",
            f"Hyrox {_fmt_source_compact(hy_sched)}",
            f"Hyrox Invoices ({inv_count}×) {_fmt_source_compact(newest_inv)}" if inv_count else "Hyrox Invoices: –",
            f"Ausstehend {_fmt_source_compact(files_found.get('ausstehend'))}",
            f"Fitness {_fmt_source_compact(files_found.get('fitness'))}",
            f"Intrum ({len(intrum_sorted)}×)" if intrum_sorted else "Intrum: –",
        ]),
        "👥 Therapeuten": (18, [
            f"Umsatzanalyse: {_fmt_source(files_found.get('umsatzanalyse'))}",
        ]),
        "👟 Runnerslab": (9, [
            f"Runnerslab: {_fmt_source(files_found.get('runnerslab'))}",
        ]),
        "🏃 Hyrox": (14, [
            f"Hyrox Schedules: {_fmt_source(hy_sched)}",
            f"Hyrox Dashboard historisch: {_fmt_source(hy_dashboard)}",
            f"Hyrox Active Passes: {_fmt_source(hy_active)}",
            f"Hyrox Past Passes: {_fmt_source(hy_past)}",
            f"Hyrox Invoices (neueste von {inv_count}): {_fmt_source(newest_inv)}" if inv_count else "Hyrox Invoices: nicht gefunden",
        ]),
        "🏋 Fitness-Abos": (8, [
            f"Fitness: {_fmt_source(files_found.get('fitness'))}",
        ]),
        "💳 Ausstehende Zhlg.": (12, [
            f"Ausstehende Zahlungen: {_fmt_source(files_found.get('ausstehend'))}",
        ]),
        "⚖️ Inkasso": (10, [
            intrum_basis,
            (
                "pdfplumber fehlt — pip install pdfplumber (Fälle nicht extrahierbar)"
                if pdfplumber is None
                else (
                    f"{inkasso_count} Fall/Fälle aus PDF (Duplikate nach Inkasso-Nr., neuere Datei gewinnt)"
                    if (intrum_files or [])
                    else "Keine Intrum-PDF in input/"
                )
            ),
        ]),
        "📋 Berichte Ärzte": (18, [
            f"Berichte Ärzte: {_fmt_source(files_found.get('berichte'))}",
        ]),
    }

    for title, (cols, lines) in sheet_sources.items():
        if title in wb.sheetnames:
            add_data_basis(wb[title], cols, lines)
            if title == "📊 Übersicht" and intrum_sorted and len(intrum_basis) > 200:
                wb[title].row_dimensions[2].height = 36


# ═══════════════════════════════════════════════════════════════════════
# HAUPTPROGRAMM
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("Kineo AG Dashboard Updater")
    print("=" * 60)

    ensure_input_output_layout()

    today_str = datetime.today().strftime("%d.%m.%Y")

    # Dateien finden
    print("\n[1/3] Suche Quelldateien in input/...")
    files_found = {}
    files_missing = []
    for key in FILE_PATTERNS:
        f = find_file(key)
        if f:
            files_found[key] = f
            print(f"  ✓ {key}: {f.name}")
        else:
            files_missing.append(key)
            print(f"  ✗ {key}: nicht gefunden (Muster: {FILE_PATTERNS[key]})")

    if not files_found.get("umsatzanalyse"):
        print("\n⚠️  Umsatzanalyse nicht gefunden — ohne diese kann kein Dashboard erstellt werden.")
        print("  Bitte Datei in den input/-Ordner legen.")
        return

    # Daten laden
    print("\n[2/3] Lade Daten...")
    ua_data = load_umsatzanalyse(files_found["umsatzanalyse"])
    print(f"  → Monate mit Daten: {ua_data['months_available']}")

    rl_data = {}
    if files_found.get("runnerslab"):
        rl_data = load_runnerslab(files_found["runnerslab"])
        print(f"  → Runnerslab-Jahre: {sorted(rl_data.keys())}")

    hy_data = {}
    hy_sched     = find_file("hyrox_schedule")
    hy_dashboard = find_file("hyrox_dashboard")
    hy_active    = find_file("hyrox_active")
    hy_past      = find_file("hyrox_past")
    hy_invoices  = find_all_files("hyrox_invoices")
    intrum_files = find_all_files("intrum")
    if hy_sched or hy_active or hy_past or hy_dashboard or hy_invoices:
        hy_data = load_hyrox(hy_sched, hy_active, hy_past, hy_dashboard)
        print(f"  → Hyrox-Wochen: {len(hy_data.get('weekly', []))}")
    else:
        print("  ✗ hyrox: keine SportsNow-Exporte gefunden")
        print("    Erwartet: schedules, active/past passes, optional *invoices*.xlsx")

    inv_data = []
    if files_found.get("ausstehend"):
        inv_data = load_ausstehend(files_found["ausstehend"])
        print(f"  → Ausstehende Rechnungen: {len(inv_data)}")

    fit_data = []
    if files_found.get("fitness"):
        fit_data = load_fitness(files_found["fitness"])
        print(f"  → Fitness-Einträge: {len(fit_data)}")

    ber_data = {}
    if files_found.get("berichte"):
        ber_data = load_berichte(files_found["berichte"])
        print(f"  → Berichte Wochen: {len(ber_data.get('data',{}))}")

    inkasso_cases = load_intrum_cases(intrum_files)
    if pdfplumber is None:
        print("  ⚠ Intrum/Inkasso: pdfplumber nicht installiert — Sheet ohne PDF-Daten")
    elif intrum_files:
        print(f"  → Intrum Inkasso-Fälle (nach Dedupe): {len(inkasso_cases)}")

    # Dashboard schreiben
    print("\n[3/3] Schreibe Dashboard...")
    wb = Workbook()
    wb.active.title = "_temp"

    write_uebersicht(wb, ua_data, rl_data, hy_data, inv_data, fit_data, today_str)
    write_therapeuten(wb, ua_data, today_str)
    write_runnerslab(wb, rl_data)

    if hy_data:
        write_hyrox(wb, hy_data)
    else:
        ws = wb.create_sheet("🏃 Hyrox")
        ws.cell(1,1,"Hyrox-Datei nicht gefunden — bitte in input/ ablegen")

    # Fitness
    ws_fit = wb.create_sheet("🏋 Fitness-Abos")
    ws_fit.sheet_view.showGridLines = False
    title_bar(ws_fit,1,"Fitness-Abos Thalwil",bg=GREEN,cols=8)
    ws_fit.merge_cells("A2:H2"); ws_fit["A2"].fill=fill("22C55E"); ws_fit.row_dimensions[2].height=4
    hdr(ws_fit,3,["Periode","2-Jahres","1-Jahr","6-Monate","MTT Plus","Total","Neue","Kündig."],bg="14532D")
    for i,row in enumerate(fit_data,4):
        bg=ALT if i%2==0 else WHITE
        kw=row.get("kw","")
        y2=row.get("Anzahl aktive Verträge 2 Jahr") or row.get("Anzahl aktive Vertr\u00e4ge 2 Jahr")
        y1=row.get("Anzahl aktive Verträge 1 Jahr") or row.get("Anzahl aktive Vertr\u00e4ge 1 Jahr")
        m6=row.get("Anzahl aktive Verträge  6 Monate") or row.get("Anzahl aktive Vertr\u00e4ge  6 Monate")
        mtt=row.get("Anzahl aktive Verträge MTT Plus") or row.get("Anzahl aktive Vertr\u00e4ge MTT Plus")
        tot=row.get("Mitglieder Gesamt")
        neu=row.get("Anzahl neue Verträge") or row.get("Anzahl neue Vertr\u00e4ge")
        kue=row.get("Anzahl Kündigungen")
        dc(ws_fit,i,1,kw,bg,bold=True,h="left")
        for c_,v in zip([2,3,4,5,6,7,8],[y2,y1,m6,mtt,tot,neu,kue]):
            if v is not None and str(v) not in ["nan","None",""]:
                try: dc(ws_fit,i,c_,int(v),bg,h="center")
                except: dc(ws_fit,i,c_,"–",bg,h="center")
            else: dc(ws_fit,i,c_,"–",bg,h="center")
    for c,w in [(1,14)]+[(i,10) for i in range(2,9)]:
        ws_fit.column_dimensions[gcl(c)].width=w

    # Ausstehende Zahlungen
    ws_az = wb.create_sheet("💳 Ausstehende Zhlg.")
    ws_az.sheet_view.showGridLines = False
    title_bar(ws_az,1,f"Ausstehende Zahlungen · {len(inv_data)} Rechnungen · Total CHF {sum(i['betrag'] for i in inv_data):,.0f}".replace(",","'"),
              bg=RED,cols=12)
    ws_az.merge_cells("A2:L2"); ws_az["A2"].fill=fill("EF4444"); ws_az.row_dimensions[2].height=4
    hdr(ws_az,3,["Re.-Nr.","Kunde","Therapeut","Typ","Rechnung an","Re.-Datum","Letzte Mahn.","Mahnstufe","Tage überf.","Standort","Betrag (CHF)","Inkasso?"],bg="7F1D1D")
    for i,inv in enumerate(inv_data,4):
        d=inv.get("tage",0)
        rb=fill("FEE2E2" if d>365 else("FED7AA" if d>180 else("FEF3C7" if d>90 else(ALT if i%2==0 else WHITE))))
        vals=[inv.get("renr"),inv.get("kunde"),inv.get("therapeut"),inv.get("an"),
              inv.get("an_wen"),inv.get("re_datum"),inv.get("letzte_m"),inv.get("mahnstufe"),
              inv.get("tage"),inv.get("standort"),inv.get("betrag"),""]
        for j,v in enumerate(vals,1):
            ce=ws_az.cell(i,j,v); ce.fill=rb; ce.font=fnt(sz=9,bold=(j==11)); ce.border=bb()
            ce.alignment=aln("right" if j in[1,8,9,10,11] else"left")
        ws_az.cell(i,11).number_format="#'##0.00"
    for c,w in [(1,8),(2,26),(3,13),(4,13),(5,26),(6,12),(7,13),(8,8),(9,10),(10,8),(11,13),(12,9)]:
        ws_az.column_dimensions[gcl(c)].width=w

    # Inkasso Intrum (Zeilen aus Intrum-Auftragsbestätigung-PDF)
    ws_ink = wb.create_sheet("⚖️ Inkasso")
    ws_ink.sheet_view.showGridLines = False
    n_ink = len(inkasso_cases)
    title_bar(
        ws_ink,
        1,
        f"Inkasso Intrum – {n_ink} Fall/Fälle (aus PDF)" if n_ink else "Inkasso Intrum (aus PDF)",
        bg=PURPLE,
        cols=10,
    )
    ws_ink.merge_cells("A2:J2"); ws_ink["A2"].fill=fill("8B5CF6"); ws_ink.row_dimensions[2].height=4
    hdr(ws_ink,3,["Inkasso-Nr.","Schuldner","Adresse","Score","Adress-Status","Ref.-Nr.","Datum","Kapital (CHF)","Fallstatus","Erfolgsquote"],bg="4C1D95")
    r_data = 4
    if pdfplumber is None:
        ws_ink.merge_cells(f"A{r_data}:J{r_data}")
        c = ws_ink.cell(r_data, 1, "pdfplumber fehlt — bitte installieren: pip install pdfplumber")
        c.font = fnt(sz=9, color="64748B", italic=True)
        c.fill = fill(HDRBG)
        c.alignment = aln("left", "center")
        ws_ink.row_dimensions[r_data].height = 18
        r_data += 1
    elif not intrum_files:
        ws_ink.merge_cells(f"A{r_data}:J{r_data}")
        c = ws_ink.cell(r_data, 1, "Keine Intrum-PDF in input/ — Datei mit Muster *ntrum*.pdf ablegen.")
        c.font = fnt(sz=9, color="64748B", italic=True)
        c.fill = fill(HDRBG)
        c.alignment = aln("left", "center")
        ws_ink.row_dimensions[r_data].height = 18
        r_data += 1
    elif not inkasso_cases:
        ws_ink.merge_cells(f"A{r_data}:J{r_data}")
        c = ws_ink.cell(
            r_data,
            1,
            "Intrum-PDF vorhanden, aber keine Fälle erkannt — Textlayout ggf. abweichend; Datei prüfen.",
        )
        c.font = fnt(sz=9, color="64748B", italic=True)
        c.fill = fill(HDRBG)
        c.alignment = aln("left", "center")
        ws_ink.row_dimensions[r_data].height = 22
        r_data += 1
    else:
        for i, cse in enumerate(inkasso_cases, start=r_data):
            bg = ALT if i % 2 == 0 else WHITE
            vals = [
                cse.get("inkasso_nr"),
                cse.get("schuldner"),
                cse.get("adresse"),
                cse.get("score"),
                cse.get("adress_status"),
                cse.get("ref_nr"),
                cse.get("datum"),
                cse.get("kapital"),
                cse.get("fallstatus"),
                cse.get("erfolgsquote"),
            ]
            for j, v in enumerate(vals, 1):
                h_align = "left" if j in (2, 3, 9, 10) else ("center" if j in (4, 5, 7) else "right")
                ce = ws_ink.cell(i, j, v if v is not None else "–")
                ce.fill = fill(bg)
                ce.font = fnt(sz=9, bold=(j == 1))
                ce.border = bb()
                ce.alignment = aln(h_align)
            ws_ink.cell(i, 8).number_format = "#'##0.00"
        r_data = 3 + len(inkasso_cases) + 1
    r_leg = r_data + 1
    sec(ws_ink, r_leg, "Score-Legende Intrum", cols=10)
    hdr(ws_ink, r_leg + 1, ["Score", "Erfolgsquote", "Bedeutung"], bg="4C1D95")
    score_colors = ["DCFCE7", "D1FAE5", "FEF3C7", "FED7AA", "FEE2E2"]
    for idx, (sc_, eq, desc) in enumerate(
        [
            ("A", "75–100%", "Bestätigte Adresse, Zahlungserfahrung"),
            ("B", "60–75%", "Bestätigte Adresse, kürzere Adresshistorie"),
            ("C", "25–60%", "Keine bestätigte Adresse"),
            ("D", "10–25%", "Offene Forderungen / Betreibungen"),
            ("E", "0–10%", "Hohe Forderungen / Pfändungen / Verlustscheine"),
        ]
    ):
        i = r_leg + 2 + idx
        bg = ALT if idx % 2 == 0 else WHITE
        fc = score_colors[idx]
        dc(ws_ink, i, 1, sc_, bg)
        ws_ink.cell(i, 1).fill = fill(fc)
        ws_ink.cell(i, 1).font = fnt(bold=True, sz=10)
        dc(ws_ink, i, 2, eq, bg, h="center")
        dc(ws_ink, i, 3, desc, bg, h="left")
    for c, w in [(1, 12), (2, 14), (3, 40), (4, 8), (5, 18), (6, 12), (7, 12), (8, 14), (9, 14), (10, 14)]:
        ws_ink.column_dimensions[gcl(c)].width = w

    # Berichte an Ärzte
    ws_ba = wb.create_sheet("📋 Berichte Ärzte")
    ws_ba.sheet_view.showGridLines = False
    title_bar(ws_ba,1,"Berichte an Ärzte",bg=TEAL,cols=18)
    therapists=ber_data.get("therapists",[])
    data=ber_data.get("data",{})
    if therapists:
        hdr(ws_ba,2,["KW"]+therapists+["Total"],bg="164E63")
        for i,kw in enumerate(sorted(data.keys()),3):
            bg=ALT if i%2==0 else WHITE
            ws_ba.cell(i,1,kw).fill=fill(bg); ws_ba.cell(i,1).font=fnt(bold=True,sz=10); ws_ba.cell(i,1).border=bb()
            vals=data[kw]
            for j,t in enumerate(therapists,2):
                v=vals.get(t,0)
                ce=ws_ba.cell(i,j,v if v else "")
                ce.fill=fill("BAE6FD" if v else bg); ce.font=fnt(sz=10,bold=bool(v)); ce.border=bb(); ce.alignment=aln("center")
            ws_ba.cell(i,len(therapists)+2,f"=SUM(B{i}:{gcl(len(therapists)+1)}{i})").fill=fill(bg)
            ws_ba.cell(i,len(therapists)+2).font=fnt(bold=True,sz=10); ws_ba.cell(i,len(therapists)+2).border=bb(); ws_ba.cell(i,len(therapists)+2).alignment=aln("center")
        ws_ba.column_dimensions["A"].width=8
        for c in range(2,len(therapists)+3): ws_ba.column_dimensions[gcl(c)].width=9

    # Leeres Sheet entfernen, Reihenfolge korrigieren
    if "_temp" in wb.sheetnames:
        del wb["_temp"]
    desired = ["📊 Übersicht","👥 Therapeuten","👟 Runnerslab","🏃 Hyrox",
               "🏋 Fitness-Abos","💳 Ausstehende Zhlg.","⚖️ Inkasso","📋 Berichte Ärzte"]
    sheet_map = {ws.title: ws for ws in wb.worksheets}
    ordered = [sheet_map[n] for n in desired if n in sheet_map]
    remaining = [ws for ws in wb.worksheets if ws not in ordered]
    wb._sheets = ordered + remaining

    # Datengrundlage pro Sheet ergänzen (welche Datei wurde verwendet)
    add_all_sheets_data_basis(
        wb,
        files_found=files_found,
        hy_sched=hy_sched,
        hy_dashboard=hy_dashboard,
        hy_active=hy_active,
        hy_past=hy_past,
        hy_invoices=hy_invoices,
        intrum_files=intrum_files,
        inkasso_count=len(inkasso_cases),
    )

    # Speichern: zeitgestempelt unter output/archive/, Kopie als AKTUELL unter output/
    ts = datetime.today().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_ARCHIVE_DIR / f"Kineo_Dashboard_{ts}.xlsx"
    fixed_path = OUTPUT_DIR / "Kineo_Dashboard_AKTUELL.xlsx"
    wb.save(out_path)
    shutil.copy2(out_path, fixed_path)

    print(f"\n✅ Dashboard gespeichert:")
    print(f"   {out_path}")
    print(f"   {fixed_path}  (Kopie der letzten Version)")
    if files_missing:
        non_hyrox = [f for f in files_missing if "hyrox" not in f]
        if non_hyrox:
            print(f"\n⚠️  Fehlende Dateien (nicht kritisch): {', '.join(non_hyrox)}")
    print("\nFertig!")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Fehler: {e}")
        traceback.print_exc()
        input("\nEnter drücken zum Beenden...")
