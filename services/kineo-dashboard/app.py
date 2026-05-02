"""
Kineo Operations Dashboard — main entrypoint.

Default view: renders the KPIs and per-sheet tables from
output/Kineo_Dashboard_AKTUELL.xlsx so the user lands on the live
operational numbers (not the upload form).

Settings drawer (gear icon in sidebar): exposes the file-upload
form (Hyrox / SportsNow exports) and the "regenerate dashboard" action,
which re-runs update_dashboard.main() to rebuild the xlsx in-place.
"""

from __future__ import annotations

import re
import subprocess
import sys
import time
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from update_dashboard import (
    INPUT_HYROX_INVOICES,
    INPUT_HYROX_SCHEDULES,
    ensure_input_output_layout,
)

ROOT = Path(__file__).parent
DASHBOARD_FILE = ROOT / "output" / "Kineo_Dashboard_AKTUELL.xlsx"

# The "Übersicht" sheet stores six headline KPIs in row 4 (labels) and row 5
# (values). Row 6 is a sub-caption ("Jan – Mär", "Letzter Monat", …).
KPI_LAYOUT = [
    ("A5", "A4", "A6"),
    ("B5", "B4", "B6"),
    ("C5", "C4", "C6"),
    ("D5", "D4", "D6"),
    ("E5", "E4", "E6"),
    ("F5", "F4", "F6"),
]


def _file_age(path: Path) -> str:
    if not path.exists():
        return "—"
    delta = time.time() - path.stat().st_mtime
    if delta < 90:
        return f"{int(delta)} s"
    if delta < 5400:
        return f"{int(delta / 60)} min"
    if delta < 86_400:
        return f"{int(delta / 3600)} h"
    return f"{int(delta / 86_400)} T"


def _sanitize(s: str) -> str:
    s = s.strip().replace(" ", "_")
    return re.sub(r"[^\w.\-]+", "_", s)[:80] or "export"


def _month_options() -> list[str]:
    today = date.today()
    y, m = today.year, today.month
    out: list[str] = []
    for _ in range(36):
        out.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m < 1:
            m = 12
            y -= 1
    return out


KINDS = {
    "Kursplan (schedules)": (
        "schedules",
        lambda ym, extra: f"kineo-hyrox_schedules_{ym}{extra}.xlsx",
    ),
    "Aktive Abos (active passes)": (
        "schedules",
        lambda ym, extra: f"kineo-hyrox_active_passes_{ym}{extra}.xlsx",
    ),
    "Abgelaufene Abos (past passes)": (
        "schedules",
        lambda ym, extra: f"kineo-hyrox_past_passes_{ym}{extra}.xlsx",
    ),
    "Rechnungen (invoices)": (
        "invoices",
        lambda ym, extra: f"kineo-hyrox_invoices_{ym}{extra}.xlsx",
    ),
}


@st.cache_data(ttl=60, show_spinner=False)
def _load_workbook(path_str: str, mtime: float) -> dict:
    """
    Reads the dashboard xlsx into:
      - title / subtitle (rows 1–2 of "Übersicht")
      - headline KPIs (six tiles from row 4–5)
      - per-sheet dataframes (header=None to keep the report's free-form layout)

    Cache-key includes mtime so the next regenerate invalidates automatically.
    """
    _ = mtime  # cache-key only; not used directly
    wb = load_workbook(path_str, data_only=True)
    out: dict = {"sheets": {}, "kpis": [], "title": "", "subtitle": "", "stand": ""}

    if "📊 Übersicht" in wb.sheetnames:
        ws = wb["📊 Übersicht"]
        out["title"] = ws["A1"].value or ""
        out["subtitle"] = ws["A2"].value or ""
        m = re.search(r"Stand\s+([\d\.]+)", out["title"])
        if m:
            out["stand"] = m.group(1)
        for value_ref, label_ref, sub_ref in KPI_LAYOUT:
            value = ws[value_ref].value
            if value is None:
                continue
            out["kpis"].append({
                "label": str(ws[label_ref].value or "").strip(),
                "value": str(value).strip(),
                "sub": str(ws[sub_ref].value or "").strip(),
            })

    for sn in wb.sheetnames:
        df = pd.read_excel(path_str, sheet_name=sn, header=None, engine="openpyxl")
        df = df.dropna(how="all").reset_index(drop=True)
        out["sheets"][sn] = df

    return out


def _render_dashboard() -> None:
    if not DASHBOARD_FILE.exists():
        st.warning(
            "Noch keine `Kineo_Dashboard_AKTUELL.xlsx` vorhanden. "
            "Über das Zahnrad-Symbol in der Seitenleiste kannst du Hyrox-Exporte "
            "hochladen und das Dashboard generieren."
        )
        return

    data = _load_workbook(str(DASHBOARD_FILE), DASHBOARD_FILE.stat().st_mtime)

    st.markdown(f"### {data['title']}")
    if data["subtitle"]:
        st.caption(data["subtitle"])

    if data["kpis"]:
        cols = st.columns(len(data["kpis"]))
        for col, kpi in zip(cols, data["kpis"]):
            col.metric(label=kpi["label"], value=kpi["value"], help=kpi["sub"] or None)

    st.divider()

    # Übersicht is already represented by the KPI tiles, so its tab shows the
    # detail rows below the header (everything from row 7 onwards).
    sheet_names = [s for s in data["sheets"].keys() if s != "📊 Übersicht"]
    tabs = st.tabs(["📊 Übersicht"] + sheet_names)

    with tabs[0]:
        df = data["sheets"]["📊 Übersicht"]
        body = df.iloc[6:].reset_index(drop=True)
        st.dataframe(body, hide_index=True, use_container_width=True, height=520)

    for tab, sn in zip(tabs[1:], sheet_names):
        with tab:
            df = data["sheets"][sn]
            st.dataframe(df, hide_index=True, use_container_width=True, height=520)


def _render_settings_drawer() -> None:
    """The gear-drawer in the sidebar: upload + regenerate."""
    ensure_input_output_layout()

    age = _file_age(DASHBOARD_FILE)
    if DASHBOARD_FILE.exists():
        ts = datetime.fromtimestamp(DASHBOARD_FILE.stat().st_mtime).strftime("%d.%m.%Y %H:%M")
        st.sidebar.caption(f"Letzter Build: **{ts}** ({age} alt)")
    else:
        st.sidebar.caption("Noch kein Dashboard generiert.")

    with st.sidebar.expander("Daten-Upload", expanded=False):
        kind = st.selectbox("Dateityp", list(KINDS.keys()))
        month = st.selectbox("Monat", _month_options())
        extra = st.text_input("Zusatz im Dateinamen (optional)", placeholder="z.B. v2")
        extra_part = f"_{_sanitize(extra)}" if extra.strip() else ""

        area, name_fn = KINDS[kind]
        if area == "schedules":
            dest_dir = INPUT_HYROX_SCHEDULES / month
        else:
            dest_dir = INPUT_HYROX_INVOICES / month
        dest_dir.mkdir(parents=True, exist_ok=True)

        target_name = name_fn(month, extra_part)
        st.caption(f"Ziel: `{dest_dir / target_name}`")

        uploaded = st.file_uploader("Excel-Datei (.xlsx)", type=["xlsx"], key="upl")
        if uploaded and st.button("Speichern", type="primary", use_container_width=True):
            path = dest_dir / target_name
            path.write_bytes(uploaded.getvalue())
            st.success(f"Gespeichert: `{path.name}`")
            st.caption("Anschliessend unten 'Dashboard neu generieren'.")

    with st.sidebar.expander("Dashboard neu generieren", expanded=False):
        st.caption(
            "Liest alle Quelldateien aus `input/` neu ein und schreibt "
            "`output/Kineo_Dashboard_AKTUELL.xlsx`. Dauert i.d.R. < 1 min."
        )
        if st.button("Jetzt generieren", use_container_width=True):
            placeholder = st.empty()
            placeholder.info("update_dashboard.py läuft …")
            try:
                proc = subprocess.run(
                    [sys.executable, "update_dashboard.py"],
                    cwd=str(ROOT),
                    capture_output=True,
                    text=True,
                    timeout=300,
                )
                if proc.returncode == 0:
                    placeholder.success("Dashboard neu generiert.")
                    _load_workbook.clear()
                    st.rerun()
                else:
                    placeholder.error(f"Fehler (rc={proc.returncode})")
                    st.code(proc.stderr[-2000:] or proc.stdout[-2000:])
            except subprocess.TimeoutExpired:
                placeholder.error("Timeout (>300s) — bitte Logs prüfen.")
            except Exception as exc:
                placeholder.error(f"Exception: {exc}")


def main() -> None:
    st.set_page_config(
        page_title="Kineo Operations Dashboard",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.sidebar.title("Einstellungen")
    _render_settings_drawer()
    _render_dashboard()


if __name__ == "__main__":
    main()
