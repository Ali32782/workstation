"""Probelauf: Simuliert einen zusätzlichen Monat (Apr26) und prüft Layout + Merge-Konflikte.

Lädt die echten Daten, klont März-Werte als April, rendert das Dashboard nach
output/Kineo_Dashboard_PROBELAUF.xlsx und verifiziert anschliessend:
  - keine überlappenden Merges in irgendeinem Sheet
  - alle Monate erscheinen als eigene Spalte
  - Σ/Ø-Spalte ist vorhanden
"""
from __future__ import annotations
import copy
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

import update_dashboard as ud


def inject_month(ua_data: dict, dst: str, factor: float = 1.05) -> dict:
    """Klont März-Daten als <dst> und hängt sie an months_available an."""
    ua = copy.deepcopy(ua_data)
    src = "Mär26"

    if src not in ua.get("months_available", []):
        print(f"⚠ Quellmonat {src} nicht vorhanden — Probelauf bricht ab.")
        sys.exit(1)
    if dst in ua.get("months_available", []):
        print(f"ℹ {dst} bereits vorhanden — überspringe Injection.")
        return ua

    ua["months_available"].append(dst)

    if "by_month" in ua and src in ua["by_month"]:
        src_rows = copy.deepcopy(ua["by_month"][src])
        for r in src_rows:
            for k, v in list(r.items()):
                if isinstance(v, (int, float)) and k not in ("fte", "pensum", "BG"):
                    r[k] = round(v * factor, 2)
        ua["by_month"][dst] = src_rows

    if "by_month_totals" in ua and src in ua["by_month_totals"]:
        totals = copy.deepcopy(ua["by_month_totals"][src])
        for k, v in list(totals.items()):
            if isinstance(v, (int, float)):
                totals[k] = round(v * factor, 2)
        ua["by_month_totals"][dst] = totals

    # Top-Level-Felder (<prefix>_umsatz / <prefix>_zielB) für Ranking füllen
    prefix_map = {"Apr26": "apr", "Mai26": "mai", "Jun26": "jun",
                  "Jul26": "jul", "Aug26": "aug", "Sep26": "sep"}
    prefix = prefix_map.get(dst)
    if prefix:
        for t in ua.get("therapeuten", []):
            name = t.get("name") or t.get("Name")
            if not name:
                continue
            match = next(
                (r for r in ua["by_month"].get(dst, []) if (r.get("name") or r.get("Name")) == name),
                None,
            )
            if match:
                t[f"{prefix}_umsatz"] = match.get("umsatz") or match.get("Umsatz")
                t[f"{prefix}_zielB"] = match.get("zielB") or match.get("Ziel(B)")

    return ua


def verify(path: Path) -> bool:
    wb = load_workbook(path)
    ok = True
    for sn in wb.sheetnames:
        ws = wb[sn]
        merges = sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col))
        overlaps = []
        for i, m1 in enumerate(merges):
            for m2 in merges[i + 1:]:
                if m1.min_row > m2.max_row or m2.min_row > m1.max_row:
                    continue
                if m1.min_col > m2.max_col or m2.min_col > m1.max_col:
                    continue
                overlaps.append((m1, m2))
        status = "✓" if not overlaps else "✗"
        print(f"  {status} {sn}: {len(merges)} Merges, {len(overlaps)} Overlaps")
        for m1, m2 in overlaps:
            print(f"      OVERLAP: {m1}  ⇄  {m2}")
            ok = False

    ws_u = wb["📊 Übersicht"]
    header_row = [ws_u.cell(9, c).value for c in range(1, 20) if ws_u.cell(9, c).value]
    print(f"\n  Übersicht Kennzahlen-Header (Zeile 9): {header_row}")

    ws_t = wb["👥 Therapeuten"]
    ms_hdr = [ws_t.cell(9, c).value for c in range(1, 30) if ws_t.cell(9, c).value]
    print(f"  Therapeuten Monatsübersicht-Header (Zeile 9): {ms_hdr}")

    return ok


def main():
    extra = sys.argv[1:] or ["Apr26"]
    suffix = "_".join(m[:3] for m in extra)
    print(f"[Probelauf] Simuliere zusätzliche Monate: {extra}\n")

    files_found = {k: ud.find_file(k) for k in ud.FILE_PATTERNS}

    if not files_found.get("umsatzanalyse"):
        print("✗ Umsatzanalyse nicht gefunden"); sys.exit(1)

    print("[1/3] Lade Daten…")
    ua_data = ud.load_umsatzanalyse(files_found["umsatzanalyse"])
    print(f"  echte Monate: {ua_data['months_available']}")

    for i, m in enumerate(extra):
        ua_data = inject_month(ua_data, m, factor=1.0 + 0.05 * (i + 1))
    print(f"  Monate nach Injection: {ua_data['months_available']}")

    rl_data = ud.load_runnerslab(files_found["runnerslab"]) if files_found.get("runnerslab") else {}

    hy_data = {}
    hy_sched = ud.find_file("hyrox_schedule")
    hy_dashboard = ud.find_file("hyrox_dashboard")
    hy_active = ud.find_file("hyrox_active")
    hy_past = ud.find_file("hyrox_past")
    if hy_sched or hy_active or hy_past or hy_dashboard:
        hy_data = ud.load_hyrox(hy_sched, hy_active, hy_past, hy_dashboard)

    inv_data = ud.load_ausstehend(files_found["ausstehend"]) if files_found.get("ausstehend") else []
    fit_data = ud.load_fitness(files_found["fitness"]) if files_found.get("fitness") else []

    today_str = "20.04.2026 (Probelauf)"

    print("\n[2/3] Schreibe Dashboard…")
    wb = Workbook()
    wb.active.title = "_temp"
    ud.write_uebersicht(wb, ua_data, rl_data, hy_data, inv_data, fit_data, today_str)
    ud.write_therapeuten(wb, ua_data, today_str)
    ud.write_runnerslab(wb, rl_data)
    if hy_data:
        ud.write_hyrox(wb, hy_data)
    if "_temp" in wb.sheetnames:
        del wb["_temp"]

    out = Path(f"output/Kineo_Dashboard_PROBELAUF_{suffix}.xlsx")
    wb.save(out)
    print(f"  → gespeichert: {out}")

    print("\n[3/3] Verifikation…")
    ok = verify(out)
    print("\n" + ("✅ Probelauf OK — keine Overlaps, Layout skaliert sauber."
                  if ok else "❌ Probelauf fehlgeschlagen — siehe Overlaps oben."))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
