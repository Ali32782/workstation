"""
Kineo AG – Dashboard Update Script
===================================
Verwendung:
  python update_dashboard.py

Das Skript liest alle Quelldateien aus dem Ordner 'input/',
erkennt automatisch die neuesten Versionen, und schreibt das
komplette Dashboard nach 'output/Kineo_Dashboard.xlsx'.

Voraussetzungen:
  pip install openpyxl pandas pdfplumber
"""

import os
import re
import glob
import sys
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, LineChart, Reference, Series

# ═══════════════════════════════════════════════════════════════════════
# KONFIGURATION — hier kannst du Dateimuster anpassen
# ═══════════════════════════════════════════════════════════════════════

INPUT_DIR  = Path("input")
OUTPUT_DIR = Path("output")

# Dateimuster: Schlüssel → Suchmuster (Glob) im input/-Ordner
# Das Skript nimmt bei mehreren Treffern immer die neueste Datei (nach Datum im Namen oder Änderungsdatum)
FILE_PATTERNS = {
    "umsatzanalyse":    "*msatzanalyse*2026*.xlsx",
    "hyrox_schedule":   "*yrox*schedules*.xlsx",          # SportsNow: Kursplan Export (aktuell)
    "hyrox_dashboard":  "*yrox*dashboard*.xlsx",          # Altes aufbereitetes Dashboard (historisch)
    "hyrox_active":     "*yrox*active*passes*.xlsx",      # SportsNow: Aktive Abonnemente
    "hyrox_past":       "*yrox*past*passes*.xlsx",        # SportsNow: Abgelaufene Abonnemente
    "runnerslab":       "*nnerslab*.xlsx",
    "ausstehend":       "*usstehend*Zahlung*.xlsx",
    "fitness":          "*itness*Thalwil*.xlsx",
    "berichte":         "*erichte*Aerzte*.xlsx",
    "intrum":           "*ntrum*.pdf",
}

# Hinweis SportsNow-Exports:
# Kursplan:   Mein Adminbereich → Stundenplan → "Exportieren" (ohne TN-Details)
# Abonnemente: Mein Adminbereich → Abonnemente → Aktive / Abgelaufene → "Exportieren"

# ═══════════════════════════════════════════════════════════════════════
# STIL-HELFER
# ═══════════════════════════════════════════════════════════════════════

DARK   = "1E293B"; BLUE  = "2563EB"; GREEN  = "16A34A"; RED   = "DC2626"
ORANGE = "D97706"; TEAL  = "0891B2"; PURPLE = "7C3AED"; WHITE = "FFFFFF"
ALT    = "F8FAFC"; LGRAY = "F1F5F9"; HDRBG  = "EFF6FF"; LBLUE = "DBEAFE"

def fill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, sz=10, color="1E293B", italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bb(c="D9D9D9"):
    s = Side(style="thin", color=c); return Border(bottom=s)
def fb(c="CBD5E1"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def title_bar(ws, row, text, bg=DARK, fg=WHITE, sz=14, cols=16):
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{gcl(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", bold=True, size=sz, color=fg)
    c.fill = fill(bg); c.alignment = aln("center", "center")

def sec(ws, row, text, cols=16, bg=LGRAY):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"A{row}:{gcl(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", bold=True, size=9, color="475569")
    c.fill = fill(bg)

def hdr(ws, row, vals, bg="1E3A5F", fg=WHITE, sz=9, h=18):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = Font(name="Arial", bold=True, size=sz, color=fg)
        c.fill = fill(bg); c.alignment = aln("center", "center"); c.border = fb("2C4A7C")

def dc(ws, row, col, val, bg=WHITE, bold=False, fmt=None, h="right", sz=10, color="1E293B"):
    c = ws.cell(row, col, val)
    c.font = Font(name="Arial", bold=bold, size=sz, color=color)
    c.fill = fill(bg); c.border = bb(); c.alignment = aln(h)
    if fmt: c.number_format = fmt
    return c

def zell(ws, row, col, val, bg=WHITE):
    """Zielerreichungs-Zelle mit Ampelfarbe."""
    if val is None or val == 0:
        ws.cell(row, col, "–").fill = fill(bg)
        ws.cell(row, col).border = bb()
        ws.cell(row, col).alignment = aln("center")
        return
    fc, tc = (("D1FAE5", GREEN)  if val >= 1.0  else
              ("DCFCE7", GREEN)  if val >= 0.9  else
              ("FEF3C7", ORANGE) if val >= 0.8  else
              ("FED7AA", ORANGE) if val >= 0.7  else
              ("FEE2E2", RED))
    c = ws.cell(row, col, val)
    c.fill = fill(fc); c.font = Font(name="Arial", bold=True, size=10, color=tc)
    c.number_format = "0.0%"; c.border = bb(); c.alignment = aln("center")

def gap(ws, row, cols=16):
    ws.row_dimensions[row].height = 5
    ws.merge_cells(f"A{row}:{gcl(cols)}{row}")
    ws.cell(row, 1).fill = fill("E2E8F0")

def kpi_tile(ws, row, col, label, val, sub, color=BLUE):
    for r, h in [(row, 13), (row+1, 22), (row+2, 11)]:
        ws.row_dimensions[r].height = h
    for r in [row, row+1, row+2]:
        ws.cell(r, col).fill = fill(HDRBG); ws.cell(r, col).border = fb("BFDBFE")
    ws.cell(row, col, label).font = fnt(sz=9, color="64748B")
    ws.cell(row+1, col, val).font = Font(name="Arial", bold=True, size=13, color=color)
    ws.cell(row+2, col, sub).font = fnt(sz=8, color="94A3B8", italic=True)

# ═══════════════════════════════════════════════════════════════════════
# DATEI-FINDER
# ═══════════════════════════════════════════════════════════════════════

def find_file(key: str) -> Path | None:
    """Findet die neueste Datei für ein Muster. Gibt None zurück wenn nicht gefunden."""
    pattern = FILE_PATTERNS.get(key, "")
    matches = list(INPUT_DIR.glob(pattern))
    if not matches:
        matches = [f for f in INPUT_DIR.iterdir()
                   if re.search(pattern.replace("*", ".*").replace("?", "."),
                                f.name, re.IGNORECASE)]
    if not matches:
        return None
    def sort_key(p):
        # Timestamp im SportsNow-Format: 20260325_0620
        m = re.search(r'(\d{8}_?\d{4})', p.name)
        if m: return m.group(0)
        m = re.search(r'(\d{4})[\._\-]?(\d{2})[\._\-]?(\d{2})', p.name)
        if m: return m.group(0)
        return str(p.stat().st_mtime)
    return sorted(matches, key=sort_key, reverse=True)[0]

def find_all_files(key: str) -> list[Path]:
    """Findet alle Dateien für ein Muster, neueste zuerst."""
    pattern = FILE_PATTERNS.get(key, "")
    matches = list(INPUT_DIR.glob(pattern))
    if not matches:
        matches = [f for f in INPUT_DIR.iterdir()
                   if re.search(pattern.replace("*", ".*").replace("?", "."),
                                f.name, re.IGNORECASE)]
    def sort_key(p):
        m = re.search(r'(\d{4})[\._\-]?(\d{2})[\._\-]?(\d{2})', p.name)
        return m.group(0) if m else str(p.stat().st_mtime)
    return sorted(matches, key=sort_key, reverse=True)

# ═══════════════════════════════════════════════════════════════════════
# DATEN-LOADER
# ═══════════════════════════════════════════════════════════════════════

def load_umsatzanalyse(path: Path) -> dict:
    """Liest alle verfügbaren Monatsdaten aus der Umsatzanalyse."""
    print(f"  Lese Umsatzanalyse: {path.name}")
    ua = pd.read_excel(path, sheet_name=None, header=None)
    result = {
        "months_available": [],  # z.B. ["Jan26","Feb26"]
        "by_month": {},          # month -> list of therapeuten dicts
        "tracking": [],          # Umsatztracking
        "overview": {},          # Übersicht-Daten
        "input_params": {},      # Parameterblatt
    }

    # Input Parameter
    if "Input Parameter" in ua:
        ip = ua["Input Parameter"].dropna(how="all")
        params = {}
        for _, row in ip.iterrows():
            if pd.notna(row[0]) and pd.notna(row[1]):
                params[str(row[0]).strip()] = row[1]
        result["input_params"] = params

    # Umsatztracking
    if "Umsatztracking" in ua:
        tr = ua["Umsatztracking"]
        tracking = []
        for i in range(2, len(tr)):
            row = tr.iloc[i]
            if pd.notna(row[0]) and str(row[0]).strip() not in ["Summe","NaN","nan",""]:
                tracking.append({
                    "name": str(row[0]).strip(),
                    "stand1": float(row[1]) if pd.notna(row[1]) else None,
                    "stand2": float(row[2]) if pd.notna(row[2]) else None,
                    "delta": float(row[3]) if pd.notna(row[3]) else 0,
                })
        result["tracking"] = tracking

    # Übersicht (YTD-Daten)
    if "Übersicht" in ua:
        ov = ua["Übersicht"]
        overview = {"therapeuten": []}
        skip_names = {"xxx","Summe","Check","NaN","nan","","0","0.0"}
        for i in range(4, len(ov)):
            row = ov.iloc[i]
            name = str(row[0]).strip() if pd.notna(row[0]) else ""

            # Rechte Seite: Summary-Werte in col 29/30/31 (unabhängig von col0)
            label29 = str(row[29]).strip() if pd.notna(row[29]) else ""
            val30 = float(row[30]) if pd.notna(row[30]) else None
            val31 = float(row[31]) if pd.notna(row[31]) else None
            if label29 == "FTE":
                overview["fte_jan"] = val30; overview["fte_feb"] = val31
            elif label29 == "Ferien":
                overview["ferien_jan"] = val30; overview["ferien_feb"] = val31
            elif label29 == "Umsatz":
                overview["umsatz_jan"] = val30; overview["umsatz_feb"] = val31
            elif label29 == "Umsatz pro FTE":
                overview["umsatz_fte_jan"] = val30; overview["umsatz_fte_feb"] = val31
            elif label29 == "Zielerreichung Ø":
                overview["ziel_jan"] = val30; overview["ziel_feb"] = val31
            elif label29 == "Krank & andere":
                overview["krank_jan"] = val30; overview["krank_feb"] = val31

            # Therapeuten-Zeilen
            if not name or name in skip_names:
                continue
            bg = float(row[1]) if pd.notna(row[1]) else None
            jan_u  = float(row[2])  if pd.notna(row[2])  else None
            feb_u  = float(row[3])  if pd.notna(row[3])  else None
            jan_zB = float(row[15]) if pd.notna(row[15]) else None
            feb_zB = float(row[16]) if pd.notna(row[16]) else None
            overview["therapeuten"].append({
                "name": name, "bg": bg,
                "jan_umsatz": jan_u, "feb_umsatz": feb_u,
                "jan_zielB": jan_zB, "feb_zielB": feb_zB,
            })
        result["overview"] = overview

    # Monatliche Detail-Sheets
    month_map = {
        "Jan26": "Umsatz pro Mitarbeiter Jan26",
        "Feb26": "Umsatz pro Mitarbeiter Feb26",
        "Mär26": "Umsatz pro Mitarbeiter Mrz26",
        "Apr26": "Umsatz pro Mitarbeiter Apr26",
        "Mai26": "Umsatz pro Mitarbeiter Mai26",
        "Jun26": "Umsatz pro Mitarbeiter Jun26",
        "Jul26": "Umsatz pro Mitarbeiter Jul26",
        "Aug26": "Umsatz pro Mitarbeiter Aug26",
        "Sep26": "Umsatz pro Mitarbeiter Sep26",
        "Okt26": "Umsatz pro Mitarbeiter Okt26",
        "Nov26": "Umsatz pro Mitarbeiter Nov26",
        "Dez26": "Umsatz pro Mitarbeiter Dez26",
    }
    taetigkeiten_map = {
        "Jan26": "Input Tätigkeiten Jan26",
        "Feb26": "Input Tätigkeiten Feb26",
        "Mär26": "Input Tätigkeiten Mrz26",
        "Apr26": "Input Tätigkeiten Apr26",
        "Mai26": "Input Tätigkeiten Mai26",
    }

    for month_key, sheet_name in month_map.items():
        if sheet_name not in ua:
            continue
        df = ua[sheet_name]
        ther_list = []
        has_data = False
        # Summen-Zeile 27: col4=Ferien, col5=Kurse, col6=Mktg, col7=Lauf, col8=Mgmt, col9=Krank
        try:
            sum_row = df.iloc[27]
            result["by_month_totals"] = result.get("by_month_totals", {})
            result["by_month_totals"][month_key] = {
                "ferien":  float(sum_row[4]) if pd.notna(sum_row[4]) else 0,
                "kurse":   float(sum_row[5]) if pd.notna(sum_row[5]) else 0,
                "mktg":    float(sum_row[6]) if pd.notna(sum_row[6]) else 0,
                "lauf":    float(sum_row[7]) if pd.notna(sum_row[7]) else 0,
                "mgmt":    float(sum_row[8]) if pd.notna(sum_row[8]) else 0,
                "krank":   float(sum_row[9]) if pd.notna(sum_row[9]) else 0,
            }
        except: pass
        for i in range(3, len(df)):
            row = df.iloc[i]
            name = str(row[0]).strip() if pd.notna(row[0]) else ""
            if not name or name in ["Summe","NaN","nan","","FTE","Ferien","0","0.0"]:
                # Summenzeilen
                label = str(row[0]).strip()
                if label == "FTE":
                    pass  # captured elsewhere
                continue
            bg = float(row[1]) if pd.notna(row[1]) else None
            umsatz = float(row[14]) if pd.notna(row[14]) else None
            if umsatz and umsatz > 0:
                has_data = True
            zA = float(row[16]) if pd.notna(row[16]) else None
            zB = float(row[18]) if pd.notna(row[18]) else None
            zC = float(row[20]) if pd.notna(row[20]) else None
            t_entry = {
                "name": name, "bg": bg,
                "umsatz": umsatz,
                "zielA": zA, "zielB": zB, "zielC": zC,
                "ferien": float(row[4]) if pd.notna(row[4]) else 0,
                "krank": float(row[9]) if pd.notna(row[9]) else 0,
                "mgmt": float(row[8]) if pd.notna(row[8]) else 0,
                "kurs_tage": float(row[5]) if pd.notna(row[5]) else 0,
                "at": float(row[2]) if pd.notna(row[2]) else 0,
                "sollat": float(row[3]) if pd.notna(row[3]) else 0,
                "tageB": float(row[11]) if pd.notna(row[11]) else 0,
                "kurs_anz": 0, "mktg_h": 0, "lauf_anz": 0,  # from Tätigkeiten
            }
            ther_list.append(t_entry)

        # Tätigkeiten-Daten hinzufügen
        taet_sheet = taetigkeiten_map.get(month_key)
        if taet_sheet and taet_sheet in ua:
            td = ua[taet_sheet]
            for i in range(3, len(td)):
                trow = td.iloc[i]
                tname = str(trow[0]).strip() if pd.notna(trow[0]) else ""
                if not tname or tname in ["0","Summe","NaN","nan",""]:
                    continue
                for t in ther_list:
                    if t["name"] == tname:
                        t["kurs_anz"] = int(trow[3]) if pd.notna(trow[3]) else 0
                        t["mktg_h"] = float(trow[4]) if pd.notna(trow[4]) else 0
                        t["lauf_anz"] = int(trow[5]) if pd.notna(trow[5]) else 0
                        break

        if has_data:
            result["months_available"].append(month_key)
            result["by_month"][month_key] = ther_list

    return result


def load_runnerslab(path: Path) -> dict:
    """Liest Runnerslab Umsätze."""
    print(f"  Lese Runnerslab: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    years = {}
    months = ["Januar","Februar","März","April","Mai","Juni",
              "Juli","August","September","Oktober","November","Dezember"]
    months_short = ["Jan","Feb","Mär","Apr","Mai","Jun",
                    "Jul","Aug","Sep","Okt","Nov","Dez"]
    for i, row in df.iterrows():
        val = str(row[0]).strip()
        if re.match(r'^20\d\d$', val):
            yr = int(val)
            pos_row = df.iloc[i+1]
            onl_row = df.iloc[i+2]
            tot_row = df.iloc[i+3]
            years[yr] = {}
            for j, m in enumerate(months_short):
                years[yr][m] = {
                    "pos":    float(pos_row[j+1]) if pd.notna(pos_row[j+1]) else 0,
                    "online": float(onl_row[j+1]) if pd.notna(onl_row[j+1]) else 0,
                    "total":  float(tot_row[j+1]) if pd.notna(tot_row[j+1]) else 0,
                }
    return years


def load_hyrox(schedule_path: Path, active_path: Path, past_path: Path,
               dashboard_path: Path = None) -> dict:
    """
    Liest Hyrox-Daten aus SportsNow-Exporten + optionalem historischen Dashboard.
    Strategie: Historische Wochen aus altem Dashboard, aktuelle aus neuem Export.
      - dashboard_path: Altes hyrox_dashboard*.xlsx (historische Wochendaten)
      - schedule_path:  Neuer SportsNow Kursplan-Export (aktuelle Wochen)
      - active_path:    SportsNow Aktive Abonnemente
      - past_path:      SportsNow Abgelaufene Abonnemente
    """
    result = {
        "weekly": [], "best_classes": [], "worst_classes": [],
        "by_kurstyp": {}, "by_time": {}, "by_day": {},
        "total_revenue": 0, "pass_sales_monthly": {},
        "conversion": {"trial": 0, "converted": 0, "rate": 0},
        "non_returning": [], "non_returning_by_type": [],
    }

    # ── Kursplan: historisch + aktuell kombinieren ────────────────────

    # Sammle alle Kursdaten in einem DataFrame
    all_sc_frames = []

    # 1. Historische Daten aus altem Dashboard (hyrox_dashboard*.xlsx)
    if dashboard_path and dashboard_path.exists():
        print(f"  Lese Hyrox historische Daten: {dashboard_path.name}")
        try:
            hist = pd.read_excel(dashboard_path, sheet_name="Source Weekly Classes", header=0)
            hist.columns = hist.columns.str.strip()
            hist["WeekStart"] = pd.to_datetime(hist["WeekStart"], errors="coerce")
            for _, row in hist.iterrows():
                if pd.notna(row["WeekStart"]):
                    kw = row["WeekStart"].isocalendar().week
                    result["weekly"].append({
                        "date": row["WeekStart"].strftime("%d.%m.%y"),
                        "kw": int(kw),
                        "kurse": int(row.get("Anzahl Kurse", 0) or 0),
                        "tn": int(row.get("Gesamtteilnehmer", 0) or 0),
                        "plaetze": int(row.get("Gesamtplätze", 0) or 0),
                        "auslastung": float(row.get("Auslastung %", 0) or 0),
                        "avg_tn": float(row.get("Durchschnitt Teilnehmer pro Kurs", 0) or 0),
                        "neue_kunden": 0,
                        "k1":0,"trial":0,"er10":0,"er30":0,"unlim":0,"pt":0,
                    })
            print(f"  → {len(result['weekly'])} historische Wochen geladen")
        except Exception as e:
            print(f"  ⚠️  Historische Daten nicht lesbar: {e}")

    # Bestimme bereits geladene Wochen anhand Datum (nicht KW, da KW sich wiederholen kann)
    loaded_week_dates = set()
    for w in result["weekly"]:
        loaded_week_dates.add(w.get("date",""))

    # 2. Alle SportsNow Kursplan-Exporte einlesen und zusammenfügen
    all_sc_parts = []
    for sp in find_all_files("hyrox_schedule"):
        try:
            df_try = pd.read_excel(sp, header=0)
            df_try.columns = df_try.columns.str.strip()
            if "Anmeldungen" in df_try.columns:
                df_try["_source"] = sp.name
                all_sc_parts.append(df_try)
                print(f"  Lese Kursplan: {sp.name} ({len(df_try)} Kurse)")
        except: pass

    sc_to_use = pd.concat(all_sc_parts, ignore_index=True) if all_sc_parts else None

    if sc_to_use is not None and not sc_to_use.empty:
        sc = sc_to_use.copy()
        sc["Datum"] = pd.to_datetime(sc["Datum"], dayfirst=True, errors="coerce")
        sc = sc[sc["Datum"].notna()].copy()
        # Duplikate entfernen (gleicher Kurs in mehreren Exporten) — neueste Quelle gewinnt
        sc = sc.sort_values("_source").drop_duplicates(
            subset=["Datum","Beginnt um","Stunde"], keep="last"
        ).reset_index(drop=True)
        sc["Auslastung"] = sc["Anmeldungen"] / sc["Plätze"].replace(0, 1)
        sc["Wochentag_DE"] = sc["Datum"].dt.day_name().map({
            "Monday":"Montag","Tuesday":"Dienstag","Wednesday":"Mittwoch",
            "Thursday":"Donnerstag","Friday":"Freitag","Saturday":"Samstag","Sunday":"Sonntag"
        })
        sc["WeekStart"] = sc["Datum"] - pd.to_timedelta(sc["Datum"].dt.weekday, unit="d")
        # Nur Wochen hinzufügen die noch nicht aus historischen Daten geladen wurden
        for ws_date, grp in sc.groupby("WeekStart"):
            kw = int(ws_date.isocalendar().week)
            ws_label = ws_date.strftime("%d.%m.%y")
            if ws_label not in loaded_week_dates:
                result["weekly"].append({
                    "date": ws_date.strftime("%d.%m.%y"),
                    "kw": kw,
                    "kurse": len(grp),
                    "tn": int(grp["Anmeldungen"].sum()),
                    "plaetze": int(grp["Plätze"].sum()),
                    "auslastung": float(grp["Anmeldungen"].sum() / grp["Plätze"].sum()) if grp["Plätze"].sum() > 0 else 0,
                    "avg_tn": float(grp["Anmeldungen"].mean()),
                    "neue_kunden": 0,
                    "k1":0,"trial":0,"er10":0,"er30":0,"unlim":0,"pt":0,
                })
        # Best / Worst aus allen verfügbaren Kursen
        ran = sc[sc["Anmeldungen"] > 0].copy()
        if len(ran) >= 3:
            result["best_classes"] = ran.nlargest(3, "Auslastung")[
                ["Datum","Beginnt um","Stunde","Ort/Raum","Anmeldungen","Plätze","Auslastung","Wochentag_DE"]
            ].rename(columns={"Wochentag_DE":"Wochentag"}).to_dict("records")
            result["worst_classes"] = ran.nsmallest(3, "Auslastung")[
                ["Datum","Beginnt um","Stunde","Ort/Raum","Anmeldungen","Plätze","Auslastung","Wochentag_DE"]
            ].rename(columns={"Wochentag_DE":"Wochentag"}).to_dict("records")
        # Auslastung-Analysen
        for key, col, attr in [("by_kurstyp","Stunde","by_kurstyp"),
                                 ("by_time","Beginnt um","by_time"),
                                 ("by_day","Wochentag_DE","by_day")]:
            result[attr] = ran.groupby(col).agg(
                n=("Anmeldungen","count"),
                avg_tn=("Anmeldungen","mean"),
                avg_ausl=("Auslastung","mean")
            ).sort_values("avg_ausl", ascending=False).to_dict("index")

    # Neue Schedule-Wochen überschreiben historische (neuere Daten sind genauer)
    # Merge: historische als Basis, neue Exporte überschreiben bei Überlappung
    if all_sc_parts:
        sc_weekly_map = {}  # date -> entry
        sc["WeekStart"] = sc["Datum"] - pd.to_timedelta(sc["Datum"].dt.weekday, unit="D")
        for ws_date, grp in sc.groupby("WeekStart"):
            ws_label = ws_date.strftime("%d.%m.%y")
            sc_weekly_map[ws_label] = {
                "date": ws_label,
                "kw": int(ws_date.isocalendar().week),
                "kurse": len(grp),
                "tn": int(grp["Anmeldungen"].sum()),
                "plaetze": int(grp["Plätze"].sum()),
                "auslastung": float(grp["Anmeldungen"].sum() / grp["Plätze"].sum()) if grp["Plätze"].sum() > 0 else 0,
                "avg_tn": float(grp["Anmeldungen"].mean()),
                "neue_kunden": 0,
                "k1":0,"trial":0,"er10":0,"er30":0,"unlim":0,"pt":0,
            }
        # Historische Wochen die NICHT in neuen Exporten sind behalten
        # Wochen die in neuen Exporten sind → neue Version verwenden
        hist_only = [w for w in result["weekly"] if w["date"] not in sc_weekly_map]
        new_weeks = list(sc_weekly_map.values())
        result["weekly"] = hist_only + new_weeks

    # Sortieren nach Datum
    result["weekly"].sort(key=lambda x: x.get("date", ""))

    # ── Pass Sales (Aktiv + Abgelaufen) ───────────────────────────────
    all_passes = []
    for path, label in [(active_path,"aktiv"),(past_path,"abgelaufen")]:
        if path and path.exists():
            print(f"  Lese Hyrox Passes ({label}): {path.name}")
            df = pd.read_excel(path, header=0)
            df.columns = df.columns.str.strip()
            all_passes.append(df)

    if all_passes:
        passes = pd.concat(all_passes, ignore_index=True)
        passes["Kaufdatum"] = pd.to_datetime(passes["Kaufdatum"], dayfirst=True, errors="coerce")
        passes["Monat"] = passes["Kaufdatum"].dt.to_period("M").astype(str)
        result["total_revenue"] = float(passes["Total"].sum())
        result["pass_sales_monthly"] = passes.groupby("Monat")["Total"].sum().to_dict()

        # Produktgruppen vereinfachen
        def produktgruppe(abo):
            abo = str(abo)
            if "Trial" in abo:      return "Trial"
            if "10 Class" in abo:   return "10er"
            if "30 Class" in abo:   return "30er"
            if "Unlimited" in abo:  return "Unlimited"
            if "Single" in abo:     return "1 Karte"
            if "PT" in abo or "Personal Training" in abo: return "PT"
            return "Andere"
        passes["Produktgruppe"] = passes["Abonnement"].apply(produktgruppe)
        result["pass_sales_by_product"] = passes.groupby("Produktgruppe")["Total"].sum().to_dict()

        # Trial Conversion
        trial_kunden = passes[passes["Produktgruppe"]=="Trial"]["Kunde"].nunique()
        trial_ids = set(passes[passes["Produktgruppe"]=="Trial"]["Kunde"])
        converted_ids = set(passes[passes["Produktgruppe"].isin(["10er","30er","Unlimited"])]["Kunde"])
        converted = len(trial_ids & converted_ids)
        result["conversion"] = {
            "trial": trial_kunden,
            "converted": converted,
            "rate": converted/trial_kunden if trial_kunden > 0 else 0,
        }

        # Non-returning customers: abgelaufen aber kein aktives Abo
        if active_path and active_path.exists() and past_path and past_path.exists():
            active_df = pd.read_excel(active_path, header=0)
            past_df   = pd.read_excel(past_path,   header=0)
            active_df.columns = active_df.columns.str.strip()
            past_df.columns   = past_df.columns.str.strip()
            active_ids_set = set(active_df["Kunde"].unique())
            past_ids_set   = set(past_df["Kunde"].unique())
            nr_ids = past_ids_set - active_ids_set
            nr = past_df[past_df["Kunde"].isin(nr_ids)].copy()
            nr["Gültig bis"] = pd.to_datetime(nr["Gültig bis"], dayfirst=True, errors="coerce")
            nr = nr.sort_values("Gültig bis", ascending=False).drop_duplicates("Kunde", keep="first")
            nr = nr.sort_values("Gültig bis", ascending=False)
            result["non_returning"] = nr[
                [c for c in ["Vorname","Name","E-Mail","Abonnement","Gültig bis","Total"] if c in nr.columns]
            ].to_dict("records")
            result["non_returning_by_type"] = nr.groupby("Abonnement").agg(
                Anzahl=("Kunde","count"), Umsatz=("Total","sum")
            ).sort_values("Anzahl", ascending=False).reset_index().to_dict("records")
            print(f"  → Nicht-wiederkehrende Kunden: {len(nr)}")

        print(f"  → Hyrox Total Revenue: CHF {result['total_revenue']:,.0f}")
        print(f"  → Trial Conversion: {converted}/{trial_kunden} = {result['conversion']['rate']:.1%}")

    return result

def load_ausstehend(path: Path) -> list:
    """Liest ausstehende Zahlungen."""
    print(f"  Lese Ausstehende Zahlungen: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    invoices = []
    header_row = None
    for i, row in df.iterrows():
        if str(row[0]).strip() in ["Re.- Nr.", "Re.-Nr."]:
            header_row = i
            break
    if header_row is None:
        return []
    for i in range(header_row+1, len(df)):
        row = df.iloc[i]
        if pd.notna(row[0]) and str(row[0]).strip() not in ["NaN","nan",""]:
            try:
                invoices.append({
                    "renr":      str(row[0]).strip(),
                    "kunde":     str(row[1]).strip() if pd.notna(row[1]) else "",
                    "therapeut": str(row[2]).strip() if pd.notna(row[2]) else "",
                    "an":        str(row[3]).strip() if pd.notna(row[3]) else "",
                    "an_wen":    str(row[4]).strip() if pd.notna(row[4]) else "",
                    "re_datum":  str(row[6]).strip() if pd.notna(row[6]) else "",
                    "letzte_m":  str(row[7]).strip() if pd.notna(row[7]) else "",
                    "mahnstufe": int(row[8]) if pd.notna(row[8]) else 0,
                    "tage":      int(row[9]) if pd.notna(row[9]) else 0,
                    "standort":  int(row[10]) if pd.notna(row[10]) else 0,
                    "betrag":    float(row[12]) if pd.notna(row[12]) else 0,
                })
            except: pass
    return invoices


def load_fitness(path: Path) -> list:
    """Liest Fitness-Abo-Daten."""
    print(f"  Lese Fitness-Abos: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    rows = []
    # Header ist Zeile mit "Was" oder KW-Muster
    header_idx = None
    for i, row in df.iterrows():
        if str(row[0]).strip() == "Was":
            header_idx = i; break
    if header_idx is None:
        return []
    labels = df.iloc[header_idx+1:header_idx+10, 0].tolist()
    kw_cols = df.iloc[header_idx].tolist()
    for j, kw in enumerate(kw_cols[1:], 1):
        if pd.isna(kw): continue
        kw_label = str(kw).strip()
        entry = {"kw": kw_label}
        for k, label in enumerate(labels):
            label_str = str(label).strip() if pd.notna(label) else ""
            val = df.iloc[header_idx+1+k, j]
            entry[label_str] = val if pd.notna(val) else None
        rows.append(entry)
    return rows


def load_berichte(path: Path) -> dict:
    """Liest Berichte an Ärzte."""
    print(f"  Lese Berichte an Ärzte: {path.name}")
    df = pd.read_excel(path, header=None, sheet_name=0)
    if len(df) < 2:
        return {}
    header_row = df.iloc[0].tolist()
    therapists = [str(h) for h in header_row[1:] if pd.notna(h)]
    data = {}
    for i in range(1, len(df)):
        row = df.iloc[i]
        kw = str(row[0]).strip() if pd.notna(row[0]) else ""
        if not kw: continue
        vals = {}
        for j, t in enumerate(therapists):
            v = row[j+1]
            vals[t] = int(v) if pd.notna(v) and v != "" else 0
        data[kw] = vals
    return {"therapists": therapists, "data": data}


# ═══════════════════════════════════════════════════════════════════════
# DASHBOARD-WRITER  — baut alle 8 Sheets
# ═══════════════════════════════════════════════════════════════════════

def write_uebersicht(wb, ua_data, rl_data, hy_data, inv_data, fit_data, today_str):
    ws = wb.active; ws.title = "📊 Übersicht"
    ws.sheet_view.showGridLines = False
    ov = ua_data.get("overview", {})
    months_avail = ua_data.get("months_available", [])
    mt = ua_data.get("by_month_totals", {})
    jan_t = mt.get("Jan26", {})
    feb_t = mt.get("Feb26", {})

    jan_u   = ov.get("umsatz_jan", 0) or 0
    feb_u   = ov.get("umsatz_feb", 0) or 0
    jan_zB  = ov.get("ziel_jan",   0) or 0
    feb_zB  = ov.get("ziel_feb",   0) or 0
    jan_fte = ov.get("fte_jan",    0) or 0
    feb_fte = ov.get("fte_feb",    0) or 0

    rl_jan26 = sum(rl_data.get(2026,{}).get(m,{}).get("total",0) for m in ["Jan","Feb"])
    hy_total = hy_data.get("total_revenue", 0)
    inv_total = sum(i["betrag"] for i in inv_data) if inv_data else 0
    inv_count = len(inv_data)
    fitness_total = None
    for r in reversed(fit_data or []):
        g = r.get("Mitglieder Gesamt")
        if g: fitness_total = int(g); break

    title_bar(ws, 1, f"Kineo AG – Management Dashboard  ·  Stand {today_str}", cols=12)
    ws.merge_cells("A2:L2"); ws["A2"].fill = fill(BLUE); ws.row_dimensions[2].height = 4

    kpis = [
        (1, "Ther. Jan 2026",     f"CHF {jan_u:,.0f}".replace(",","'"),   "Gesamtumsatz",     BLUE),
        (2, "Ther. Feb 2026",     f"CHF {feb_u:,.0f}".replace(",","'"),   "Gesamtumsatz",     BLUE),
        (3, "Ø Zielerr. Jan",     f"{jan_zB:.1%}",                         "ZielB pensenber.", ORANGE),
        (4, "Ø Zielerr. Feb",     f"{feb_zB:.1%}",                         "ZielB pensenber.", ORANGE),
        (5, "Runnerslab Jan–Feb", f"CHF {rl_jan26:,.0f}".replace(",","'"), "2026 aktuell",     GREEN),
        (6, "Hyrox total",        f"CHF {hy_total:,.0f}".replace(",","'"), "seit Start",       TEAL),
        (7, "Ausstehend",         f"CHF {inv_total:,.0f}".replace(",","'"), f"{inv_count} Re.", RED),
        (8, "Fitness-Abos",       f"{fitness_total or '–'} Mitgl.",       "Thalwil aktuell",  GREEN),
    ]
    for col, label, val, sub, color in kpis:
        kpi_tile(ws, 4, col, label, val, sub, color)
    gap(ws, 7, 12)

    # A – Therapeuten Kennzahlen
    sec(ws, 8, "A │ THERAPEUTEN – Kennzahlen aus Umsatzanalyse", cols=12)
    hdr(ws, 9, ["Kennzahl","Jan 2026","Feb 2026","Jan+Feb / Ø"])

    def r(v): return round(v, 1) if isinstance(v, float) else v

    # Tage + Stunden Format (8.4h pro Tag)
    def th(tage, std): return f"{r(tage)} T / {r(std)} h" if tage else "0 T"
    j_krank_h = r(jan_t.get("krank",0)*8.4); f_krank_h = r(feb_t.get("krank",0)*8.4)
    j_kurs_h  = r(jan_t.get("kurse",0)*8.4); f_kurs_h  = r(feb_t.get("kurse",0)*8.4)
    j_mktg_h  = r(jan_t.get("mktg",0)*8.4);  f_mktg_h  = r(feb_t.get("mktg",0)*8.4)
    j_lauf_h  = r(jan_t.get("lauf",0)*8.4);  f_lauf_h  = r(feb_t.get("lauf",0)*8.4)
    j_mgmt_h  = r(jan_t.get("mgmt",0)*8.4);  f_mgmt_h  = r(feb_t.get("mgmt",0)*8.4)

    a_rows = [
        ("Gesamtumsatz (CHF)",    jan_u,               feb_u,               jan_u+feb_u,         "#'##0.00"),
        ("FTE",                   jan_fte,              feb_fte,             (jan_fte+feb_fte)/2,  "0.0"),
        ("Umsatz pro FTE (CHF)",  ov.get("umsatz_fte_jan",0) or 0,
                                  ov.get("umsatz_fte_feb",0) or 0,
                                  ((ov.get("umsatz_fte_jan",0) or 0)+(ov.get("umsatz_fte_feb",0) or 0))/2, "#'##0.00"),
        ("Ø Zielerreichung (B)",  jan_zB,              feb_zB,              (jan_zB+feb_zB)/2,   "0.0%"),
        ("Krankheitstage",
         th(jan_t.get("krank",0), j_krank_h),
         th(feb_t.get("krank",0), f_krank_h),
         th(jan_t.get("krank",0)+feb_t.get("krank",0), r(j_krank_h+f_krank_h)), "0"),
        ("Fitnesskurse",
         th(jan_t.get("kurse",0), j_kurs_h),
         th(feb_t.get("kurse",0), f_kurs_h),
         th(jan_t.get("kurse",0)+feb_t.get("kurse",0), r(j_kurs_h+f_kurs_h)), "0"),
        ("Marketing",
         th(jan_t.get("mktg",0), j_mktg_h),
         th(feb_t.get("mktg",0), f_mktg_h),
         th(jan_t.get("mktg",0)+feb_t.get("mktg",0), r(j_mktg_h+f_mktg_h)), "0"),
        ("Laufanalyse",
         th(jan_t.get("lauf",0), j_lauf_h),
         th(feb_t.get("lauf",0), f_lauf_h),
         th(jan_t.get("lauf",0)+feb_t.get("lauf",0), r(j_lauf_h+f_lauf_h)), "0"),
        ("Management",
         th(jan_t.get("mgmt",0), j_mgmt_h),
         th(feb_t.get("mgmt",0), f_mgmt_h),
         th(jan_t.get("mgmt",0)+feb_t.get("mgmt",0), r(j_mgmt_h+f_mgmt_h)), "0"),
        ("Abwesenheiten Total",
         th(sum(jan_t.get(k,0) for k in ["krank","kurse","mktg","lauf","mgmt"]),
            r(sum([j_krank_h,j_kurs_h,j_mktg_h,j_lauf_h,j_mgmt_h]))),
         th(sum(feb_t.get(k,0) for k in ["krank","kurse","mktg","lauf","mgmt"]),
            r(sum([f_krank_h,f_kurs_h,f_mktg_h,f_lauf_h,f_mgmt_h]))),
         th(sum(jan_t.get(k,0)+feb_t.get(k,0) for k in ["krank","kurse","mktg","lauf","mgmt"]),
            r(sum([j_krank_h,j_kurs_h,j_mktg_h,j_lauf_h,j_mgmt_h,
                   f_krank_h,f_kurs_h,f_mktg_h,f_lauf_h,f_mgmt_h]))), "0"),
        ("Monate mit Daten",      len([m for m in months_avail if "26" in m]), None, None, "0"),
    ]

    r_start = 10
    for i, (label, jv, fv, tot, fmt) in enumerate(a_rows, r_start):
        bg = ALT if i%2==0 else WHITE
        dc(ws, i, 1, label, bg, bold=True, h="left")
        # String values (Tage/Std format) get left alignment, no number format
        is_str = isinstance(jv, str)
        dc(ws, i, 2, jv,  bg, fmt=None if is_str else fmt, h="left" if is_str else "right") if jv  is not None else dc(ws, i, 2, "–", bg, h="center")
        dc(ws, i, 3, fv,  bg, fmt=None if isinstance(fv,str) else fmt, h="left" if isinstance(fv,str) else "right") if fv  is not None else dc(ws, i, 3, "–", bg, h="center")
        dc(ws, i, 4, tot, bg, bold=True, fmt=None if isinstance(tot,str) else fmt, h="left" if isinstance(tot,str) else "right") if tot is not None else dc(ws, i, 4, "–", bg, h="center")

    # Dynamic row positions
    r_b  = r_start + len(a_rows) + 1   # Section B start
    r_b1 = r_b + 1                      # B header
    r_b2 = r_b + 2                      # B data start
    gap(ws, r_b - 1, 12)

    # B – Runnerslab Jahresvergleich
    sec(ws, r_b, "B │ RUNNERSLAB – Jan+Feb Jahresvergleich", cols=12)
    hdr(ws, r_b1, ["Jahr","Jan Total","Feb Total","Jan+Feb","Δ% vs. Vorjahr"])
    years_to_show = sorted([y for y in rl_data.keys() if y >= 2021])
    prev_jf = None
    for i, yr in enumerate(years_to_show, r_b2):
        bg = ALT if i%2==0 else WHITE
        rl_jan = rl_data[yr].get("Jan",{}).get("total",0)
        rl_feb = rl_data[yr].get("Feb",{}).get("total",0)
        jf = rl_jan + rl_feb
        dc(ws, i, 1, str(yr), bg, bold=True, h="center")
        dc(ws, i, 2, rl_jan, bg, fmt="#'##0")
        dc(ws, i, 3, rl_feb, bg, fmt="#'##0")
        dc(ws, i, 4, jf, bg, bold=True, fmt="#'##0")
        if prev_jf and prev_jf > 0:
            delta = (jf - prev_jf) / prev_jf
            c1 = ws.cell(i, 5, delta); c1.number_format="0.0%"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,size=10,color=GREEN if delta>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
        else:
            dc(ws, i, 5, "–", bg, h="center")
        prev_jf = jf

    r_c_gap = r_b2 + len(years_to_show)
    gap(ws, r_c_gap, 12)

    # C – Runnerslab Monat für Monat
    r_c = r_c_gap + 1
    months_s = ["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"]
    show_years = sorted([y for y in rl_data.keys() if y >= 2024])
    sec(ws, r_c, "C │ RUNNERSLAB – Monat für Monat (alle verfügbaren Jahre)", cols=12)
    hdr(ws, r_c+1, ["Monat"] + [str(y) for y in show_years] + ["Δ letzt.2J","Δ%"])
    for i, m in enumerate(months_s):
        rr = r_c+2+i; bg = ALT if rr%2==0 else WHITE
        dc(ws, rr, 1, m, bg, bold=True, h="left")
        vals = []
        for j, yr in enumerate(show_years):
            v = rl_data[yr].get(m,{}).get("total",0) or 0
            dc(ws, rr, j+2, v if v else None, bg, fmt="#'##0", bold=(j==len(show_years)-1))
            vals.append(v)
        if len(vals) >= 2 and vals[-2] and vals[-1]:
            d = vals[-1]-vals[-2]; dp = d/vals[-2]
            c1=ws.cell(rr,len(show_years)+2,d); c1.number_format="#'##0"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,size=10,color=GREEN if d>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
            c2=ws.cell(rr,len(show_years)+3,dp); c2.number_format="0.0%"; c2.fill=fill(bg)
            c2.border=bb(); c2.alignment=aln("right")
        else:
            for cn in [len(show_years)+2, len(show_years)+3]:
                dc(ws, rr, cn, "–", bg, h="center")

    r_d_gap = r_c + 2 + 12
    gap(ws, r_d_gap, 12)

    # D – Hyrox
    r_d = r_d_gap + 1
    sec(ws, r_d, "D │ HYROX – Monatsumsatz & Auslastung 2026", cols=12)
    hdr(ws, r_d+1, ["Monat","Umsatz (CHF)","Käufe","Ø CHF/Kauf","Neue Kunden","Auslastung Ø","Kurse"])
    hy_m = [
        ("Dezember 2025", 4928.0,  15, None, None, None, 2),
        ("Januar 2026",  13605.8,  58, None, 50,   0.582, 37),
        ("Februar 2026",  9185.8,  67, None, 64,   0.649, 48),
        ("März 2026",     8254.0,  58, None, 19,   0.472, 35),
    ]
    for i, (m, rev, n, _, nc, au, ku) in enumerate(hy_m, r_d+2):
        bg = ALT if i%2==0 else WHITE
        dc(ws, i, 1, m, bg, bold=True, h="left"); dc(ws, i, 2, rev, bg, fmt="#'##0.00"); dc(ws, i, 3, n, bg, h="center")
        ws.cell(i,4,f"=B{i}/C{i}").fill=fill(bg); ws.cell(i,4).font=fnt(); ws.cell(i,4).border=bb()
        ws.cell(i,4).alignment=aln("right"); ws.cell(i,4).number_format="#'##0.00"
        dc(ws, i, 5, nc if nc else "–", bg, h="right" if nc else "center")
        if au:
            c_ = ws.cell(i, 6, au); c_.number_format="0.0%"; c_.border=bb(); c_.alignment=aln("center")
            c_.fill=fill("D1FAE5" if au>=0.65 else("FEF3C7" if au>=0.5 else"FEE2E2"))
            c_.font=Font(name="Arial",bold=True,size=10,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED))
        else:
            dc(ws, i, 6, "–", bg, h="center")
        dc(ws, i, 7, ku, bg, h="center")

    for c, w in [(1,24),(2,15),(3,12),(4,13),(5,13),(6,12),(7,10),(8,12),(9,10),(10,8),(11,8),(12,8)]:
        ws.column_dimensions[gcl(c)].width = w

def write_therapeuten(wb, ua_data, today_str):
    ws = wb.create_sheet("👥 Therapeuten")
    ws.sheet_view.showGridLines = False
    months = ua_data.get("months_available", [])
    ov = ua_data.get("overview", {})
    ther_list = ov.get("therapeuten", [])

    jan_u = ov.get("umsatz_jan",0) or 0
    feb_u = ov.get("umsatz_feb",0) or 0
    jan_zB = ov.get("ziel_jan",0) or 0
    feb_zB = ov.get("ziel_feb",0) or 0

    title_bar(ws,1,f"Therapeuten – Umsatz & Zielerreichung  ·  {today_str}",cols=11)
    ws.merge_cells("A2:K2"); ws["A2"].fill=fill(BLUE); ws.row_dimensions[2].height=4

    kpi_ma=[
        (1,"Jan Umsatz",f"CHF {jan_u:,.0f}".replace(",","'"),BLUE),
        (2,"Feb Umsatz",f"CHF {feb_u:,.0f}".replace(",","'"),BLUE),
        (3,"Ø Ziel(B) Jan",f"{jan_zB:.1%}",ORANGE),
        (4,"Ø Ziel(B) Feb",f"{feb_zB:.1%}",ORANGE),
        (5,"Jan FTE",f"{ov.get('fte_jan',0):.1f}",GREEN),
        (6,"Feb FTE",f"{ov.get('fte_feb',0):.1f}",GREEN),
        (7,"Jan Umsatz/FTE",f"CHF {ov.get('umsatz_fte_jan',0):,.0f}".replace(",","'"),TEAL),
        (8,"Feb Umsatz/FTE",f"CHF {ov.get('umsatz_fte_feb',0):,.0f}".replace(",","'"),TEAL),
        (9,"Monate mit Daten",str(len(months)),DARK),
        (10,"Jan Krank",f"{ov.get('krank_jan',0):.0f} T",RED),
        (11,"Feb Krank",f"{ov.get('krank_feb',0):.0f} T",RED),
    ]
    for col,label,val,color in kpi_ma: kpi_tile(ws,4,col,label,val,"2026",color)
    gap(ws,7,11)

    # Ranking nach Ø ZielB
    def avg_zB(t): 
        vals=[v for v in [t.get("jan_zielB"), t.get("feb_zielB")] if v]
        return sum(vals)/len(vals) if vals else 0
    ranked = sorted(ther_list, key=avg_zB, reverse=True)

    sec(ws,8,"RANKING – Ø Zielerreichung (B) · pensenbereinigt · fairster Vergleich",cols=11)
    hdr(ws,9,["Rang","Therapeut/in","BG%","Jan Umsatz","Jan Ziel(B)",
              "Feb Umsatz","Feb Ziel(B)","Total Umsatz","Ø Ziel(B)","Trend","Note"],h=20)
    for rank,t in enumerate(ranked,1):
        i=9+rank; bg=ALT if i%2==0 else WHITE
        medal="🥇" if rank==1 else("🥈" if rank==2 else("🥉" if rank==3 else str(rank)))
        jzb=t.get("jan_zielB"); fzb=t.get("feb_zielB")
        avg_z=avg_zB(t)
        trend=("▲" if (fzb or 0)>(jzb or 0) else("▼" if (fzb or 0)<(jzb or 0) else "=")) if jzb and fzb else "–"
        trend_c=GREEN if trend=="▲" else(RED if trend=="▼" else "94A3B8")
        ws.cell(i,1,rank if rank>3 else medal).fill=fill(bg)
        ws.cell(i,1).font=fnt(bold=True,sz=10); ws.cell(i,1).border=bb(); ws.cell(i,1).alignment=aln("center")
        dc(ws,i,2,t["name"],bg,bold=True,h="left")
        dc(ws,i,3,(t.get("bg") or 0)/100,bg,fmt="0%",h="center")
        dc(ws,i,4,t.get("jan_umsatz"),bg,fmt="#'##0.00") if t.get("jan_umsatz") else dc(ws,i,4,"–",bg,h="center")
        zell(ws,i,5,jzb,bg)
        dc(ws,i,6,t.get("feb_umsatz"),bg,fmt="#'##0.00") if t.get("feb_umsatz") else dc(ws,i,6,"–",bg,h="center")
        zell(ws,i,7,fzb,bg)
        dc(ws,i,8,(t.get("jan_umsatz") or 0)+(t.get("feb_umsatz") or 0),bg,bold=True,fmt="#'##0.00")
        zell(ws,i,9,avg_z if avg_z>0 else None,bg)
        dc(ws,i,10,trend,bg,h="center",bold=True,color=trend_c)
        note="Teilzeit" if (t.get("bg") or 100)<=50 else ""
        dc(ws,i,11,note,bg,h="center",sz=9,color="64748B")

    r_rt=9+len(ranked)+1
    hdr(ws,r_rt,["","TOTAL/Ø","",
                 f"=SUM(D10:D{r_rt-1})",f"=AVERAGE(E10:E{r_rt-1})",
                 f"=SUM(F10:F{r_rt-1})",f"=AVERAGE(G10:G{r_rt-1})",
                 f"=SUM(H10:H{r_rt-1})",f"=AVERAGE(I10:I{r_rt-1})","",""])
    for c in [4,6,8]: ws.cell(r_rt,c).number_format="#'##0.00"
    for c in [5,7,9]: ws.cell(r_rt,c).number_format="0.0%"
    gap(ws,r_rt+1,11)

    # Detail-Tabellen pro Monat
    row_cur = r_rt+2
    for month in months:
        month_data = ua_data["by_month"].get(month, [])
        if not month_data: continue
        sec(ws,row_cur,f"{month} – Detail pro Therapeut/in",cols=11)
        hdr(ws,row_cur+1,["Therapeut/in","BG%","Ferien","Krank","Mgmt(T)",
                          "Kurse(T)","TageB","Umsatz","CHF/TagB","Ziel%(A)","Ziel%(B)"],h=18,sz=8)
        r=row_cur+2
        for t in month_data:
            bg=ALT if r%2==0 else WHITE
            dc(ws,r,1,t["name"],bg,bold=True,h="left")
            dc(ws,r,2,(t.get("bg") or 0)/100,bg,fmt="0%",h="center")
            dc(ws,r,3,t.get("ferien",0),bg,fmt="0.0",h="center")
            dc(ws,r,4,t.get("krank",0),bg,h="center")
            dc(ws,r,5,t.get("mgmt",0),bg,h="center")
            dc(ws,r,6,round(t.get("kurs_tage",0),2),bg,fmt="0.00",h="center")
            dc(ws,r,7,round(t.get("tageB",0),1),bg,fmt="0.0")
            u=t.get("umsatz"); dc(ws,r,8,u,bg,bold=True,fmt="#'##0.00") if u else dc(ws,r,8,"–",bg,h="center")
            fte=round(u/t["tageB"],2) if u and t.get("tageB") else None
            dc(ws,r,9,fte,bg,fmt="#'##0.00") if fte else dc(ws,r,9,"–",bg,h="center")
            zell(ws,r,10,t.get("zielA"),bg); zell(ws,r,11,t.get("zielB"),bg)
            r+=1
        # Totals
        r_mt=r
        hdr(ws,r_mt,[f"{month} TOTAL","","",
                     f"=SUM(D{row_cur+2}:D{r_mt-1})",f"=SUM(E{row_cur+2}:E{r_mt-1})","","",
                     f"=SUM(H{row_cur+2}:H{r_mt-1})","",
                     f"=AVERAGE(J{row_cur+2}:J{r_mt-1})",f"=AVERAGE(K{row_cur+2}:K{r_mt-1})"])
        for c in [8]: ws.cell(r_mt,c).number_format="#'##0.00"
        for c in [10,11]: ws.cell(r_mt,c).number_format="0.0%"
        row_cur = r_mt+2
        gap(ws,row_cur-1,11)

    # Umsatztracking
    tracking = ua_data.get("tracking",[])
    if tracking:
        sec(ws,row_cur,"UMSATZTRACKING – Monatsabschluss-Vergleich",cols=11)
        hdr(ws,row_cur+1,["Therapeut/in","Stand 1","Stand 2 (final)","Delta","Trend"])
        for i,t in enumerate(tracking,row_cur+2):
            bg=ALT if i%2==0 else WHITE
            dc(ws,i,1,t["name"],bg,bold=True,h="left")
            dc(ws,i,2,t.get("stand1"),bg,fmt="#'##0.00")
            dc(ws,i,3,t.get("stand2"),bg,bold=True,fmt="#'##0.00")
            d=t.get("delta",0)
            dc(ws,i,4,d,bg,fmt="#'##0.00",bold=True,color=GREEN if d>0 else(RED if d<0 else "94A3B8"))
            dc(ws,i,5,"▲" if d>0 else("▼" if d<0 else "="),bg,h="center",bold=True,
               color=GREEN if d>0 else(RED if d<0 else "94A3B8"))

    ws.merge_cells(f"A{row_cur+len(tracking)+3}:K{row_cur+len(tracking)+3}")
    ws.cell(row_cur+len(tracking)+3,1,
        "ℹ️  Ranking-Basis: Ziel(B) = Umsatz ÷ (TageB × CHF 1'040)  |  "
        "TageB = SollAT − Ferien, inkl. Kurse & Weiterb., exkl. Krank & Mgmt  |  "
        "Pensenbereinigt = fairer Vergleich Vollzeit/Teilzeit").font=fnt(sz=8,color="475569",italic=True)
    ws.cell(row_cur+len(tracking)+3,1).fill=fill(LBLUE)
    ws.row_dimensions[row_cur+len(tracking)+3].height=14

    for c,w in [(1,16),(2,7),(3,8),(4,8),(5,8),(6,9),(7,8),(8,14),(9,12),(10,10),(11,10)]:
        ws.column_dimensions[gcl(c)].width=w


def write_runnerslab(wb, rl_data):
    ws = wb.create_sheet("👟 Runnerslab")
    ws.sheet_view.showGridLines = False
    title_bar(ws,1,"Runnerslab – Jahresvergleich & Monat-für-Monat",bg="1D4ED8",cols=10)

    months_s=["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"]
    years_sorted=sorted(rl_data.keys())

    # Jahrestabelle
    sec(ws,3,"Jahresgesamtumsatz",cols=10)
    hdr(ws,4,["Jahr","POS (CHF)","Online (CHF)","Total (CHF)","Δ CHF","Δ%"])
    prev_tot=None
    for i,yr in enumerate(years_sorted,5):
        bg=ALT if i%2==0 else WHITE
        pos=sum(rl_data[yr].get(m,{}).get("pos",0) for m in months_s)
        onl=sum(rl_data[yr].get(m,{}).get("online",0) for m in months_s)
        tot=sum(rl_data[yr].get(m,{}).get("total",0) for m in months_s)
        dc(ws,i,1,str(yr),bg,bold=True,h="center")
        dc(ws,i,2,pos,bg,fmt="#'##0"); dc(ws,i,3,onl,bg,fmt="#'##0")
        dc(ws,i,4,tot,bg,bold=True,fmt="#'##0")
        if prev_tot and prev_tot>0:
            d=tot-prev_tot; dp=d/prev_tot
            c1=ws.cell(i,5,d); c1.number_format="#'##0"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,sz=10,color=GREEN if d>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
            c2=ws.cell(i,6,dp); c2.number_format="0.0%"; c2.fill=fill(bg)
            c2.border=bb(); c2.alignment=aln("right")
        else:
            dc(ws,i,5,"–",bg,h="center"); dc(ws,i,6,"–",bg,h="center")
        prev_tot=tot

    gap(ws,5+len(years_sorted),10)

    # Monat für Monat
    r_mfm=6+len(years_sorted)
    show_yrs=sorted([y for y in years_sorted if y>=2024])
    sec(ws,r_mfm,"Monat für Monat",cols=10)
    hdr(ws,r_mfm+1,["Monat"]+[str(y) for y in show_yrs]+["Δ letzt.","Δ%"])
    for i,m in enumerate(months_s):
        r=r_mfm+2+i; bg=ALT if r%2==0 else WHITE
        dc(ws,r,1,m,bg,bold=True,h="left")
        vals=[]
        for j,yr in enumerate(show_yrs):
            v=rl_data[yr].get(m,{}).get("total",0) or 0
            dc(ws,r,j+2,v if v else None,bg,fmt="#'##0",bold=(j==len(show_yrs)-1))
            vals.append(v)
        if len(vals)>=2 and vals[-2] and vals[-1]:
            d=vals[-1]-vals[-2]; dp=d/vals[-2]
            c1=ws.cell(r,len(show_yrs)+2,d); c1.number_format="#'##0"; c1.fill=fill(bg)
            c1.font=Font(name="Arial",bold=True,sz=10,color=GREEN if d>0 else RED)
            c1.border=bb(); c1.alignment=aln("right")
            c2=ws.cell(r,len(show_yrs)+3,dp); c2.number_format="0.0%"; c2.fill=fill(bg)
            c2.border=bb(); c2.alignment=aln("right")
        else:
            for c_ in [len(show_yrs)+2,len(show_yrs)+3]:
                dc(ws,r,c_,"–",bg,h="center")

    # Chart Monatsverlauf
    r_chart=r_mfm+2+12+2
    chart=LineChart(); chart.title="Monatsverlauf"; chart.width=20; chart.height=12
    for j,yr in enumerate(show_yrs):
        s=Series(Reference(ws,min_col=j+2,min_row=r_mfm+1,max_row=r_mfm+13),title=str(yr))
        chart.series.append(s)
    chart.set_categories(Reference(ws,min_col=1,min_row=r_mfm+2,max_row=r_mfm+13))
    ws.add_chart(chart,"A"+str(r_chart))

    for c,w in [(1,12)]+[(i,12) for i in range(2,10)]:
        ws.column_dimensions[gcl(c)].width=w


def write_hyrox(wb, hy_data):
    ws = wb.create_sheet("🏃 Hyrox")
    ws.sheet_view.showGridLines = False
    title_bar(ws,1,"Hyrox – Dashboard",bg=TEAL,cols=14)
    ws.merge_cells("A2:N2"); ws["A2"].fill=fill("06B6D4"); ws.row_dimensions[2].height=4

    weekly = hy_data.get("weekly",[])
    total_tn=sum(w.get("tn",0) for w in weekly)
    total_kurse=sum(w.get("kurse",0) for w in weekly)
    avg_ausl=sum(w.get("auslastung",0) for w in weekly)/len(weekly) if weekly else 0
    neue_k=sum(w.get("neue_kunden",0) for w in weekly)
    cv=hy_data.get("conversion",{})

    kpis=[(1,"Total Umsatz",f"CHF {hy_data.get('total_revenue',0):,.0f}".replace(",","'")),
          (3,"Teilnehmer",str(total_tn)),(5,"Kurse",str(total_kurse)),
          (7,"Ø Auslastung",f"{avg_ausl:.1%}"),(9,"Neue Kunden",str(neue_k)),
          (11,"Trial Conv.",f"{cv.get('rate',0):.1%}"),(13,"Trial→10er",f"{cv.get('converted',0)} von {cv.get('trial',0)}")]
    for col,lbl,val in kpis: kpi_tile(ws,4,col,lbl,val,"aktuell",TEAL)
    gap(ws,7,14)

    # Wochentabelle
    sec(ws,8,"Wöchentliche Übersicht",cols=14)
    hdr(ws,9,["Woche ab","KW","Kurse","TN","Plätze","Ausl.%","Ø TN","Neue K.",
              "1Karte","Trial","10er","30er","Unlim.","PT"],bg="0E4F5C")
    for i,w in enumerate(weekly,10):
        bg=ALT if i%2==0 else WHITE
        dc(ws,i,1,w.get("date","")[:10],bg,h="left"); dc(ws,i,2,w.get("kw"),bg,h="center")
        dc(ws,i,3,w.get("kurse"),bg,h="center"); dc(ws,i,4,w.get("tn"),bg,h="center")
        dc(ws,i,5,w.get("plaetze"),bg,h="center")
        au=w.get("auslastung",0)
        c_=ws.cell(i,6,au); c_.number_format="0.0%"; c_.border=bb(); c_.alignment=aln("center")
        c_.fill=fill("D1FAE5" if au>=0.65 else("FEF3C7" if au>=0.5 else"FEE2E2"))
        c_.font=Font(name="Arial",bold=True,sz=10,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED))
        dc(ws,i,7,round(w.get("avg_tn",0),2),bg,fmt="0.00",h="center")
        dc(ws,i,8,w.get("neue_kunden"),bg,h="center")
        for c_n,k in zip(range(9,15),["k1","trial","er10","er30","unlim","pt"]):
            dc(ws,i,c_n,w.get(k,0) or None,bg,h="center")
    r_ht=10+len(weekly)
    hdr(ws,r_ht,["TOTAL","",
                 f"=SUM(C10:C{r_ht-1})",f"=SUM(D10:D{r_ht-1})","",
                 f"=AVERAGE(F10:F{r_ht-1})",f"=AVERAGE(G10:G{r_ht-1})",
                 f"=SUM(H10:H{r_ht-1})"]+[f"=SUM({gcl(c)}10:{gcl(c)}{r_ht-1})" for c in range(9,15)],bg="0E4F5C")
    ws.cell(r_ht,6).number_format="0.0%"
    gap(ws,r_ht+1,14)

    # Best / Worst Kurse
    r_bw=r_ht+2
    sec(ws,r_bw,"TOP 3 BESTE KURSE (höchste Auslastung)",cols=14,bg="D1FAE5")
    hdr(ws,r_bw+1,["Datum","Uhrzeit","Kurstyp","Standort","Anm.","Plätze","Auslastung","Wochentag"],bg="14532D")
    for i,cls in enumerate(hy_data.get("best_classes",[]),r_bw+2):
        bg="F0FDF4"
        dc(ws,i,1,str(cls.get("Datum",""))[:10],bg,bold=True,h="left")
        dc(ws,i,2,cls.get("Beginnt um",""),bg,h="center")
        dc(ws,i,3,cls.get("Stunde",""),bg,h="left")
        dc(ws,i,4,cls.get("Ort/Raum",""),bg,h="left")
        dc(ws,i,5,cls.get("Anmeldungen"),bg,h="center"); dc(ws,i,6,cls.get("Plätze"),bg,h="center")
        c_=ws.cell(i,7,cls.get("Auslastung",0)); c_.number_format="0.0%"; c_.fill=fill("D1FAE5")
        c_.font=Font(name="Arial",bold=True,sz=10,color=GREEN); c_.border=bb(); c_.alignment=aln("center")
        dc(ws,i,8,cls.get("Wochentag",""),bg,h="center")

    r_ww=r_bw+6
    sec(ws,r_ww,"FLOP 3 SCHLECHTESTE KURSE (tiefste Auslastung, mind. 1 TN)",cols=14,bg="FEE2E2")
    hdr(ws,r_ww+1,["Datum","Uhrzeit","Kurstyp","Standort","Anm.","Plätze","Auslastung","Wochentag","Hinweis"],bg="7F1D1D")
    hints={"06:30":"Früh-Slot","06:00":"Früh-Slot"}
    for i,cls in enumerate(hy_data.get("worst_classes",[]),r_ww+2):
        bg="FFF1F1"
        dc(ws,i,1,str(cls.get("Datum",""))[:10],bg,bold=True,h="left")
        dc(ws,i,2,cls.get("Beginnt um",""),bg,h="center")
        dc(ws,i,3,cls.get("Stunde",""),bg,h="left")
        dc(ws,i,4,cls.get("Ort/Raum",""),bg,h="left")
        dc(ws,i,5,cls.get("Anmeldungen"),bg,h="center"); dc(ws,i,6,cls.get("Plätze"),bg,h="center")
        c_=ws.cell(i,7,cls.get("Auslastung",0)); c_.number_format="0.0%"; c_.fill=fill("FEE2E2")
        c_.font=Font(name="Arial",bold=True,sz=10,color=RED); c_.border=bb(); c_.alignment=aln("center")
        dc(ws,i,8,cls.get("Wochentag",""),bg,h="center")
        hint=hints.get(cls.get("Beginnt um",""),"Zu wenig Nachfrage")
        dc(ws,i,9,hint,bg,sz=9,color=ORANGE,h="left")

    r_an=r_ww+6
    sec(ws,r_an,"AUSLASTUNGS-ANALYSE",cols=14)
    hdr(ws,r_an+1,["Kategorie","Anz. Kurse","Ø TN/Kurs","Ø Auslastung","Empfehlung"])
    row_cur=r_an+2
    for label, data_dict in [("nach Kurstyp", hy_data.get("by_kurstyp",{})),
                              ("nach Uhrzeit", hy_data.get("by_time",{})),
                              ("nach Wochentag", hy_data.get("by_day",{}))]:
        ws.cell(row_cur,1,f"— {label} —").font=fnt(bold=True,sz=9,color=TEAL)
        ws.cell(row_cur,1).fill=fill("E0F7FA"); ws.row_dimensions[row_cur].height=14
        row_cur+=1
        for key, vals in data_dict.items():
            bg=ALT if row_cur%2==0 else WHITE
            au=vals.get("avg_ausl",0)
            dc(ws,row_cur,1,str(key),bg,bold=True,h="left")
            dc(ws,row_cur,2,vals.get("n",0),bg,h="center")
            dc(ws,row_cur,3,round(vals.get("avg_tn",0),1),bg,fmt="0.0")
            c_=ws.cell(row_cur,4,au); c_.number_format="0.0%"; c_.fill=fill("D1FAE5" if au>=0.65 else("FEF3C7" if au>=0.5 else"FEE2E2"))
            c_.font=Font(name="Arial",bold=True,sz=10,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED))
            c_.border=bb(); c_.alignment=aln("center")
            rec="⭐ Top-Slot" if au>=0.65 else("Gut" if au>=0.5 else"⚠️ Prüfen")
            dc(ws,row_cur,5,rec,bg,sz=9,color=GREEN if au>=0.65 else(ORANGE if au>=0.5 else RED),h="left")
            row_cur+=1

    for c,w in [(1,14),(2,7),(3,8),(4,10),(5,8),(6,8),(7,9),(8,12),(9,14),(10,8),(11,8),(12,8),(13,8),(14,8)]:
        ws.column_dimensions[gcl(c)].width=w

    # ── Nicht-wiederkehrende Kunden ───────────────────────────────────
    # Erste freie Zeile bestimmen
    r_nr_start = ws.max_row + 2
    gap(ws, r_nr_start - 1, 14)
    nr_list      = hy_data.get("non_returning", [])
    nr_type_list = hy_data.get("non_returning_by_type", [])

    sec(ws, r_nr_start,
        f"NICHT-WIEDERKEHRENDE KUNDEN – Abgelaufenes Abo, kein Folge-Kauf · Total: {len(nr_list)} Kunden",
        cols=14, bg="FEE2E2")

    # Zusammenfassung nach Typ
    hdr(ws, r_nr_start+1, ["Abo-Typ","Anzahl Kunden","Umsatz letztes Abo (CHF)","Empfehlung"], bg="7F1D1D")
    for idx, row_d in enumerate(nr_type_list, r_nr_start+2):
        bg = ALT if idx % 2 == 0 else WHITE
        dc(ws, idx, 1, row_d.get("Abonnement",""), bg, bold=True, h="left")
        dc(ws, idx, 2, row_d.get("Anzahl", 0), bg, h="center")
        dc(ws, idx, 3, row_d.get("Umsatz", 0), bg, fmt="#'##0.00")
        rec = "Re-Aktivierung per E-Mail" if "Trial" in str(row_d.get("Abonnement","")) else "Angebot senden"
        dc(ws, idx, 4, rec, bg, h="left", sz=9, color=ORANGE)

    r_detail_start = r_nr_start + 2 + len(nr_type_list) + 1
    gap(ws, r_detail_start - 1, 14)
    sec(ws, r_detail_start, "DETAIL – Retargeting-Liste (Name, E-Mail, letztes Abo)", cols=14, bg="FEF2F2")
    hdr(ws, r_detail_start+1,
        ["Vorname","Name","E-Mail","Letztes Abo","Gültig bis","Betrag (CHF)"], bg="7F1D1D")
    for idx, row_d in enumerate(nr_list, r_detail_start+2):
        bg = ALT if idx % 2 == 0 else WHITE
        dc(ws, idx, 1, row_d.get("Vorname",""), bg, h="left")
        dc(ws, idx, 2, row_d.get("Name",""), bg, h="left")
        dc(ws, idx, 3, row_d.get("E-Mail",""), bg, h="left", sz=9)
        dc(ws, idx, 4, row_d.get("Abonnement",""), bg, h="left", sz=9)
        gueltig = row_d.get("Gültig bis")
        dc(ws, idx, 5, str(gueltig)[:10] if gueltig else "–", bg, h="center")
        dc(ws, idx, 6, row_d.get("Total", 0), bg, fmt="#'##0.00")

    r_total = r_detail_start + 2 + len(nr_list)
    hdr(ws, r_total,
        ["TOTAL", "", f"{len(nr_list)} Kunden reaktivieren", "", "",
         f"=SUM(F{r_detail_start+2}:F{r_total-1})"], bg="7F1D1D")
    ws.cell(r_total, 6).number_format = "#'##0.00"


def write_simple_sheet(wb, title, bg_color, data_rows, col_headers, col_widths=None, cols=12):
    """Generischer Sheet-Writer für einfache Datentabellen."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    title_bar(ws,1,title.replace("💳 ","").replace("🏋 ","").replace("⚖️ ","").replace("📋 ",""),
              bg=bg_color,cols=cols)
    ws.merge_cells(f"A2:{gcl(cols)}2"); ws["A2"].fill=fill(bg_color); ws.row_dimensions[2].height=4
    hdr(ws,3,col_headers)
    for i,row in enumerate(data_rows,4):
        bg=ALT if i%2==0 else WHITE
        for j,val in enumerate(row,1):
            dc(ws,i,j,val,bg,h="left" if j==1 else "right")
    if col_widths:
        for c,w in enumerate(col_widths,1):
            ws.column_dimensions[gcl(c)].width=w
    return ws


# ═══════════════════════════════════════════════════════════════════════
# HAUPTPROGRAMM
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("Kineo AG Dashboard Updater")
    print("=" * 60)

    # Verzeichnisse erstellen
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    today_str = datetime.today().strftime("%d.%m.%Y")

    # Dateien finden
    print("\n[1/3] Suche Quelldateien in input/...")
    files_found = {}
    files_missing = []
    for key in FILE_PATTERNS:
        f = find_file(key)
        if f:
            files_found[key] = f
            print(f"  ✓ {key}: {f.name}")
        else:
            files_missing.append(key)
            print(f"  ✗ {key}: nicht gefunden (Muster: {FILE_PATTERNS[key]})")

    if not files_found.get("umsatzanalyse"):
        print("\n⚠️  Umsatzanalyse nicht gefunden — ohne diese kann kein Dashboard erstellt werden.")
        print("  Bitte Datei in den input/-Ordner legen.")
        return

    # Daten laden
    print("\n[2/3] Lade Daten...")
    ua_data = load_umsatzanalyse(files_found["umsatzanalyse"])
    print(f"  → Monate mit Daten: {ua_data['months_available']}")

    rl_data = {}
    if files_found.get("runnerslab"):
        rl_data = load_runnerslab(files_found["runnerslab"])
        print(f"  → Runnerslab-Jahre: {sorted(rl_data.keys())}")

    hy_data = {}
    hy_sched     = find_file("hyrox_schedule")
    hy_dashboard = find_file("hyrox_dashboard")
    hy_active    = find_file("hyrox_active")
    hy_past      = find_file("hyrox_past")
    if hy_sched or hy_active or hy_past or hy_dashboard:
        hy_data = load_hyrox(hy_sched, hy_active, hy_past, hy_dashboard)
        print(f"  → Hyrox-Wochen: {len(hy_data.get('weekly', []))}")
    else:
        print("  ✗ hyrox: keine SportsNow-Exporte gefunden")
        print("    Erwartet: *hyrox*schedules*.xlsx, *hyrox*active*passes*.xlsx, *hyrox*past*passes*.xlsx")

    inv_data = []
    if files_found.get("ausstehend"):
        inv_data = load_ausstehend(files_found["ausstehend"])
        print(f"  → Ausstehende Rechnungen: {len(inv_data)}")

    fit_data = []
    if files_found.get("fitness"):
        fit_data = load_fitness(files_found["fitness"])
        print(f"  → Fitness-Einträge: {len(fit_data)}")

    ber_data = {}
    if files_found.get("berichte"):
        ber_data = load_berichte(files_found["berichte"])
        print(f"  → Berichte Wochen: {len(ber_data.get('data',{}))}")

    # Dashboard schreiben
    print("\n[3/3] Schreibe Dashboard...")
    wb = Workbook()
    wb.active.title = "_temp"

    write_uebersicht(wb, ua_data, rl_data, hy_data, inv_data, fit_data, today_str)
    write_therapeuten(wb, ua_data, today_str)
    write_runnerslab(wb, rl_data)

    if hy_data:
        write_hyrox(wb, hy_data)
    else:
        ws = wb.create_sheet("🏃 Hyrox")
        ws.cell(1,1,"Hyrox-Datei nicht gefunden — bitte in input/ ablegen")

    # Fitness
    ws_fit = wb.create_sheet("🏋 Fitness-Abos")
    ws_fit.sheet_view.showGridLines = False
    title_bar(ws_fit,1,"Fitness-Abos Thalwil",bg=GREEN,cols=8)
    ws_fit.merge_cells("A2:H2"); ws_fit["A2"].fill=fill("22C55E"); ws_fit.row_dimensions[2].height=4
    hdr(ws_fit,3,["Periode","2-Jahres","1-Jahr","6-Monate","MTT Plus","Total","Neue","Kündig."],bg="14532D")
    for i,row in enumerate(fit_data,4):
        bg=ALT if i%2==0 else WHITE
        kw=row.get("kw","")
        y2=row.get("Anzahl aktive Verträge 2 Jahr") or row.get("Anzahl aktive Vertr\u00e4ge 2 Jahr")
        y1=row.get("Anzahl aktive Verträge 1 Jahr") or row.get("Anzahl aktive Vertr\u00e4ge 1 Jahr")
        m6=row.get("Anzahl aktive Verträge  6 Monate") or row.get("Anzahl aktive Vertr\u00e4ge  6 Monate")
        mtt=row.get("Anzahl aktive Verträge MTT Plus") or row.get("Anzahl aktive Vertr\u00e4ge MTT Plus")
        tot=row.get("Mitglieder Gesamt")
        neu=row.get("Anzahl neue Verträge") or row.get("Anzahl neue Vertr\u00e4ge")
        kue=row.get("Anzahl Kündigungen")
        dc(ws_fit,i,1,kw,bg,bold=True,h="left")
        for c_,v in zip([2,3,4,5,6,7,8],[y2,y1,m6,mtt,tot,neu,kue]):
            if v is not None and str(v) not in ["nan","None",""]:
                try: dc(ws_fit,i,c_,int(v),bg,h="center")
                except: dc(ws_fit,i,c_,"–",bg,h="center")
            else: dc(ws_fit,i,c_,"–",bg,h="center")
    for c,w in [(1,14)]+[(i,10) for i in range(2,9)]:
        ws_fit.column_dimensions[gcl(c)].width=w

    # Ausstehende Zahlungen
    ws_az = wb.create_sheet("💳 Ausstehende Zhlg.")
    ws_az.sheet_view.showGridLines = False
    title_bar(ws_az,1,f"Ausstehende Zahlungen · {len(inv_data)} Rechnungen · Total CHF {sum(i['betrag'] for i in inv_data):,.0f}".replace(",","'"),
              bg=RED,cols=12)
    ws_az.merge_cells("A2:L2"); ws_az["A2"].fill=fill("EF4444"); ws_az.row_dimensions[2].height=4
    hdr(ws_az,3,["Re.-Nr.","Kunde","Therapeut","Typ","Rechnung an","Re.-Datum","Letzte Mahn.","Mahnstufe","Tage überf.","Standort","Betrag (CHF)","Inkasso?"],bg="7F1D1D")
    for i,inv in enumerate(inv_data,4):
        d=inv.get("tage",0)
        rb=fill("FEE2E2" if d>365 else("FED7AA" if d>180 else("FEF3C7" if d>90 else(ALT if i%2==0 else WHITE))))
        vals=[inv.get("renr"),inv.get("kunde"),inv.get("therapeut"),inv.get("an"),
              inv.get("an_wen"),inv.get("re_datum"),inv.get("letzte_m"),inv.get("mahnstufe"),
              inv.get("tage"),inv.get("standort"),inv.get("betrag"),""]
        for j,v in enumerate(vals,1):
            ce=ws_az.cell(i,j,v); ce.fill=rb; ce.font=fnt(sz=9,bold=(j==11)); ce.border=bb()
            ce.alignment=aln("right" if j in[1,8,9,10,11] else"left")
        ws_az.cell(i,11).number_format="#'##0.00"
    for c,w in [(1,8),(2,26),(3,13),(4,13),(5,26),(6,12),(7,13),(8,8),(9,10),(10,8),(11,13),(12,9)]:
        ws_az.column_dimensions[gcl(c)].width=w

    # Inkasso Intrum (statisch — wird aus Intrum-PDF nicht automatisch gelesen)
    ws_ink = wb.create_sheet("⚖️ Inkasso")
    ws_ink.sheet_view.showGridLines = False
    title_bar(ws_ink,1,"Inkasso Intrum – offene Fälle (manuell pflegen)",bg=PURPLE,cols=10)
    ws_ink.merge_cells("A2:J2"); ws_ink["A2"].fill=fill("8B5CF6"); ws_ink.row_dimensions[2].height=4
    hdr(ws_ink,3,["Inkasso-Nr.","Schuldner","Adresse","Score","Adress-Status","Ref.-Nr.","Datum","Kapital (CHF)","Fallstatus","Erfolgsquote"],bg="4C1D95")
    ws_ink.cell(4,1,"→ Bitte Inkasso-Fälle aus Intrum-Bestätigung hier manuell eintragen").font = fnt(sz=9,color="64748B",italic=True)
    ws_ink.merge_cells("A4:J4"); ws_ink.cell(4,1).fill = fill(HDRBG)
    ws_ink.row_dimensions[4].height = 14
    # Score-Legende
    sec(ws_ink,6,"Score-Legende Intrum",cols=10)
    hdr(ws_ink,7,["Score","Erfolgsquote","Bedeutung"],bg="4C1D95")
    score_colors = ["DCFCE7","D1FAE5","FEF3C7","FED7AA","FEE2E2"]
    for idx,(sc_,eq,desc) in enumerate([
        ("A","75–100%","Bestätigte Adresse, Zahlungserfahrung"),
        ("B","60–75%","Bestätigte Adresse, kürzere Adresshistorie"),
        ("C","25–60%","Keine bestätigte Adresse"),
        ("D","10–25%","Offene Forderungen / Betreibungen"),
        ("E","0–10%","Hohe Forderungen / Pfändungen / Verlustscheine"),
    ]):
        i = 8 + idx
        bg = ALT if idx%2==0 else WHITE
        fc = score_colors[idx]
        dc(ws_ink,i,1,sc_,bg); ws_ink.cell(i,1).fill = fill(fc)
        ws_ink.cell(i,1).font = fnt(bold=True,sz=10)
        dc(ws_ink,i,2,eq,bg,h="center"); dc(ws_ink,i,3,desc,bg,h="left")
    for c,w in [(1,12),(2,14),(3,40),(4,8),(5,18),(6,12),(7,12),(8,14),(9,14),(10,14)]:
        ws_ink.column_dimensions[gcl(c)].width = w

    # Berichte an Ärzte
    ws_ba = wb.create_sheet("📋 Berichte Ärzte")
    ws_ba.sheet_view.showGridLines = False
    title_bar(ws_ba,1,"Berichte an Ärzte",bg=TEAL,cols=18)
    therapists=ber_data.get("therapists",[])
    data=ber_data.get("data",{})
    if therapists:
        hdr(ws_ba,2,["KW"]+therapists+["Total"],bg="164E63")
        for i,kw in enumerate(sorted(data.keys()),3):
            bg=ALT if i%2==0 else WHITE
            ws_ba.cell(i,1,kw).fill=fill(bg); ws_ba.cell(i,1).font=fnt(bold=True,sz=10); ws_ba.cell(i,1).border=bb()
            vals=data[kw]
            for j,t in enumerate(therapists,2):
                v=vals.get(t,0)
                ce=ws_ba.cell(i,j,v if v else "")
                ce.fill=fill("BAE6FD" if v else bg); ce.font=fnt(sz=10,bold=bool(v)); ce.border=bb(); ce.alignment=aln("center")
            ws_ba.cell(i,len(therapists)+2,f"=SUM(B{i}:{gcl(len(therapists)+1)}{i})").fill=fill(bg)
            ws_ba.cell(i,len(therapists)+2).font=fnt(bold=True,sz=10); ws_ba.cell(i,len(therapists)+2).border=bb(); ws_ba.cell(i,len(therapists)+2).alignment=aln("center")
        ws_ba.column_dimensions["A"].width=8
        for c in range(2,len(therapists)+3): ws_ba.column_dimensions[gcl(c)].width=9

    # Leeres Sheet entfernen, Reihenfolge korrigieren
    if "_temp" in wb.sheetnames:
        del wb["_temp"]
    desired = ["📊 Übersicht","👥 Therapeuten","👟 Runnerslab","🏃 Hyrox",
               "🏋 Fitness-Abos","💳 Ausstehende Zhlg.","⚖️ Inkasso","📋 Berichte Ärzte"]
    sheet_map = {ws.title: ws for ws in wb.worksheets}
    ordered = [sheet_map[n] for n in desired if n in sheet_map]
    remaining = [ws for ws in wb.worksheets if ws not in ordered]
    wb._sheets = ordered + remaining

    # Speichern
    ts = datetime.today().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_DIR / f"Kineo_Dashboard_{ts}.xlsx"
    # Auch als fixen Namen speichern (einfacher Zugriff)
    fixed_path = OUTPUT_DIR / "Kineo_Dashboard_AKTUELL.xlsx"
    wb.save(out_path)
    wb.save(fixed_path)

    print(f"\n✅ Dashboard gespeichert:")
    print(f"   {out_path}")
    print(f"   {fixed_path}  (immer aktuellste Version)")
    if files_missing:
        non_hyrox = [f for f in files_missing if "hyrox" not in f]
        if non_hyrox:
            print(f"\n⚠️  Fehlende Dateien (nicht kritisch): {', '.join(non_hyrox)}")
    print("\nFertig!")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Fehler: {e}")
        traceback.print_exc()
        input("\nEnter drücken zum Beenden...")
